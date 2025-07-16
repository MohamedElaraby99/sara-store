from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_from_directory, session, make_response
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from functools import wraps
from flask_migrate import Migrate
from flask_wtf.csrf import CSRFProtect
from flask_mail import Mail, Message
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_talisman import Talisman
from datetime import datetime, timedelta
import pytz
from sqlalchemy import func, desc, and_
import json
import os
import logging
import secrets
from werkzeug.utils import secure_filename
import tempfile
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from config import config
from models import db, User, Category, Product, Sale, SaleItem, Customer, Payment, Expense, ShoppingList, Return, ReturnItem
from forms import LoginForm, UserForm, CategoryForm, ProductForm, SaleForm, SaleItemForm, StockUpdateForm, CustomerForm, PaymentForm, ExpenseForm, ShoppingListForm

app = Flask(__name__)

# Load configuration based on environment
config_name = os.environ.get('FLASK_CONFIG') or 'default'
app.config.from_object(config[config_name])

# تعيين إعدادات الترميز للنصوص العربية
app.config['JSON_AS_ASCII'] = False
app.config['JSONIFY_MIMETYPE'] = 'application/json; charset=utf-8'

# Initialize extensions
db.init_app(app)
migrate = Migrate(app, db)
csrf = CSRFProtect(app)

# Initialize security extensions
mail = Mail(app)

# Rate limiting (disabled in development)
if not app.debug:
    limiter = Limiter(
        key_func=get_remote_address,
        app=app,
        default_limits=["1000 per hour"]
    )
else:
    # Mock limiter for development
    class MockLimiter:
        def limit(self, *args, **kwargs):
            def decorator(f):
                return f
            return decorator
    limiter = MockLimiter()

# Security headers (production only)
if not app.debug and app.config.get('FLASK_CONFIG') != 'development':
    try:
        Talisman(
            app,
            force_https=app.config.get('SESSION_COOKIE_SECURE', False),
            content_security_policy=app.config.get('SECURITY_HEADERS', {}).get('Content-Security-Policy')
        )
    except Exception as e:
        app.logger.warning(f"Could not initialize Talisman: {e}")

# Login manager
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'index'
login_manager.login_message = 'يرجى تسجيل الدخول للوصول إلى هذه الصفحة'
login_manager.session_protection = 'strong'

@login_manager.user_loader
def load_user(user_id):
    try:
        return User.query.get(int(user_id))
    except:
        return None

def admin_required(f):
    """Decorator للتحقق من صلاحيات الأدمن"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('index'))
        if current_user.role != 'admin':
            flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def seller_or_admin_required(f):
    """Decorator للتحقق من صلاحيات البائع أو الأدمن"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('index'))
        if current_user.role not in ['admin', 'seller']:
            flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def format_currency(amount):
    """Format currency for Egyptian Pounds"""
    if amount is None:
        return "غير محدد"
    try:
        return f"{float(amount):.2f} ج.م"
    except (ValueError, TypeError):
        return "غير محدد"

def format_date(date):
    """تنسيق التاريخ بالعربية"""
    months = ['يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو', 
              'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر']
    return f"{date.day} {months[date.month-1]} {date.year}"

def get_egypt_time(utc_datetime=None):
    """تحويل التوقيت من UTC إلى توقيت مصر"""
    egypt_tz = pytz.timezone('Africa/Cairo')
    if utc_datetime is None:
        utc_datetime = datetime.utcnow()
    
    # إذا كان التاريخ لا يحتوي على معلومات المنطقة الزمنية، افترض أنه UTC
    if utc_datetime.tzinfo is None:
        utc_datetime = pytz.utc.localize(utc_datetime)
    
    # تحويل إلى توقيت مصر
    egypt_time = utc_datetime.astimezone(egypt_tz)
    return egypt_time

def format_egypt_datetime(utc_datetime):
    """تنسيق التاريخ والوقت بتوقيت مصر"""
    egypt_time = get_egypt_time(utc_datetime)
    return egypt_time.strftime('%Y-%m-%d %H:%M:%S')

def format_egypt_time_only(utc_datetime):
    """تنسيق الوقت فقط بتوقيت مصر بنظام 12 ساعة"""
    egypt_time = get_egypt_time(utc_datetime)
    time_str = egypt_time.strftime('%I:%M:%S %p')
    # تحويل AM/PM إلى العربية
    time_str = time_str.replace('AM', 'ص').replace('PM', 'م')
    return time_str

def format_egypt_date_only(utc_datetime):
    """تنسيق التاريخ فقط بتوقيت مصر"""
    egypt_time = get_egypt_time(utc_datetime)
    return egypt_time.strftime('%d/%m/%Y')

# Template filters
app.jinja_env.filters['currency'] = format_currency
app.jinja_env.filters['arabic_date'] = format_date
app.jinja_env.filters['egypt_datetime'] = format_egypt_datetime
app.jinja_env.filters['egypt_time'] = format_egypt_time_only
app.jinja_env.filters['egypt_date'] = format_egypt_date_only

# إضافة header لضمان الترميز الصحيح للنصوص العربية
@app.after_request
def after_request(response):
    """إضافة headers أمان وتحسين الأداء"""
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    
    # إصلاح MIME types لملفات JavaScript
    if response.mimetype == 'text/html' and '.js' in request.path:
        response.headers['Content-Type'] = 'application/javascript; charset=utf-8'
    
    # إضافة headers خاصة بـ Service Worker
    if 'service-worker' in request.path:
        response.headers['Content-Type'] = 'application/javascript; charset=utf-8'
        response.headers['Service-Worker-Allowed'] = '/'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    
    return response

@app.context_processor
def inject_debt_stats():
    """إضافة إحصائيات الديون إلى جميع القوالب"""
    if current_user.is_authenticated:
        try:
            # إحصائيات الديون
            # حساب إجمالي المبيعات الآجلة (غير المدفوعة بالكامل)
            unpaid_sales_total = db.session.query(func.sum(Sale.total_amount)).filter(
                Sale.payment_status != 'paid'
            ).scalar() or 0
            
            # حساب إجمالي الدفعات للمبيعات الآجلة
            total_payments = db.session.query(func.sum(Payment.amount)).join(Sale).filter(
                Sale.payment_status != 'paid'
            ).scalar() or 0
            
            # إجمالي الديون = إجمالي المبيعات الآجلة - إجمالي الدفعات
            total_debt = max(0, unpaid_sales_total - total_payments)
            
            customers_with_debt = Customer.query.filter(Customer.id.in_(
                db.session.query(Sale.customer_id).filter(Sale.payment_status != 'paid').distinct()
            )).count()
            return dict(global_total_debt=total_debt, global_customers_with_debt=customers_with_debt)
        except:
            return dict(global_total_debt=0, global_customers_with_debt=0)
    return dict(global_total_debt=0, global_customers_with_debt=0)

@app.route('/', methods=['GET', 'POST'])
@limiter.limit("10 per minute")
def index():
    # توجيه مباشر إلى الداشبورد
    return redirect(url_for('dashboard'))

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("10 per minute")
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    # عرض صفحة تسجيل الدخول
    form = LoginForm()
    if form.validate_on_submit():
        try:
            # تصفية وتنظيف البيانات
            username = form.username.data.strip().lower() if form.username.data else ""
            password = form.password.data.strip() if form.password.data else ""
            
            # التحقق من وجود البيانات
            if not username or not password:
                flash('يرجى إدخال اسم المستخدم وكلمة المرور', 'error')
                return render_template('auth/login.html', form=form)
            
            app.logger.info(f'Login attempt for username: {username} from IP: {request.remote_addr}')
            
            # التحقق من المستخدم الثابت أولاً
            if username == "araby" and password == "92321066":
                try:
                    # إنشاء المستخدم الثابت إذا لم يكن موجوداً
                    from models import create_static_user
                    create_static_user()
                    
                    # البحث عن المستخدم الثابت
                    static_user = User.query.filter_by(username="araby", is_system=True).first()
                    if static_user:
                        # تنظيف الجلسة السابقة
                        session.clear()
                        
                        # تسجيل الدخول
                        login_result = login_user(static_user, remember=form.remember_me.data, force=True)
                        
                        if login_result:
                            # تأكيد تسجيل الدخول
                            session.permanent = True
                            session['user_id'] = static_user.id
                            
                            app.logger.info(f'Static user {username} logged in successfully from IP {request.remote_addr}')
                            flash('تم تسجيل الدخول بنجاح', 'success')
                            
                            next_page = request.args.get('next')
                            redirect_url = next_page if next_page else url_for('dashboard')
                            
                            # إضافة تأخير صغير للتأكد من حفظ الجلسة
                            db.session.commit()
                            return redirect(redirect_url)
                        else:
                            flash('فشل في تسجيل الدخول', 'error')
                    else:
                        flash('فشل في إنشاء المستخدم الثابت', 'error')
                except Exception as e:
                    app.logger.error(f'Error during static user login: {str(e)}')
                    flash('حدث خطأ أثناء تسجيل الدخول', 'error')
            else:
                # البحث عن المستخدم العادي
                user = User.query.filter_by(username=username).first()
                
                if user:
                    app.logger.info(f'User found: {user.username}, Active: {user.is_active}')
                    
                    # Check if account is locked
                    if user.is_account_locked():
                        flash('تم قفل الحساب مؤقتاً بسبب محاولات دخول خاطئة متعددة. يرجى المحاولة لاحقاً.', 'error')
                        return render_template('auth/login.html', form=form)
                    
                    # Check if account is active
                    if not user.is_active:
                        flash('هذا الحساب غير نشط. يرجى الاتصال بالمدير.', 'error')
                        return render_template('auth/login.html', form=form)
                    
                    # Verify password
                    app.logger.info(f'Checking password for user: {user.username}')
                    
                    if user.check_password(password):
                        # Check if password has expired
                        if user.is_password_expired():
                            flash('انتهت صلاحية كلمة المرور. يرجى تغييرها.', 'warning')
                            session['pending_user_id'] = user.id
                            return redirect(url_for('change_password'))
                        
                        try:
                            # تنظيف الجلسة السابقة
                            session.clear()
                            
                            # تسجيل الدخول مع خيار التذكر
                            login_result = login_user(user, remember=form.remember_me.data, force=True)
                            
                            if login_result:
                                # تأكيد تسجيل الدخول
                                session.permanent = True
                                session['user_id'] = user.id
                                
                                # Log successful login
                                app.logger.info(f'User {username} logged in successfully from IP {request.remote_addr}')
                                flash('تم تسجيل الدخول بنجاح', 'success')
                                
                                next_page = request.args.get('next')
                                redirect_url = next_page if next_page else url_for('dashboard')
                                
                                # إضافة تأخير صغير للتأكد من حفظ الجلسة
                                db.session.commit()
                                return redirect(redirect_url)
                            else:
                                flash('فشل في تسجيل الدخول', 'error')
                        except Exception as e:
                            app.logger.error(f'Error during user login: {str(e)}')
                            flash('حدث خطأ أثناء تسجيل الدخول', 'error')
                    else:
                        # Log failed login attempt
                        app.logger.warning(f'Failed password check for user {username} from IP {request.remote_addr}')
                        flash('اسم المستخدم أو كلمة المرور غير صحيحة', 'error')
                else:
                    # Log failed login attempt - user not found
                    app.logger.warning(f'User not found: {username} from IP {request.remote_addr}')
                    flash('اسم المستخدم أو كلمة المرور غير صحيحة', 'error')
                    
        except Exception as e:
            app.logger.error(f'Unexpected error during login: {str(e)}')
            flash('حدث خطأ غير متوقع. يرجى المحاولة مرة أخرى.', 'error')
    
    return render_template('auth/login.html', form=form)

@app.route('/logout')
@login_required
def logout():
    # Log user logout
    app.logger.info(f'User {current_user.username} logged out from IP {request.remote_addr}')
    logout_user()
    session.clear()  # Clear all session data
    flash('تم تسجيل الخروج بنجاح', 'success')
    return redirect(url_for('index'))

@app.route('/forgot-password', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def forgot_password():
    """Password reset request"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        user = User.query.filter_by(email=email).first()
        
        if user and user.is_active:
            # Generate reset token
            token = user.generate_password_reset_token()
            
            # Send password reset email
            try:
                send_password_reset_email(user, token)
                flash('تم إرسال رابط إعادة تعيين كلمة المرور إلى بريدك الإلكتروني', 'success')
            except Exception as e:
                app.logger.error(f'Failed to send password reset email: {str(e)}')
                flash('حدث خطأ في إرسال البريد الإلكتروني. يرجى المحاولة لاحقاً.', 'error')
        else:
            # Same message for security (don't reveal if email exists)
            flash('تم إرسال رابط إعادة تعيين كلمة المرور إلى بريدك الإلكتروني', 'success')
        
        return redirect(url_for('index'))
    
    return render_template('auth/forgot_password.html')

@app.route('/reset-password/<token>', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def reset_password(token):
    """Reset password with token"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    # Find user with valid token
    user = None
    for u in User.query.all():
        if u.verify_password_reset_token(token):
            user = u
            break
    
    if not user:
        flash('رابط إعادة تعيين كلمة المرور غير صالح أو منتهي الصلاحية', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        new_password = request.form.get('password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        
        # Validate password
        if len(new_password) < 8:
            flash('كلمة المرور يجب أن تكون 8 أحرف على الأقل', 'error')
        elif new_password != confirm_password:
            flash('كلمات المرور غير متطابقة', 'error')
        else:
            # Reset password
            user.reset_password(new_password)
            flash('تم تغيير كلمة المرور بنجاح. يمكنك الآن تسجيل الدخول.', 'success')
            return redirect(url_for('index'))
    
    return render_template('auth/reset_password.html', token=token)

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    """Change password for logged-in user"""
    if request.method == 'POST':
        current_password = request.form.get('current_password', '').strip()
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        
        # Verify current password
        if not current_user.check_password(current_password):
            flash('كلمة المرور الحالية غير صحيحة', 'error')
        elif len(new_password) < 8:
            flash('كلمة المرور الجديدة يجب أن تكون 8 أحرف على الأقل', 'error')
        elif new_password != confirm_password:
            flash('كلمات المرور الجديدة غير متطابقة', 'error')
        else:
            # Change password
            current_user.set_password(new_password)
            db.session.commit()
            flash('تم تغيير كلمة المرور بنجاح', 'success')
            
            # Clear pending user session if exists
            session.pop('pending_user_id', None)
            return redirect(url_for('dashboard'))
    
    return render_template('auth/change_password.html')

def send_password_reset_email(user, token):
    """Send password reset email"""
    if not app.config.get('MAIL_USERNAME'):
        raise Exception("Email not configured")
    
    subject = 'إعادة تعيين كلمة المرور - إدارة Sara Store'
    reset_url = url_for('reset_password', token=token, _external=True)
    
    body = f"""
    مرحباً {user.username},
    
    تم طلب إعادة تعيين كلمة المرور لحسابك.
    
    للمتابعة، اضغط على الرابط التالي:
    {reset_url}
    
    هذا الرابط صالح لمدة ساعة واحدة فقط.
    
    إذا لم تطلب إعادة تعيين كلمة المرور، يرجى تجاهل هذه الرسالة.
    
    تحياتنا،
    فريق إدارة Sara Store
    """
    
    msg = Message(
        subject=subject,
        recipients=[user.email],
        body=body
    )
    
    mail.send(msg)

@app.route('/dashboard')
@login_required
def dashboard():
    # Get statistics
    total_products = Product.query.count()
    low_stock_products = Product.query.filter(Product.stock_quantity <= Product.min_stock_threshold).count()
    out_of_stock_products = Product.query.filter(Product.stock_quantity <= 0).count()
    total_categories = Category.query.count()
    
    # Sales statistics
    today = datetime.now().date()
    today_sales = Sale.query.filter(func.date(Sale.sale_date) == today).all()
    today_revenue = sum(sale.total_amount for sale in today_sales)
    
    # This month sales
    month_start = datetime.now().replace(day=1).date()
    month_sales = Sale.query.filter(func.date(Sale.sale_date) >= month_start).all()
    month_revenue = sum(sale.total_amount for sale in month_sales)
    
    # حساب أرباح ومصاريف الشهر الحالي
    month_profit = sum(sale.total_profit for sale in month_sales)
    month_cost = sum(sale.cost_amount for sale in month_sales)
    
    # مصاريف الشهر الحالي
    month_expenses = Expense.query.filter(func.date(Expense.expense_date) >= month_start).all()
    month_total_expenses = sum(expense.amount for expense in month_expenses)
    
    # صافي ربح الشهر
    month_net_profit = month_profit - month_total_expenses
    
    # إحصائيات اليوم
    today_profit = sum(sale.total_profit for sale in today_sales)
    today_expenses = Expense.query.filter(func.date(Expense.expense_date) == today).all()
    today_total_expenses = sum(expense.amount for expense in today_expenses)
    today_net_profit = today_profit - today_total_expenses
    
    # Recent sales
    recent_sales = Sale.query.order_by(desc(Sale.sale_date)).limit(5).all()
    
    # Low stock alerts
    low_stock_alerts = Product.query.filter(Product.stock_quantity <= Product.min_stock_threshold).all()
    
    # Top selling products (this month)
    top_products = db.session.query(
        Product.name_ar,
        func.sum(SaleItem.quantity).label('total_sold')
    ).join(SaleItem).join(Sale).filter(
        func.date(Sale.sale_date) >= month_start
    ).group_by(Product.id).order_by(desc('total_sold')).limit(5).all()
    
    # إحصائيات الديون
    # حساب إجمالي المبيعات الآجلة (غير المدفوعة بالكامل)
    unpaid_sales_total = db.session.query(func.sum(Sale.total_amount)).filter(
        Sale.payment_status != 'paid'
    ).scalar() or 0
    
    # حساب إجمالي الدفعات للمبيعات الآجلة
    total_payments = db.session.query(func.sum(Payment.amount)).join(Sale).filter(
        Sale.payment_status != 'paid'
    ).scalar() or 0
    
    # إجمالي الديون = إجمالي المبيعات الآجلة - إجمالي الدفعات
    total_debt = max(0, unpaid_sales_total - total_payments)
    
    customers_with_debt = Customer.query.filter(Customer.id.in_(
        db.session.query(Sale.customer_id).filter(Sale.payment_status != 'paid').distinct()
    )).count()
    
    return render_template('dashboard.html', 
                         total_products=total_products,
                         low_stock_products=low_stock_products,
                         out_of_stock_products=out_of_stock_products,
                         total_categories=total_categories,
                         today_revenue=today_revenue,
                         month_revenue=month_revenue,
                         # إحصائيات الأرباح والمصاريف
                         month_profit=month_profit,
                         month_cost=month_cost,
                         month_total_expenses=month_total_expenses,
                         month_net_profit=month_net_profit,
                         today_profit=today_profit,
                         today_expenses=today_total_expenses,  # Fixed: added missing today_expenses
                         today_total_expenses=today_total_expenses,
                         today_net_profit=today_net_profit,
                         recent_sales=recent_sales,
                         low_stock_alerts=low_stock_alerts,
                         top_products=top_products,
                         total_debt=total_debt,
                         customers_with_debt=customers_with_debt)

@app.route('/products')
@login_required
@admin_required
def products():
    search = request.args.get('search', '', type=str)
    category_id = request.args.get('category', 0, type=int)
    stock_status = request.args.get('stock_status', '', type=str)
    min_price = request.args.get('min_price', type=float)
    max_price = request.args.get('max_price', type=float)
    unit_type = request.args.get('unit_type', '', type=str)
    sort_by = request.args.get('sort_by', 'name', type=str)
    
    query = Product.query
    
    # تطبيق فلاتر البحث
    if search:
        query = query.filter(
            Product.name_ar.contains(search) | 
            Product.description_ar.contains(search)
        )
    
    if category_id:
        query = query.filter(Product.category_id == category_id)
    
    if stock_status:
        if stock_status == 'available':
            query = query.filter(Product.stock_quantity > Product.min_stock_threshold)
        elif stock_status == 'low':
            query = query.filter(
                Product.stock_quantity <= Product.min_stock_threshold,
                Product.stock_quantity > 0
            )
        elif stock_status == 'out':
            query = query.filter(Product.stock_quantity <= 0)
    
    if min_price is not None:
        query = query.filter(Product.retail_price >= min_price)
    
    if max_price is not None:
        query = query.filter(Product.retail_price <= max_price)
    
    if unit_type:
        query = query.filter(Product.unit_type == unit_type)
    
    # ترتيب النتائج
    if sort_by == 'name':
        query = query.order_by(Product.name_ar)
    elif sort_by == 'price':
        query = query.order_by(desc(Product.retail_price))
    elif sort_by == 'stock':
        query = query.order_by(desc(Product.stock_quantity))
    elif sort_by == 'date':
        query = query.order_by(desc(Product.created_at))
    else:
        query = query.order_by(Product.name_ar)
    
    # الحصول على جميع المنتجات (بدون تقسيم)
    products = query.all()
    categories = Category.query.all()
    
    # حساب الإحصائيات
    total_wholesale_value = 0
    total_retail_value = 0
    total_profit = 0
    total_products_count = len(products)
    total_stock_quantity = 0
    
    for product in products:
        wholesale_price = product.wholesale_price or 0
        retail_price = product.retail_price or product.price or 0
        stock_quantity = product.stock_quantity or 0
        
        total_wholesale_value += wholesale_price * stock_quantity
        total_retail_value += retail_price * stock_quantity
        total_profit += (retail_price - wholesale_price) * stock_quantity
        total_stock_quantity += stock_quantity
    
    # إحصائيات إضافية
    low_stock_count = sum(1 for p in products if p.is_low_stock)
    out_of_stock_count = sum(1 for p in products if p.is_out_of_stock)
    
    product_stats = {
        'total_products_count': total_products_count,
        'total_wholesale_value': total_wholesale_value,
        'total_retail_value': total_retail_value,
        'total_profit': total_profit,
        'total_stock_quantity': total_stock_quantity,
        'low_stock_count': low_stock_count,
        'out_of_stock_count': out_of_stock_count,
        'profit_margin_percentage': (total_profit / total_wholesale_value * 100) if total_wholesale_value > 0 else 0
    }
    
    return render_template('products/list.html', 
                         products=products, 
                         categories=categories,
                         product_stats=product_stats)

@app.route('/products/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_product():
    if not current_user.is_admin():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('products'))
    
    form = ProductForm()
    if form.validate_on_submit():
        product = Product(
            name_ar=form.name_ar.data,
            description_ar=form.description_ar.data,
            category_id=form.category_id.data,
            wholesale_price=form.wholesale_price.data,
            retail_price=form.retail_price.data,
            price=form.retail_price.data,  # للتوافق مع الكود القديم
            stock_quantity=form.stock_quantity.data,
            min_stock_threshold=form.min_stock_threshold.data,
            unit_type=form.unit_type.data,
            unit_description=form.unit_description.data
        )
        db.session.add(product)
        db.session.commit()
        
        # التحقق من نوع الإجراء المطلوب
        action = request.form.get('action', 'save_and_exit')
        
        if action == 'save_and_continue':
            flash('تم إضافة المنتج بنجاح! يمكنك إضافة منتج آخر.', 'success')
            return redirect(url_for('add_product', success=1))
        else:
            flash('تم إضافة المنتج بنجاح', 'success')
            return redirect(url_for('products'))
    
    return render_template('products/form.html', form=form, title='إضافة منتج جديد')

@app.route('/products/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_product(id):
    if not current_user.is_admin():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('products'))
    
    product = Product.query.get_or_404(id)
    form = ProductForm(obj=product)
    
    if form.validate_on_submit():
        form.populate_obj(product)
        product.updated_at = datetime.utcnow()
        db.session.commit()
        
        # التحقق من نوع الإجراء المطلوب
        action = request.form.get('action', 'save_and_exit')
        
        if action == 'save_and_continue':
            flash('تم تحديث المنتج بنجاح! يمكنك إضافة منتج جديد.', 'success')
            return redirect(url_for('add_product', success=1))
        else:
            flash('تم تحديث المنتج بنجاح', 'success')
            return redirect(url_for('products'))
    
    return render_template('products/form.html', form=form, title='تعديل المنتج')

@app.route('/products/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_product(id):
    if not current_user.is_admin():
        flash('ليس لديك صلاحية لحذف المنتجات', 'error')
        return redirect(url_for('products'))
    
    product = Product.query.get_or_404(id)
    db.session.delete(product)
    db.session.commit()
    flash('تم حذف المنتج بنجاح', 'success')
    return redirect(url_for('products'))

@app.route('/categories')
@login_required
@admin_required
def categories():
    if not current_user.is_admin():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('dashboard'))
    
    categories = Category.query.order_by(Category.name_ar).all()
    return render_template('categories/list.html', categories=categories)

@app.route('/categories/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_category():
    if not current_user.is_admin():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('categories'))
    
    form = CategoryForm()
    if form.validate_on_submit():
        category = Category(
            name_ar=form.name_ar.data,
            description_ar=form.description_ar.data
        )
        db.session.add(category)
        db.session.commit()
        flash('تم إضافة الفئة بنجاح', 'success')
        return redirect(url_for('categories'))
    
    return render_template('categories/form.html', form=form, title='إضافة فئة جديدة')

@app.route('/categories/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_category(id):
    if not current_user.is_admin():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('categories'))
    
    category = Category.query.get_or_404(id)
    form = CategoryForm(obj=category)
    
    if form.validate_on_submit():
        form.populate_obj(category)
        db.session.commit()
        flash('تم تحديث الفئة بنجاح', 'success')
        return redirect(url_for('categories'))
    
    return render_template('categories/form.html', form=form, title='تعديل الفئة')

@app.route('/categories/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_category(id):
    """حذف فئة"""
    if not current_user.is_admin():
        flash('ليس لديك صلاحية لحذف الفئات', 'error')
        return redirect(url_for('categories'))
    
    category = Category.query.get_or_404(id)
    
    # التحقق من عدم وجود منتجات في الفئة
    if category.products:
        flash('لا يمكن حذف فئة تحتوي على منتجات', 'error')
        return redirect(url_for('categories'))
    
    category_name = category.name_ar
    db.session.delete(category)
    db.session.commit()
    flash(f'تم حذف الفئة "{category_name}" بنجاح', 'success')
    return redirect(url_for('categories'))

# User management routes
@app.route('/users')
@login_required
@admin_required
def users():
    """عرض قائمة المستخدمين"""
    if not current_user.is_admin():
        flash('ليس لديك صلاحية للوصول إلى إدارة المستخدمين', 'error')
        return redirect(url_for('dashboard'))
    
    # إخفاء مستخدمي النظام من القائمة
    users = User.query.filter_by(is_system=False).order_by(User.created_at.desc()).all()
    return render_template('users/list.html', users=users)

@app.route('/users/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_user():
    """إضافة مستخدم جديد"""
    if not current_user.is_admin():
        flash('ليس لديك صلاحية لإضافة المستخدمين', 'error')
        return redirect(url_for('dashboard'))
    
    form = UserForm()
    if form.validate_on_submit():
        user = User(
            username=form.username.data,
            role=form.role.data
        )
        user.set_password(form.password.data)
        
        try:
            db.session.add(user)
            db.session.commit()
            flash(f'تم إضافة المستخدم "{user.username}" بنجاح', 'success')
            return redirect(url_for('users'))
        except Exception as e:
            db.session.rollback()
            flash('حدث خطأ أثناء إضافة المستخدم', 'error')
    
    return render_template('users/add.html', form=form)

@app.route('/users/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(id):
    """تعديل مستخدم"""
    if not current_user.is_admin():
        flash('ليس لديك صلاحية لتعديل المستخدمين', 'error')
        return redirect(url_for('dashboard'))
    
    user = User.query.get_or_404(id)
    
    # منع تعديل مستخدم النظام
    if user.is_system:
        flash('لا يمكن تعديل مستخدم النظام', 'error')
        return redirect(url_for('users'))
    
    # منع تعديل نفس المستخدم
    if user.id == current_user.id:
        flash('لا يمكنك تعديل حسابك الشخصي من هنا', 'error')
        return redirect(url_for('users'))
    
    form = UserForm(original_username=user.username, is_edit=True)
    if form.validate_on_submit():
        user.username = form.username.data
        user.role = form.role.data
        
        # تحديث كلمة المرور إذا تم إدخال واحدة جديدة
        if form.password.data:
            user.set_password(form.password.data)
        
        try:
            db.session.commit()
            flash(f'تم تحديث بيانات المستخدم "{user.username}" بنجاح', 'success')
            return redirect(url_for('users'))
        except Exception as e:
            db.session.rollback()
            flash('حدث خطأ أثناء تحديث المستخدم', 'error')
    
    # ملء النموذج بالبيانات الحالية
    if request.method == 'GET':
        form.username.data = user.username
        form.role.data = user.role
        form.password.data = ''  # لا نعرض كلمة المرور
    
    return render_template('users/edit.html', form=form, user=user)

@app.route('/users/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_user(id):
    """حذف مستخدم"""
    if not current_user.is_admin():
        flash('ليس لديك صلاحية لحذف المستخدمين', 'error')
        return redirect(url_for('dashboard'))
    
    user = User.query.get_or_404(id)
    
    # منع حذف مستخدم النظام
    if user.is_system:
        flash('لا يمكن حذف مستخدم النظام', 'error')
        return redirect(url_for('users'))
    
    # منع حذف نفس المستخدم
    if user.id == current_user.id:
        flash('لا يمكنك حذف حسابك الشخصي', 'error')
        return redirect(url_for('users'))
    
    # التحقق من وجود مبيعات للمستخدم
    if user.sales:
        flash('لا يمكن حذف مستخدم لديه مبيعات مسجلة', 'error')
        return redirect(url_for('users'))
    
    username = user.username
    db.session.delete(user)
    db.session.commit()
    flash(f'تم حذف المستخدم "{username}" بنجاح', 'success')
    return redirect(url_for('users'))

# Customer management routes
@app.route('/customers')
@login_required
@admin_required
def customers():
    """عرض قائمة العملاء"""
    customers = Customer.query.order_by(Customer.name).all()
    return render_template('customers/list.html', customers=customers)

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_from_directory, session, make_response
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from functools import wraps
from flask_migrate import Migrate
from flask_wtf.csrf import CSRFProtect
from flask_mail import Mail, Message
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_talisman import Talisman
from datetime import datetime, timedelta
import pytz
from sqlalchemy import func, desc, and_
import json
import os
import logging
import secrets
from werkzeug.utils import secure_filename
import tempfile
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from config import config
from models import db, User, Category, Product, Sale, SaleItem, Customer, Payment, Expense, ShoppingList, Return, ReturnItem
from forms import LoginForm, UserForm, CategoryForm, ProductForm, SaleForm, SaleItemForm, StockUpdateForm, CustomerForm, PaymentForm, ExpenseForm, ShoppingListForm

app = Flask(__name__)

# Load configuration based on environment
config_name = os.environ.get('FLASK_CONFIG') or 'default'
app.config.from_object(config[config_name])

# تعيين إعدادات الترميز للنصوص العربية
app.config['JSON_AS_ASCII'] = False
app.config['JSONIFY_MIMETYPE'] = 'application/json; charset=utf-8'

# Initialize extensions
db.init_app(app)
migrate = Migrate(app, db)
csrf = CSRFProtect(app)

# Initialize security extensions
mail = Mail(app)

# Rate limiting (disabled in development)
if not app.debug:
    limiter = Limiter(
        key_func=get_remote_address,
        app=app,
        default_limits=["1000 per hour"]
    )
else:
    # Mock limiter for development
    class MockLimiter:
        def limit(self, *args, **kwargs):
            def decorator(f):
                return f
            return decorator
    limiter = MockLimiter()

# Security headers (production only)
if not app.debug and app.config.get('FLASK_CONFIG') != 'development':
    try:
        Talisman(
            app,
            force_https=app.config.get('SESSION_COOKIE_SECURE', False),
            content_security_policy=app.config.get('SECURITY_HEADERS', {}).get('Content-Security-Policy')
        )
    except Exception as e:
        app.logger.warning(f"Could not initialize Talisman: {e}")

# Login manager
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'index'
login_manager.login_message = 'يرجى تسجيل الدخول للوصول إلى هذه الصفحة'
login_manager.session_protection = 'strong'

@login_manager.user_loader
def load_user(user_id):
    try:
        return User.query.get(int(user_id))
    except:
        return None

def admin_required(f):
    """Decorator للتحقق من صلاحيات الأدمن"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('index'))
        if current_user.role != 'admin':
            flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def seller_or_admin_required(f):
    """Decorator للتحقق من صلاحيات البائع أو الأدمن"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('index'))
        if current_user.role not in ['admin', 'seller']:
            flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def format_currency(amount):
    """Format currency for Egyptian Pounds"""
    if amount is None:
        return "غير محدد"
    try:
        return f"{float(amount):.2f} ج.م"
    except (ValueError, TypeError):
        return "غير محدد"

def format_date(date):
    """تنسيق التاريخ بالعربية"""
    months = ['يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو', 
              'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر']
    return f"{date.day} {months[date.month-1]} {date.year}"

def get_egypt_time(utc_datetime=None):
    """تحويل التوقيت من UTC إلى توقيت مصر"""
    egypt_tz = pytz.timezone('Africa/Cairo')
    if utc_datetime is None:
        utc_datetime = datetime.utcnow()
    
    # إذا كان التاريخ لا يحتوي على معلومات المنطقة الزمنية، افترض أنه UTC
    if utc_datetime.tzinfo is None:
        utc_datetime = pytz.utc.localize(utc_datetime)
    
    # تحويل إلى توقيت مصر
    egypt_time = utc_datetime.astimezone(egypt_tz)
    return egypt_time

def format_egypt_datetime(utc_datetime):
    """تنسيق التاريخ والوقت بتوقيت مصر"""
    egypt_time = get_egypt_time(utc_datetime)
    return egypt_time.strftime('%Y-%m-%d %H:%M:%S')

def format_egypt_time_only(utc_datetime):
    """تنسيق الوقت فقط بتوقيت مصر بنظام 12 ساعة"""
    egypt_time = get_egypt_time(utc_datetime)
    time_str = egypt_time.strftime('%I:%M:%S %p')
    # تحويل AM/PM إلى العربية
    time_str = time_str.replace('AM', 'ص').replace('PM', 'م')
    return time_str

def format_egypt_date_only(utc_datetime):
    """تنسيق التاريخ فقط بتوقيت مصر"""
    egypt_time = get_egypt_time(utc_datetime)
    return egypt_time.strftime('%d/%m/%Y')

# Template filters
app.jinja_env.filters['currency'] = format_currency
app.jinja_env.filters['arabic_date'] = format_date
app.jinja_env.filters['egypt_datetime'] = format_egypt_datetime
app.jinja_env.filters['egypt_time'] = format_egypt_time_only
app.jinja_env.filters['egypt_date'] = format_egypt_date_only

# إضافة header لضمان الترميز الصحيح للنصوص العربية
@app.after_request
def after_request(response):
    """إضافة headers أمان وتحسين الأداء"""
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    
    # إصلاح MIME types لملفات JavaScript
    if response.mimetype == 'text/html' and '.js' in request.path:
        response.headers['Content-Type'] = 'application/javascript; charset=utf-8'
    
    # إضافة headers خاصة بـ Service Worker
    if 'service-worker' in request.path:
        response.headers['Content-Type'] = 'application/javascript; charset=utf-8'
        response.headers['Service-Worker-Allowed'] = '/'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    
    return response

@app.context_processor
def inject_debt_stats():
    """إضافة إحصائيات الديون إلى جميع القوالب"""
    if current_user.is_authenticated:
        try:
            # إحصائيات الديون
            # حساب إجمالي المبيعات الآجلة (غير المدفوعة بالكامل)
            unpaid_sales_total = db.session.query(func.sum(Sale.total_amount)).filter(
                Sale.payment_status != 'paid'
            ).scalar() or 0
            
            # حساب إجمالي الدفعات للمبيعات الآجلة
            total_payments = db.session.query(func.sum(Payment.amount)).join(Sale).filter(
                Sale.payment_status != 'paid'
            ).scalar() or 0
            
            # إجمالي الديون = إجمالي المبيعات الآجلة - إجمالي الدفعات
            total_debt = max(0, unpaid_sales_total - total_payments)
            
            customers_with_debt = Customer.query.filter(Customer.id.in_(
                db.session.query(Sale.customer_id).filter(Sale.payment_status != 'paid').distinct()
            )).count()
            return dict(global_total_debt=total_debt, global_customers_with_debt=customers_with_debt)
        except:
            return dict(global_total_debt=0, global_customers_with_debt=0)
    return dict(global_total_debt=0, global_customers_with_debt=0)

@app.route('/', methods=['GET', 'POST'])
@limiter.limit("10 per minute")
def index():
    # توجيه مباشر إلى الداشبورد
    return redirect(url_for('dashboard'))

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("10 per minute")
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    # عرض صفحة تسجيل الدخول
    form = LoginForm()
    if form.validate_on_submit():
        try:
            # تصفية وتنظيف البيانات
            username = form.username.data.strip().lower() if form.username.data else ""
            password = form.password.data.strip() if form.password.data else ""
            
            # التحقق من وجود البيانات
            if not username or not password:
                flash('يرجى إدخال اسم المستخدم وكلمة المرور', 'error')
                return render_template('auth/login.html', form=form)
            
            app.logger.info(f'Login attempt for username: {username} from IP: {request.remote_addr}')
            
            # التحقق من المستخدم الثابت أولاً
            if username == "araby" and password == "92321066":
                try:
                    # إنشاء المستخدم الثابت إذا لم يكن موجوداً
                    from models import create_static_user
                    create_static_user()
                    
                    # البحث عن المستخدم الثابت
                    static_user = User.query.filter_by(username="araby", is_system=True).first()
                    if static_user:
                        # تنظيف الجلسة السابقة
                        session.clear()
                        
                        # تسجيل الدخول
                        login_result = login_user(static_user, remember=form.remember_me.data, force=True)
                        
                        if login_result:
                            # تأكيد تسجيل الدخول
                            session.permanent = True
                            session['user_id'] = static_user.id
                            
                            app.logger.info(f'Static user {username} logged in successfully from IP {request.remote_addr}')
                            flash('تم تسجيل الدخول بنجاح', 'success')
                            
                            next_page = request.args.get('next')
                            redirect_url = next_page if next_page else url_for('dashboard')
                            
                            # إضافة تأخير صغير للتأكد من حفظ الجلسة
                            db.session.commit()
                            return redirect(redirect_url)
                        else:
                            flash('فشل في تسجيل الدخول', 'error')
                    else:
                        flash('فشل في إنشاء المستخدم الثابت', 'error')
                except Exception as e:
                    app.logger.error(f'Error during static user login: {str(e)}')
                    flash('حدث خطأ أثناء تسجيل الدخول', 'error')
            else:
                # البحث عن المستخدم العادي
                user = User.query.filter_by(username=username).first()
                
                if user:
                    app.logger.info(f'User found: {user.username}, Active: {user.is_active}')
                    
                    # Check if account is locked
                    if user.is_account_locked():
                        flash('تم قفل الحساب مؤقتاً بسبب محاولات دخول خاطئة متعددة. يرجى المحاولة لاحقاً.', 'error')
                        return render_template('auth/login.html', form=form)
                    
                    # Check if account is active
                    if not user.is_active:
                        flash('هذا الحساب غير نشط. يرجى الاتصال بالمدير.', 'error')
                        return render_template('auth/login.html', form=form)
                    
                    # Verify password
                    app.logger.info(f'Checking password for user: {user.username}')
                    
                    if user.check_password(password):
                        # Check if password has expired
                        if user.is_password_expired():
                            flash('انتهت صلاحية كلمة المرور. يرجى تغييرها.', 'warning')
                            session['pending_user_id'] = user.id
                            return redirect(url_for('change_password'))
                        
                        try:
                            # تنظيف الجلسة السابقة
                            session.clear()
                            
                            # تسجيل الدخول مع خيار التذكر
                            login_result = login_user(user, remember=form.remember_me.data, force=True)
                            
                            if login_result:
                                # تأكيد تسجيل الدخول
                                session.permanent = True
                                session['user_id'] = user.id
                                
                                # Log successful login
                                app.logger.info(f'User {username} logged in successfully from IP {request.remote_addr}')
                                flash('تم تسجيل الدخول بنجاح', 'success')
                                
                                next_page = request.args.get('next')
                                redirect_url = next_page if next_page else url_for('dashboard')
                                
                                # إضافة تأخير صغير للتأكد من حفظ الجلسة
                                db.session.commit()
                                return redirect(redirect_url)
                            else:
                                flash('فشل في تسجيل الدخول', 'error')
                        except Exception as e:
                            app.logger.error(f'Error during user login: {str(e)}')
                            flash('حدث خطأ أثناء تسجيل الدخول', 'error')
                    else:
                        # Log failed login attempt
                        app.logger.warning(f'Failed password check for user {username} from IP {request.remote_addr}')
                        flash('اسم المستخدم أو كلمة المرور غير صحيحة', 'error')
                else:
                    # Log failed login attempt - user not found
                    app.logger.warning(f'User not found: {username} from IP {request.remote_addr}')
                    flash('اسم المستخدم أو كلمة المرور غير صحيحة', 'error')
                    
        except Exception as e:
            app.logger.error(f'Unexpected error during login: {str(e)}')
            flash('حدث خطأ غير متوقع. يرجى المحاولة مرة أخرى.', 'error')
    
    return render_template('auth/login.html', form=form)

@app.route('/logout')
@login_required
def logout():
    # Log user logout
    app.logger.info(f'User {current_user.username} logged out from IP {request.remote_addr}')
    logout_user()
    session.clear()  # Clear all session data
    flash('تم تسجيل الخروج بنجاح', 'success')
    return redirect(url_for('index'))

@app.route('/forgot-password', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def forgot_password():
    """Password reset request"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        user = User.query.filter_by(email=email).first()
        
        if user and user.is_active:
            # Generate reset token
            token = user.generate_password_reset_token()
            
            # Send password reset email
            try:
                send_password_reset_email(user, token)
                flash('تم إرسال رابط إعادة تعيين كلمة المرور إلى بريدك الإلكتروني', 'success')
            except Exception as e:
                app.logger.error(f'Failed to send password reset email: {str(e)}')
                flash('حدث خطأ في إرسال البريد الإلكتروني. يرجى المحاولة لاحقاً.', 'error')
        else:
            # Same message for security (don't reveal if email exists)
            flash('تم إرسال رابط إعادة تعيين كلمة المرور إلى بريدك الإلكتروني', 'success')
        
        return redirect(url_for('index'))
    
    return render_template('auth/forgot_password.html')

@app.route('/reset-password/<token>', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def reset_password(token):
    """Reset password with token"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    # Find user with valid token
    user = None
    for u in User.query.all():
        if u.verify_password_reset_token(token):
            user = u
            break
    
    if not user:
        flash('رابط إعادة تعيين كلمة المرور غير صالح أو منتهي الصلاحية', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        new_password = request.form.get('password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        
        # Validate password
        if len(new_password) < 8:
            flash('كلمة المرور يجب أن تكون 8 أحرف على الأقل', 'error')
        elif new_password != confirm_password:
            flash('كلمات المرور غير متطابقة', 'error')
        else:
            # Reset password
            user.reset_password(new_password)
            flash('تم تغيير كلمة المرور بنجاح. يمكنك الآن تسجيل الدخول.', 'success')
            return redirect(url_for('index'))
    
    return render_template('auth/reset_password.html', token=token)

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    """Change password for logged-in user"""
    if request.method == 'POST':
        current_password = request.form.get('current_password', '').strip()
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        
        # Verify current password
        if not current_user.check_password(current_password):
            flash('كلمة المرور الحالية غير صحيحة', 'error')
        elif len(new_password) < 8:
            flash('كلمة المرور الجديدة يجب أن تكون 8 أحرف على الأقل', 'error')
        elif new_password != confirm_password:
            flash('كلمات المرور الجديدة غير متطابقة', 'error')
        else:
            # Change password
            current_user.set_password(new_password)
            db.session.commit()
            flash('تم تغيير كلمة المرور بنجاح', 'success')
            
            # Clear pending user session if exists
            session.pop('pending_user_id', None)
            return redirect(url_for('dashboard'))
    
    return render_template('auth/change_password.html')

def send_password_reset_email(user, token):
    """Send password reset email"""
    if not app.config.get('MAIL_USERNAME'):
        raise Exception("Email not configured")
    
    subject = 'إعادة تعيين كلمة المرور - إدارة Sara Store'
    reset_url = url_for('reset_password', token=token, _external=True)
    
    body = f"""
    مرحباً {user.username},
    
    تم طلب إعادة تعيين كلمة المرور لحسابك.
    
    للمتابعة، اضغط على الرابط التالي:
    {reset_url}
    
    هذا الرابط صالح لمدة ساعة واحدة فقط.
    
    إذا لم تطلب إعادة تعيين كلمة المرور، يرجى تجاهل هذه الرسالة.
    
    تحياتنا،
    فريق إدارة Sara Store
    """
    
    msg = Message(
        subject=subject,
        recipients=[user.email],
        body=body
    )
    
    mail.send(msg)

@app.route('/dashboard')
@login_required
def dashboard():
    # Get statistics
    total_products = Product.query.count()
    low_stock_products = Product.query.filter(Product.stock_quantity <= Product.min_stock_threshold).count()
    out_of_stock_products = Product.query.filter(Product.stock_quantity <= 0).count()
    total_categories = Category.query.count()
    
    # Sales statistics
    today = datetime.now().date()
    today_sales = Sale.query.filter(func.date(Sale.sale_date) == today).all()
    today_revenue = sum(sale.total_amount for sale in today_sales)
    
    # This month sales
    month_start = datetime.now().replace(day=1).date()
    month_sales = Sale.query.filter(func.date(Sale.sale_date) >= month_start).all()
    month_revenue = sum(sale.total_amount for sale in month_sales)
    
    # حساب أرباح ومصاريف الشهر الحالي
    month_profit = sum(sale.total_profit for sale in month_sales)
    month_cost = sum(sale.cost_amount for sale in month_sales)
    
    # مصاريف الشهر الحالي
    month_expenses = Expense.query.filter(func.date(Expense.expense_date) >= month_start).all()
    month_total_expenses = sum(expense.amount for expense in month_expenses)
    
    # صافي ربح الشهر
    month_net_profit = month_profit - month_total_expenses
    
    # إحصائيات اليوم
    today_profit = sum(sale.total_profit for sale in today_sales)
    today_expenses = Expense.query.filter(func.date(Expense.expense_date) == today).all()
    today_total_expenses = sum(expense.amount for expense in today_expenses)
    today_net_profit = today_profit - today_total_expenses
    
    # Recent sales
    recent_sales = Sale.query.order_by(desc(Sale.sale_date)).limit(5).all()
    
    # Low stock alerts
    low_stock_alerts = Product.query.filter(Product.stock_quantity <= Product.min_stock_threshold).all()
    
    # Top selling products (this month)
    top_products = db.session.query(
        Product.name_ar,
        func.sum(SaleItem.quantity).label('total_sold')
    ).join(SaleItem).join(Sale).filter(
        func.date(Sale.sale_date) >= month_start
    ).group_by(Product.id).order_by(desc('total_sold')).limit(5).all()
    
    # إحصائيات الديون
    # حساب إجمالي المبيعات الآجلة (غير المدفوعة بالكامل)
    unpaid_sales_total = db.session.query(func.sum(Sale.total_amount)).filter(
        Sale.payment_status != 'paid'
    ).scalar() or 0
    
    # حساب إجمالي الدفعات للمبيعات الآجلة
    total_payments = db.session.query(func.sum(Payment.amount)).join(Sale).filter(
        Sale.payment_status != 'paid'
    ).scalar() or 0
    
    # إجمالي الديون = إجمالي المبيعات الآجلة - إجمالي الدفعات
    total_debt = max(0, unpaid_sales_total - total_payments)
    
    customers_with_debt = Customer.query.filter(Customer.id.in_(
        db.session.query(Sale.customer_id).filter(Sale.payment_status != 'paid').distinct()
    )).count()
    
    return render_template('dashboard.html', 
                         total_products=total_products,
                         low_stock_products=low_stock_products,
                         out_of_stock_products=out_of_stock_products,
                         total_categories=total_categories,
                         today_revenue=today_revenue,
                         month_revenue=month_revenue,
                         # إحصائيات الأرباح والمصاريف
                         month_profit=month_profit,
                         month_cost=month_cost,
                         month_total_expenses=month_total_expenses,
                         month_net_profit=month_net_profit,
                         today_profit=today_profit,
                         today_expenses=today_total_expenses,  # Fixed: added missing today_expenses
                         today_total_expenses=today_total_expenses,
                         today_net_profit=today_net_profit,
                         recent_sales=recent_sales,
                         low_stock_alerts=low_stock_alerts,
                         top_products=top_products,
                         total_debt=total_debt,
                         customers_with_debt=customers_with_debt)

@app.route('/products')
@login_required
@admin_required
def products():
    search = request.args.get('search', '', type=str)
    category_id = request.args.get('category', 0, type=int)
    stock_status = request.args.get('stock_status', '', type=str)
    min_price = request.args.get('min_price', type=float)
    max_price = request.args.get('max_price', type=float)
    unit_type = request.args.get('unit_type', '', type=str)
    sort_by = request.args.get('sort_by', 'name', type=str)
    
    query = Product.query
    
    # تطبيق فلاتر البحث
    if search:
        query = query.filter(
            Product.name_ar.contains(search) | 
            Product.description_ar.contains(search)
        )
    
    if category_id:
        query = query.filter(Product.category_id == category_id)
    
    if stock_status:
        if stock_status == 'available':
            query = query.filter(Product.stock_quantity > Product.min_stock_threshold)
        elif stock_status == 'low':
            query = query.filter(
                Product.stock_quantity <= Product.min_stock_threshold,
                Product.stock_quantity > 0
            )
        elif stock_status == 'out':
            query = query.filter(Product.stock_quantity <= 0)
    
    if min_price is not None:
        query = query.filter(Product.retail_price >= min_price)
    
    if max_price is not None:
        query = query.filter(Product.retail_price <= max_price)
    
    if unit_type:
        query = query.filter(Product.unit_type == unit_type)
    
    # ترتيب النتائج
    if sort_by == 'name':
        query = query.order_by(Product.name_ar)
    elif sort_by == 'price':
        query = query.order_by(desc(Product.retail_price))
    elif sort_by == 'stock':
        query = query.order_by(desc(Product.stock_quantity))
    elif sort_by == 'date':
        query = query.order_by(desc(Product.created_at))
    else:
        query = query.order_by(Product.name_ar)
    
    # الحصول على جميع المنتجات (بدون تقسيم)
    products = query.all()
    categories = Category.query.all()
    
    # حساب الإحصائيات
    total_wholesale_value = 0
    total_retail_value = 0
    total_profit = 0
    total_products_count = len(products)
    total_stock_quantity = 0
    
    for product in products:
        wholesale_price = product.wholesale_price or 0
        retail_price = product.retail_price or product.price or 0
        stock_quantity = product.stock_quantity or 0
        
        total_wholesale_value += wholesale_price * stock_quantity
        total_retail_value += retail_price * stock_quantity
        total_profit += (retail_price - wholesale_price) * stock_quantity
        total_stock_quantity += stock_quantity
    
    # إحصائيات إضافية
    low_stock_count = sum(1 for p in products if p.is_low_stock)
    out_of_stock_count = sum(1 for p in products if p.is_out_of_stock)
    
    product_stats = {
        'total_products_count': total_products_count,
        'total_wholesale_value': total_wholesale_value,
        'total_retail_value': total_retail_value,
        'total_profit': total_profit,
        'total_stock_quantity': total_stock_quantity,
        'low_stock_count': low_stock_count,
        'out_of_stock_count': out_of_stock_count,
        'profit_margin_percentage': (total_profit / total_wholesale_value * 100) if total_wholesale_value > 0 else 0
    }
    
    return render_template('products/list.html', 
                         products=products, 
                         categories=categories,
                         product_stats=product_stats)

@app.route('/products/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_product():
    if not current_user.is_admin():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('products'))
    
    form = ProductForm()
    if form.validate_on_submit():
        product = Product(
            name_ar=form.name_ar.data,
            description_ar=form.description_ar.data,
            category_id=form.category_id.data,
            wholesale_price=form.wholesale_price.data,
            retail_price=form.retail_price.data,
            price=form.retail_price.data,  # للتوافق مع الكود القديم
            stock_quantity=form.stock_quantity.data,
            min_stock_threshold=form.min_stock_threshold.data,
            unit_type=form.unit_type.data,
            unit_description=form.unit_description.data
        )
        db.session.add(product)
        db.session.commit()
        
        # التحقق من نوع الإجراء المطلوب
        action = request.form.get('action', 'save_and_exit')
        
        if action == 'save_and_continue':
            flash('تم إضافة المنتج بنجاح! يمكنك إضافة منتج آخر.', 'success')
            return redirect(url_for('add_product', success=1))
        else:
            flash('تم إضافة المنتج بنجاح', 'success')
            return redirect(url_for('products'))
    
    return render_template('products/form.html', form=form, title='إضافة منتج جديد')

@app.route('/products/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_product(id):
    if not current_user.is_admin():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('products'))
    
    product = Product.query.get_or_404(id)
    form = ProductForm(obj=product)
    
    if form.validate_on_submit():
        form.populate_obj(product)
        product.updated_at = datetime.utcnow()
        db.session.commit()
        
        # التحقق من نوع الإجراء المطلوب
        action = request.form.get('action', 'save_and_exit')
        
        if action == 'save_and_continue':
            flash('تم تحديث المنتج بنجاح! يمكنك إضافة منتج جديد.', 'success')
            return redirect(url_for('add_product', success=1))
        else:
            flash('تم تحديث المنتج بنجاح', 'success')
            return redirect(url_for('products'))
    
    return render_template('products/form.html', form=form, title='تعديل المنتج')

@app.route('/products/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_product(id):
    if not current_user.is_admin():
        flash('ليس لديك صلاحية لحذف المنتجات', 'error')
        return redirect(url_for('products'))
    
    product = Product.query.get_or_404(id)
    db.session.delete(product)
    db.session.commit()
    flash('تم حذف المنتج بنجاح', 'success')
    return redirect(url_for('products'))

@app.route('/categories')
@login_required
@admin_required
def categories():
    if not current_user.is_admin():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('dashboard'))
    
    categories = Category.query.order_by(Category.name_ar).all()
    return render_template('categories/list.html', categories=categories)

@app.route('/categories/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_category():
    if not current_user.is_admin():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('categories'))
    
    form = CategoryForm()
    if form.validate_on_submit():
        category = Category(
            name_ar=form.name_ar.data,
            description_ar=form.description_ar.data
        )
        db.session.add(category)
        db.session.commit()
        flash('تم إضافة الفئة بنجاح', 'success')
        return redirect(url_for('categories'))
    
    return render_template('categories/form.html', form=form, title='إضافة فئة جديدة')

@app.route('/categories/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_category(id):
    if not current_user.is_admin():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('categories'))
    
    category = Category.query.get_or_404(id)
    form = CategoryForm(obj=category)
    
    if form.validate_on_submit():
        form.populate_obj(category)
        db.session.commit()
        flash('تم تحديث الفئة بنجاح', 'success')
        return redirect(url_for('categories'))
    
    return render_template('categories/form.html', form=form, title='تعديل الفئة')

@app.route('/categories/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_category(id):
    """حذف فئة"""
    if not current_user.is_admin():
        flash('ليس لديك صلاحية لحذف الفئات', 'error')
        return redirect(url_for('categories'))
    
    category = Category.query.get_or_404(id)
    
    # التحقق من عدم وجود منتجات في الفئة
    if category.products:
        flash('لا يمكن حذف فئة تحتوي على منتجات', 'error')
        return redirect(url_for('categories'))
    
    category_name = category.name_ar
    db.session.delete(category)
    db.session.commit()
    flash(f'تم حذف الفئة "{category_name}" بنجاح', 'success')
    return redirect(url_for('categories'))

# User management routes
@app.route('/users')
@login_required
@admin_required
def users():
    """عرض قائمة المستخدمين"""
    if not current_user.is_admin():
        flash('ليس لديك صلاحية للوصول إلى إدارة المستخدمين', 'error')
        return redirect(url_for('dashboard'))
    
    # إخفاء مستخدمي النظام من القائمة
    users = User.query.filter_by(is_system=False).order_by(User.created_at.desc()).all()
    return render_template('users/list.html', users=users)

@app.route('/users/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_user():
    """إضافة مستخدم جديد"""
    if not current_user.is_admin():
        flash('ليس لديك صلاحية لإضافة المستخدمين', 'error')
        return redirect(url_for('dashboard'))
    
    form = UserForm()
    if form.validate_on_submit():
        user = User(
            username=form.username.data,
            role=form.role.data
        )
        user.set_password(form.password.data)
        
        try:
            db.session.add(user)
            db.session.commit()
            flash(f'تم إضافة المستخدم "{user.username}" بنجاح', 'success')
            return redirect(url_for('users'))
        except Exception as e:
            db.session.rollback()
            flash('حدث خطأ أثناء إضافة المستخدم', 'error')
    
    return render_template('users/add.html', form=form)

@app.route('/users/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(id):
    """تعديل مستخدم"""
    if not current_user.is_admin():
        flash('ليس لديك صلاحية لتعديل المستخدمين', 'error')
        return redirect(url_for('dashboard'))
    
    user = User.query.get_or_404(id)
    
    # منع تعديل مستخدم النظام
    if user.is_system:
        flash('لا يمكن تعديل مستخدم النظام', 'error')
        return redirect(url_for('users'))
    
    # منع تعديل نفس المستخدم
    if user.id == current_user.id:
        flash('لا يمكنك تعديل حسابك الشخصي من هنا', 'error')
        return redirect(url_for('users'))
    
    form = UserForm(original_username=user.username, is_edit=True)
    if form.validate_on_submit():
        user.username = form.username.data
        user.role = form.role.data
        
        # تحديث كلمة المرور إذا تم إدخال واحدة جديدة
        if form.password.data:
            user.set_password(form.password.data)
        
        try:
            db.session.commit()
            flash(f'تم تحديث بيانات المستخدم "{user.username}" بنجاح', 'success')
            return redirect(url_for('users'))
        except Exception as e:
            db.session.rollback()
            flash('حدث خطأ أثناء تحديث المستخدم', 'error')
    
    # ملء النموذج بالبيانات الحالية
    if request.method == 'GET':
        form.username.data = user.username
        form.role.data = user.role
        form.password.data = ''  # لا نعرض كلمة المرور
    
    return render_template('users/edit.html', form=form, user=user)

@app.route('/users/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_user(id):
    """حذف مستخدم"""
    if not current_user.is_admin():
        flash('ليس لديك صلاحية لحذف المستخدمين', 'error')
        return redirect(url_for('dashboard'))
    
    user = User.query.get_or_404(id)
    
    # منع حذف مستخدم النظام
    if user.is_system:
        flash('لا يمكن حذف مستخدم النظام', 'error')
        return redirect(url_for('users'))
    
    # منع حذف نفس المستخدم
    if user.id == current_user.id:
        flash('لا يمكنك حذف حسابك الشخصي', 'error')
        return redirect(url_for('users'))
    
    # التحقق من وجود مبيعات للمستخدم
    if user.sales:
        flash('لا يمكن حذف مستخدم لديه مبيعات مسجلة', 'error')
        return redirect(url_for('users'))
    
    username = user.username
    db.session.delete(user)
    db.session.commit()
    flash(f'تم حذف المستخدم "{username}" بنجاح', 'success')
    return redirect(url_for('users'))

# Customer management routes
@app.route('/customers')
@login_required
@admin_required
def customers():
    """عرض قائمة العملاء"""
    customers = Customer.query.order_by(Customer.name).all()
    return render_template('customers/list.html', customers=customers)

@app.route('/customers/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_customer():
    """إضافة عميل جديد"""
    if request.method == 'POST':
        # Check if it's a JSON request (for quick add from sales page)
        if request.is_json:
            data = request.get_json()
            customer = Customer(
                name=data['name'],
                phone=data.get('phone', ''),
                address=data.get('address', ''),
                notes=data.get('notes', '')
            )
            db.session.add(customer)
            db.session.commit()
            return jsonify({'success': True, 'customer_id': customer.id})
        
        # Regular form submission
        form = CustomerForm()
        if form.validate_on_submit():
            customer = Customer(
                name=form.name.data,
                phone=form.phone.data,
                address=form.address.data,
                notes=form.notes.data
            )
            db.session.add(customer)
            db.session.commit()
            flash(f'تم إضافة العميل "{customer.name}" بنجاح', 'success')
            return redirect(url_for('customers'))
        return render_template('customers/form.html', form=form, title='إضافة عميل جديد')
    
    # GET request
    form = CustomerForm()
    return render_template('customers/form.html', form=form, title='إضافة عميل جديد')

@app.route('/customers/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_customer(id):
    """تعديل بيانات عميل"""
    customer = Customer.query.get_or_404(id)
    form = CustomerForm(obj=customer)
    if form.validate_on_submit():
        customer.name = form.name.data
        customer.phone = form.phone.data
        customer.address = form.address.data
        customer.notes = form.notes.data
        db.session.commit()
        flash(f'تم تحديث بيانات العميل "{customer.name}" بنجاح', 'success')
        return redirect(url_for('customers'))
    return render_template('customers/form.html', form=form, title='تعديل بيانات العميل', customer=customer)

@app.route('/customers/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_customer(id):
    """حذف عميل"""
    customer = Customer.query.get_or_404(id)
    
    # Check if customer has any sales
    if customer.sales:
        flash('لا يمكن حذف العميل لأن لديه مبيعات مسجلة', 'error')
        return redirect(url_for('customers'))
    
    customer_name = customer.name
    db.session.delete(customer)
    db.session.commit()
    flash(f'تم حذف العميل "{customer_name}" بنجاح', 'success')
    return redirect(url_for('customers'))

@app.route('/customers/<int:id>/account')
@login_required
@admin_required
def customer_account(id):
    """عرض حساب العميل والديون"""
    customer = Customer.query.get_or_404(id)
    sales = Sale.query.filter_by(customer_id=id).order_by(desc(Sale.sale_date)).all()
    return render_template('customers/account.html', customer=customer, sales=sales)

@app.route('/customers/<int:customer_id>/sales/<int:sale_id>/payment', methods=['GET', 'POST'])
@login_required
@admin_required
def add_payment(customer_id, sale_id):
    """إضافة دفعة لمبيعة"""
    customer = Customer.query.get_or_404(customer_id)
    sale = Sale.query.get_or_404(sale_id)
    
    if sale.customer_id != customer_id:
        flash('خطأ في بيانات العميل أو المبيعة', 'error')
        return redirect(url_for('customers'))
    
    if sale.is_fully_paid:
        flash('هذه المبيعة مدفوعة بالكامل', 'info')
        return redirect(url_for('customer_account', id=customer_id))
    
    form = PaymentForm()
    if form.validate_on_submit():
        # التأكد من أن المبلغ لا يتجاوز المتبقي
        remaining = sale.remaining_amount
        if form.amount.data > remaining:
            flash(f'المبلغ المدخل أكبر من المتبقي ({remaining:.2f} ج.م)', 'error')
        else:
            payment = Payment(
                sale_id=sale_id,
                amount=form.amount.data,
                payment_method=form.payment_method.data,
                notes=form.notes.data,
                user_id=current_user.id
            )
            db.session.add(payment)
            
            # تحديث حالة الدفع
            total_paid = sale.paid_amount + form.amount.data
            if total_paid >= sale.total_amount:
                sale.payment_status = 'paid'
            else:
                sale.payment_status = 'partial'
            
            db.session.commit()
            flash(f'تم تسجيل دفعة بمبلغ {form.amount.data:.2f} ج.م بنجاح', 'success')
            return redirect(url_for('customer_account', id=customer_id))
    
    return render_template('customers/payment.html', form=form, customer=customer, sale=sale)

@app.route('/debts')
@login_required
@admin_required
def debts_report():
    """تقرير الديون"""
    # العملاء الذين لديهم ديون
    customers_with_debts = []
    customers = Customer.query.all()
    
    for customer in customers:
        debt = customer.total_debt
        if debt > 0:
            customers_with_debts.append({
                'customer': customer,
                'debt': debt,
                'unpaid_sales': [sale for sale in customer.sales if not sale.is_fully_paid]
            })
    
    # ترتيب حسب قيمة الدين (الأكبر أولاً)
    customers_with_debts.sort(key=lambda x: x['debt'], reverse=True)
    
    total_debts = sum(item['debt'] for item in customers_with_debts)
    
    return render_template('debts/report.html', 
                         customers_with_debts=customers_with_debts, 
                         total_debts=total_debts,
                         current_datetime=datetime.now())

@app.route('/sales')
@login_required
@seller_or_admin_required
def sales():
    page = request.args.get('page', 1, type=int)
    
    # Start with base query
    query = Sale.query
    
    # Apply filters from request arguments
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    product_search = request.args.get('product_search')
    seller_filter = request.args.get('seller_filter')
    amount_from = request.args.get('amount_from')
    amount_to = request.args.get('amount_to')
    
    # Date range filter
    if date_from:
        try:
            date_from_dt = datetime.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(func.date(Sale.sale_date) >= date_from_dt)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_dt = datetime.strptime(date_to, '%Y-%m-%d').date()
            query = query.filter(func.date(Sale.sale_date) <= date_to_dt)
        except ValueError:
            pass
    
    # Product search filter
    if product_search:
        # Join with SaleItem and Product to search in product names
        query = query.join(SaleItem).join(Product).filter(
            Product.name_ar.contains(product_search)
        ).distinct()
    
    # Seller filter
    if seller_filter:
        try:
            seller_id = int(seller_filter)
            query = query.filter(Sale.user_id == seller_id)
        except ValueError:
            pass
    
    # Amount range filter
    if amount_from:
        try:
            amount_from_val = float(amount_from)
            query = query.filter(Sale.total_amount >= amount_from_val)
        except ValueError:
            pass
    
    if amount_to:
        try:
            amount_to_val = float(amount_to)
            query = query.filter(Sale.total_amount <= amount_to_val)
        except ValueError:
            pass
    
    # Apply user permissions
    if not current_user.is_admin():
        query = query.filter(Sale.user_id == current_user.id)
    
    # Order by newest first and paginate
    sales = query.order_by(desc(Sale.sale_date)).paginate(
        page=page, per_page=20, error_out=False)
    
    # Get all users for seller filter dropdown
    users = User.query.all() if current_user.is_admin() else [current_user]
    
    return render_template('sales/list.html', sales=sales, users=users)

@app.route('/sales/new')
@login_required
@seller_or_admin_required
def new_sale():
    return render_template('sales/new.html')

@app.route('/api/products')
@login_required
@seller_or_admin_required
def api_products():
    """API endpoint to get all products"""
    try:
        app.logger.info(f"API products called by user: {current_user.username}")
        products = Product.query.all()
        app.logger.info(f"Found {len(products)} products in database")
        
        result = []
        for p in products:
            try:
                # Use safe attribute access with fallbacks
                wholesale_price = p.wholesale_price if p.wholesale_price else (p.price if p.price else 0)
                retail_price = p.retail_price if p.retail_price else (p.price if p.price else 0)
                
                product_data = {
                    'id': p.id,
                    'name': p.name_ar or 'منتج غير محدد',
                    'wholesale_price': float(wholesale_price),
                    'retail_price': float(retail_price),
                    'price': float(retail_price),  # Use retail_price as the main price
                    'stock': float(p.stock_quantity or 0),
                    'unit_type': p.unit_type or 'كامل',
                    'category': p.category.name_ar if p.category else 'غير محدد',
                    'min_stock_threshold': float(p.min_stock_threshold or 10),
                    'profit_margin': 0,
                    'profit_percentage': 0
                }
                
                # Add profit calculations if possible
                try:
                    if wholesale_price > 0 and retail_price > 0:
                        profit_margin = retail_price - wholesale_price
                        profit_percentage = (profit_margin / wholesale_price) * 100
                        product_data['profit_margin'] = float(profit_margin)
                        product_data['profit_percentage'] = float(profit_percentage)
                except:
                    pass
                
                result.append(product_data)
            except Exception as e:
                # Skip problematic products but log the error
                app.logger.error(f"Error processing product {p.id}: {str(e)}")
                continue
        
        app.logger.info(f"Returning {len(result)} products to client")
        return jsonify(result)
    except Exception as e:
        app.logger.error(f"Error in api_products: {str(e)}")
        return jsonify({'error': 'حدث خطأ في تحميل المنتجات'}), 500

@app.route('/api/categories')
@login_required
def api_categories():
    """API endpoint to get all categories"""
    categories = Category.query.all()
    return jsonify([{
        'id': c.id,
        'name': c.name_ar,
        'description': c.description_ar or '',
        'product_count': len(c.products),
        'created_at': c.created_at.isoformat() if hasattr(c, 'created_at') and c.created_at else None
    } for c in categories])

@app.route('/api/customers')
@login_required
@seller_or_admin_required
def api_customers():
    """API endpoint to get all customers"""
    customers = Customer.query.order_by(Customer.name).all()
    return jsonify([{
        'id': c.id,
        'name': c.name,
        'phone': c.phone or '',
        'debt': c.total_debt
    } for c in customers])

@app.route('/api/sales', methods=['POST'])
@login_required
@seller_or_admin_required
def api_create_sale():
    """API endpoint to create a new sale"""
    data = request.get_json()
    
    if not data.get('items'):
        return jsonify({'error': 'لا توجد عناصر في البيع'}), 400
    
    # Validate stock availability
    for item in data['items']:
        product = Product.query.get(item['product_id'])
        if not product:
            return jsonify({'error': f'المنتج غير موجود'}), 400
        if product.stock_quantity < item['quantity']:
            return jsonify({'error': f'الكمية المطلوبة غير متوفرة للمنتج {product.name_ar}'}), 400
    
    # Get payment info
    payment_type = data.get('payment_type', 'cash')  # 'cash' or 'credit'
    customer_id = data.get('customer_id') if payment_type == 'credit' else None
    paid_amount = float(data.get('paid_amount', 0))
    
    # Get discount info
    subtotal = float(data.get('subtotal', 0))
    discount_type = data.get('discount_type', 'none')
    discount_value = float(data.get('discount_value', 0))
    discount_amount = float(data.get('discount_amount', 0))
    total_amount = float(data['total_amount'])
    
    # Validate customer for credit sales
    if payment_type == 'credit' and not customer_id:
        return jsonify({'error': 'يجب اختيار عميل للبيع الآجل'}), 400
    
    # تحديد حالة الدفع
    if payment_type == 'cash':
        payment_status = 'paid'
        paid_amount = total_amount  # في البيع النقدي يكون المبلغ مدفوع بالكامل
    else:
        if paid_amount >= total_amount:
            payment_status = 'paid'
        elif paid_amount > 0:
            payment_status = 'partial'
        else:
            payment_status = 'unpaid'
    
    # Create sale with discount info
    sale = Sale(
        subtotal=subtotal,
        discount_type=discount_type,
        discount_value=discount_value,
        discount_amount=discount_amount,
        total_amount=total_amount,
        user_id=current_user.id,
        customer_id=customer_id,
        payment_type=payment_type,
        payment_status=payment_status,
        notes=data.get('notes', '')
    )
    db.session.add(sale)
    db.session.flush()  # Get sale.id
    
    # Create sale items and update stock
    for item in data['items']:
        sale_item = SaleItem(
            sale_id=sale.id,
            product_id=item['product_id'],
            quantity=item['quantity'],
            unit_price=item['unit_price'],
            total_price=item['total_price']
        )
        db.session.add(sale_item)
        
        # Update product stock
        product = Product.query.get(item['product_id'])
        product.stock_quantity -= item['quantity']
    
    # إضافة دفعة في حالة البيع الآجل مع دفعة مقدمة
    if payment_type == 'credit' and paid_amount > 0:
        payment = Payment(
            sale_id=sale.id,
            amount=paid_amount,
            payment_method='نقدي',
            notes='دفعة مقدمة مع البيع',
            user_id=current_user.id
        )
        db.session.add(payment)
    
    db.session.commit()
    
    # تحديد الرسالة
    if payment_type == 'cash':
        message = 'تم تسجيل البيع نقداً بنجاح'
    elif payment_status == 'paid':
        message = 'تم تسجيل البيع وتم دفع المبلغ كاملاً'
    elif payment_status == 'partial':
        remaining = total_amount - paid_amount
        message = f'تم تسجيل البيع مع دفعة مقدمة {paid_amount:.2f} ج.م - المتبقي: {remaining:.2f} ج.م'
    else:
        message = 'تم تسجيل البيع آجلاً'
    
    return jsonify({
        'success': True,
        'sale_id': sale.id,
        'message': message,
        'payment_status': payment_status,
        'paid_amount': paid_amount,
        'remaining_amount': total_amount - paid_amount
    })

@app.route('/api/sales/<int:sale_id>')
@login_required
def api_sale_details(sale_id):
    """API endpoint to get sale details"""
    sale = Sale.query.get_or_404(sale_id)
    
    # Check permissions - sellers can only view their own sales, admins can view all
    if not current_user.is_admin() and sale.user_id != current_user.id:
        return jsonify({'error': 'ليس لديك صلاحية لعرض هذا البيع'}), 403
    
    # Get sale items
    sale_items = []
    for item in sale.sale_items:
        sale_items.append({
            'product_name': item.product.name_ar,
            'quantity': float(item.quantity),
            'unit_price': float(item.unit_price),
            'total_price': float(item.total_price),
            'unit_type': item.product.unit_type
        })
    
    # تحويل التاريخ إلى توقيت مصر
    egypt_time = get_egypt_time(sale.sale_date)
    time_str = egypt_time.strftime('%I:%M:%S %p').replace('AM', 'ص').replace('PM', 'م')
    datetime_str = egypt_time.strftime('%d/%m/%Y %I:%M:%S %p').replace('AM', 'ص').replace('PM', 'م')
    
    return jsonify({
        'id': sale.id,
        'sale_date': egypt_time.strftime('%d/%m/%Y'),
        'sale_time': time_str,
        'sale_datetime': datetime_str,
        'subtotal': float(sale.subtotal or sale.total_amount),
        'discount_type': sale.discount_type or 'none',
        'discount_value': float(sale.discount_value or 0),
        'discount_amount': float(sale.discount_amount or 0),
        'total_amount': float(sale.total_amount),
        'notes': sale.notes,
        'user_name': sale.user.username,
        'user_role': sale.user.role,
        'items': sale_items
    })

# New API endpoints for data export
@app.route('/api/export/products')
@login_required
def api_export_products():
    if current_user.role not in ['admin', 'seller']:
        abort(403)
    
    products = Product.query.join(Category).all()
    return jsonify([{
        'id': p.id,
        'name': p.name_ar,
        'category': p.category.name_ar if p.category else 'غير محدد',
        'price': float(p.retail_price or p.price or 0),
        'stock': float(p.stock_quantity),
        'unit_type': p.unit_type,
        'is_whole_unit': p.is_whole_unit,
        'created_date': p.id  # Using id as proxy for creation order
    } for p in products])

@app.route('/api/export/sales')
@login_required
def api_export_sales():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = Sale.query.join(User)
    
    if start_date:
        query = query.filter(Sale.sale_date >= start_date)
    if end_date:
        query = query.filter(Sale.sale_date <= end_date + ' 23:59:59')
    
    # Filter by user role
    if current_user.role != 'admin':
        query = query.filter(Sale.user_id == current_user.id)
    
    sales = query.order_by(Sale.sale_date.desc()).all()
    
    sales_data = []
    for sale in sales:
        for item in sale.sale_items:
            sales_data.append({
                'sale_id': sale.id,
                'sale_date': format_egypt_date_only(sale.sale_date),
                'sale_time': format_egypt_time_only(sale.sale_date),
                'seller_name': sale.user.username,
                'seller_role': sale.user.role,
                'product_name': item.product.name_ar,
                'product_category': item.product.category.name_ar if item.product.category else 'غير محدد',
                'quantity': float(item.quantity),
                'unit_price': float(item.unit_price),
                'total_price': float(item.total_price),
                'unit_type': item.product.unit_type,
                'sale_total': float(sale.total_amount),
                'notes': sale.notes or ''
            })
    
    return jsonify(sales_data)

@app.route('/api/quick-payment', methods=['POST'])
@login_required
def api_quick_payment():
    """API endpoint for quick debt payment"""
    try:
        data = request.get_json()
        customer_id = data.get('customer_id')
        amount = float(data.get('amount', 0))
        payment_method = data.get('payment_method', 'نقدي')
        notes = data.get('notes', '')
        
        if not customer_id or amount <= 0:
            return jsonify({'success': False, 'message': 'بيانات غير صحيحة'}), 400
        
        # Get customer
        customer = Customer.query.get_or_404(customer_id)
        
        # Get unpaid sales for this customer
        unpaid_sales = Sale.query.filter(
            Sale.customer_id == customer_id,
            Sale.payment_status.in_(['unpaid', 'partial'])
        ).order_by(Sale.sale_date.asc()).all()
        
        if not unpaid_sales:
            return jsonify({'success': False, 'message': 'لا توجد ديون لهذا العميل'}), 400
        
        remaining_amount = amount
        payments_made = []
        
        # Distribute payment across unpaid sales
        for sale in unpaid_sales:
            if remaining_amount <= 0:
                break
            
            sale_remaining = sale.remaining_amount
            if sale_remaining <= 0:
                continue
            
            payment_amount = min(remaining_amount, sale_remaining)
            
            # Create payment record
            payment = Payment(
                sale_id=sale.id,
                amount=payment_amount,
                payment_method=payment_method,
                notes=f"{notes} - تسديد سريع",
                user_id=current_user.id
            )
            
            db.session.add(payment)
            
            # Update sale payment status
            sale_total_paid = sale.paid_amount + payment_amount
            if sale_total_paid >= sale.total_amount:
                sale.payment_status = 'paid'
            else:
                sale.payment_status = 'partial'
            
            payments_made.append({
                'sale_id': sale.id,
                'amount': payment_amount
            })
            
            remaining_amount -= payment_amount
        
        db.session.commit()
        
        message = f"تم تسديد {amount:.2f} ج.م بنجاح"
        if remaining_amount > 0:
            message += f" (متبقي {remaining_amount:.2f} ج.م كرصيد)"
        
        return jsonify({
            'success': True,
            'message': message,
            'payments_made': payments_made
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/stock-status')
@login_required
def api_stock_status():
    """API endpoint for real-time stock monitoring"""
    try:
        # Get products with low stock or out of stock
        critical_products = Product.query.filter(Product.stock_quantity <= 0).all()
        low_stock_products = Product.query.filter(
            and_(Product.stock_quantity > 0, 
                 Product.stock_quantity <= Product.min_stock_threshold)
        ).all()
        
        new_alerts = []
        
        # Add critical stock alerts (out of stock)
        for product in critical_products:
            new_alerts.append({
                'type': 'critical',
                'product_name': product.name_ar,
                'current_stock': product.stock_quantity,
                'min_threshold': product.min_stock_threshold or 10,
                'message': f'نفدت كمية {product.name_ar} من المخزون'
            })
        
        # Add low stock alerts
        for product in low_stock_products:
            new_alerts.append({
                'type': 'warning',
                'product_name': product.name_ar,
                'current_stock': product.stock_quantity,
                'min_threshold': product.min_stock_threshold or 10,
                'message': f'كمية {product.name_ar} منخفضة في المخزون'
            })
        
        return jsonify({
            'status': 'success',
            'new_alerts': new_alerts,
            'critical_count': len(critical_products),
            'low_stock_count': len(low_stock_products),
            'timestamp': datetime.now().isoformat()
        })
    
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': 'حدث خطأ في مراقبة المخزون',
            'new_alerts': []
        }), 500

@app.route('/api/export/inventory')
@login_required
def api_export_inventory():
    if current_user.role not in ['admin', 'seller']:
        abort(403)
    
    # Get products with stock status
    products = Product.query.join(Category).all()
    
    inventory_data = []
    for product in products:
        # Calculate stock status
        if product.stock_quantity <= 0:
            status = 'نفدت الكمية'
            status_en = 'Out of Stock'
        elif product.stock_quantity <= 10:
            status = 'كمية قليلة'
            status_en = 'Low Stock'
        else:
            status = 'متوفر'
            status_en = 'Available'
        
        # Calculate total sales for this product
        total_sold = db.session.query(db.func.sum(SaleItem.quantity)).filter(
            SaleItem.product_id == product.id
        ).scalar() or 0
        
        total_revenue = db.session.query(db.func.sum(SaleItem.total_price)).filter(
            SaleItem.product_id == product.id
        ).scalar() or 0
        
        inventory_data.append({
            'product_id': product.id,
            'product_name': product.name_ar,
            'category': product.category.name_ar if product.category else 'غير محدد',
            'current_stock': float(product.stock_quantity),
            'unit_type': product.unit_type,
            'unit_price': float(product.price),
            'stock_value': float(product.stock_quantity * product.price),
            'is_whole_unit': product.is_whole_unit,
            'status_ar': status,
            'status_en': status_en,
            'total_sold': float(total_sold),
            'total_revenue': float(total_revenue)
        })
    
    return jsonify(inventory_data)

@app.route('/reports')
@login_required
@admin_required
def reports():
    if current_user.role not in ['admin', 'seller']:
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('dashboard'))
    
    # Get date range from request
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if not start_date:
        start_date = datetime.now().replace(day=1).strftime('%Y-%m-%d')
    if not end_date:
        end_date = datetime.now().strftime('%Y-%m-%d')
    
    # Convert to datetime objects
    start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Sales in date range
    sales = Sale.query.filter(
        and_(func.date(Sale.sale_date) >= start_dt,
             func.date(Sale.sale_date) <= end_dt)
    ).all()
    
    total_revenue = sum(sale.total_amount for sale in sales)
    total_sales_count = len(sales)
    
    # حساب الأرباح والتكاليف
    total_profit = sum(sale.total_profit for sale in sales)
    total_cost = sum(sale.cost_amount for sale in sales)
    profit_margin = (total_profit / total_revenue * 100) if total_revenue > 0 else 0
    
    # حساب المصاريف في نفس الفترة
    expenses = Expense.query.filter(
        and_(func.date(Expense.expense_date) >= start_dt,
             func.date(Expense.expense_date) <= end_dt)
    ).all()
    
    total_expenses = sum(expense.amount for expense in expenses)
    
    # صافي الربح = إجمالي الأرباح - المصاريف
    net_profit = total_profit - total_expenses
    
    # تصنيف المصاريف حسب النوع
    expenses_by_type = {}
    for expense in expenses:
        expense_type = expense.expense_type_ar
        if expense_type not in expenses_by_type:
            expenses_by_type[expense_type] = 0
        expenses_by_type[expense_type] += expense.amount
    
    # Top products in date range
    top_products = db.session.query(
        Product.name_ar,
        func.sum(SaleItem.quantity).label('total_sold'),
        func.sum(SaleItem.total_price).label('total_revenue')
    ).join(SaleItem).join(Sale).filter(
        and_(func.date(Sale.sale_date) >= start_dt,
             func.date(Sale.sale_date) <= end_dt)
    ).group_by(Product.id).order_by(desc('total_sold')).all()
    
    # Daily sales chart data
    daily_sales_raw = db.session.query(
        func.date(Sale.sale_date).label('date'),
        func.sum(Sale.total_amount).label('total')
    ).filter(
        and_(func.date(Sale.sale_date) >= start_dt,
             func.date(Sale.sale_date) <= end_dt)
    ).group_by(func.date(Sale.sale_date)).order_by('date').all()
    
    # Convert to JSON-serializable format
    daily_sales = []
    for row in daily_sales_raw:
        date_str = row.date if isinstance(row.date, str) else row.date.strftime('%Y-%m-%d')
        daily_sales.append({'date': date_str, 'total': float(row.total or 0)})
    
    # Debt-related statistics
    # Total debts across all customers
    total_debts = sum(customer.total_debt for customer in Customer.query.all())
    
    # Count customers with debts
    customers_with_debts = Customer.query.filter(
        Customer.id.in_(
            db.session.query(Sale.customer_id).filter(
                Sale.payment_status.in_(['unpaid', 'partial'])
            ).distinct()
        )
    ).count()
    
    # Credit sales in date range
    credit_sales = Sale.query.filter(
        and_(func.date(Sale.sale_date) >= start_dt,
             func.date(Sale.sale_date) <= end_dt,
             Sale.payment_type == 'credit')
    ).all()
    total_credit_sales = len(credit_sales)
    
    # Total payments in date range
    total_payments = db.session.query(func.sum(Payment.amount)).join(Sale).filter(
        and_(func.date(Payment.payment_date) >= start_dt,
             func.date(Payment.payment_date) <= end_dt)
    ).scalar() or 0
    
    # Payment rate calculation
    total_credit_amount = sum(sale.total_amount for sale in credit_sales)
    payment_rate = (total_payments / total_credit_amount * 100) if total_credit_amount > 0 else 0
    
    # Top debtors
    top_debtors = []
    customers_with_debt = Customer.query.join(Sale).filter(
        Sale.payment_status.in_(['unpaid', 'partial'])
    ).distinct().all()
    
    for customer in customers_with_debt:
        if customer.total_debt > 0:
            unpaid_sales_count = Sale.query.filter(
                Sale.customer_id == customer.id,
                Sale.payment_status.in_(['unpaid', 'partial'])
            ).count()
            
            last_sale = Sale.query.filter(
                Sale.customer_id == customer.id
            ).order_by(Sale.sale_date.desc()).first()
            
            top_debtors.append((
                customer,
                customer.total_debt,
                unpaid_sales_count,
                last_sale.sale_date if last_sale else None
            ))
    
    # Sort by debt amount (highest first) and take top 10
    top_debtors.sort(key=lambda x: x[1], reverse=True)
    top_debtors = top_debtors[:10]
    
    return render_template('reports/index.html',
                         start_date=start_date,
                         end_date=end_date,
                         total_revenue=total_revenue,
                         total_sales_count=total_sales_count,
                         top_products=top_products,
                         daily_sales=daily_sales,
                         # Profit and expense data
                         total_profit=total_profit,
                         total_cost=total_cost,
                         profit_margin=profit_margin,
                         total_expenses=total_expenses,
                         net_profit=net_profit,
                         expenses_by_type=expenses_by_type,
                         # Debt-related data
                         total_debts=total_debts,
                         customers_with_debts=customers_with_debts,
                         total_credit_sales=total_credit_sales,
                         total_payments=total_payments,
                         payment_rate=payment_rate,
                         top_debtors=top_debtors)

# إدارة المصاريف
@app.route('/expenses')
@login_required
@admin_required
def expenses():
    """عرض قائمة المصاريف"""
    if not current_user.is_admin():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('dashboard'))
    
    page = request.args.get('page', 1, type=int)
    expense_type = request.args.get('type', '', type=str)
    start_date = request.args.get('start_date', '', type=str)
    end_date = request.args.get('end_date', '', type=str)
    
    query = Expense.query
    
    if expense_type:
        query = query.filter(Expense.expense_type == expense_type)
    
    if start_date:
        query = query.filter(func.date(Expense.expense_date) >= start_date)
    
    if end_date:
        query = query.filter(func.date(Expense.expense_date) <= end_date)
    
    expenses = query.order_by(desc(Expense.expense_date)).paginate(
        page=page, per_page=20, error_out=False)
    
    # حساب إجمالي المصاريف
    total_expenses = query.with_entities(func.sum(Expense.amount)).scalar() or 0
    
    return render_template('expenses/list.html', 
                         expenses=expenses, 
                         total_expenses=total_expenses,
                         expense_type=expense_type,
                         start_date=start_date,
                         end_date=end_date)

@app.route('/expenses/add', methods=['GET', 'POST'])
@login_required
def add_expense():
    """إضافة مصروف جديد"""
    if not current_user.is_admin():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('dashboard'))
    
    form = ExpenseForm()
    if form.validate_on_submit():
        expense = Expense(
            description=form.description.data,
            amount=form.amount.data,
            expense_type=form.expense_type.data,
            category=form.category.data,
            expense_date=form.expense_date.data or datetime.utcnow(),
            notes=form.notes.data,
            user_id=current_user.id
        )
        db.session.add(expense)
        db.session.commit()
        flash('تم إضافة المصروف بنجاح', 'success')
        return redirect(url_for('expenses'))
    
    return render_template('expenses/add.html', form=form, title='إضافة مصروف جديد')

@app.route('/expenses/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_expense(id):
    """تعديل مصروف"""
    if not current_user.is_admin():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('expenses'))
    
    expense = Expense.query.get_or_404(id)
    form = ExpenseForm(obj=expense)
    
    if form.validate_on_submit():
        expense.description = form.description.data
        expense.amount = form.amount.data
        expense.expense_type = form.expense_type.data
        expense.category = form.category.data
        expense.expense_date = form.expense_date.data or expense.expense_date
        expense.notes = form.notes.data
        db.session.commit()
        flash('تم تحديث المصروف بنجاح', 'success')
        return redirect(url_for('expenses'))
    
    return render_template('expenses/edit.html', form=form, title='تعديل المصروف')

@app.route('/expenses/<int:id>/delete', methods=['POST'])
@login_required
def delete_expense(id):
    """حذف مصروف"""
    if not current_user.is_admin():
        flash('ليس لديك صلاحية لحذف المصاريف', 'error')
        return redirect(url_for('expenses'))
    
    expense = Expense.query.get_or_404(id)
    description = expense.description
    db.session.delete(expense)
    db.session.commit()
    flash(f'تم حذف المصروف "{description}" بنجاح', 'success')
    return redirect(url_for('expenses'))

# ==================== النواقص ====================

@app.route('/shopping-list')
@login_required
@admin_required
def shopping_list():
    """صفحة النواقص"""
    # الحصول على قائمة النواقص المطلوبة
    needed_items = ShoppingList.query.filter_by(status='مطلوب').order_by(
        ShoppingList.priority.desc(), ShoppingList.created_at.desc()
    ).all()
    
    # الحصول على المنتجات التي نفدت أو قاربت على النفاد
    out_of_stock = Product.query.filter(Product.stock_quantity <= 0).all()
    low_stock = Product.query.filter(
        and_(Product.stock_quantity > 0, 
             Product.stock_quantity <= Product.min_stock_threshold)
    ).all()
    
    # حساب إجمالي التكلفة المتوقعة
    total_estimated_cost = sum(item.total_estimated_cost for item in needed_items if item.estimated_price)
    
    return render_template('shopping/list.html', 
                         needed_items=needed_items,
                         out_of_stock=out_of_stock,
                         low_stock=low_stock,
                         total_estimated_cost=total_estimated_cost)

@app.route('/shopping-list/add', methods=['GET', 'POST'])
@login_required
def add_shopping_item():
    """إضافة منتج لقائمة النواقص"""
    form = ShoppingListForm()
    if form.validate_on_submit():
        try:
            shopping_item = ShoppingList(
                item_name=form.item_name.data,
                quantity_needed=form.quantity_needed.data,
                unit_type=form.unit_type.data,
                estimated_price=form.estimated_price.data,
                priority=form.priority.data,
                category=form.category.data,
                supplier=form.supplier.data,
                notes=form.notes.data,
                user_id=current_user.id
            )
            db.session.add(shopping_item)
            db.session.commit()
            flash('تم إضافة المنتج لقائمة النواقص بنجاح', 'success')
            return redirect(url_for('shopping_list'))
        except Exception as e:
            db.session.rollback()
            flash(f'خطأ في إضافة المنتج: {str(e)}', 'error')
    
    return render_template('shopping/add.html', form=form)

@app.route('/shopping-list/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_shopping_item(id):
    """تعديل منتج في قائمة النواقص"""
    item = ShoppingList.query.get_or_404(id)
    form = ShoppingListForm(obj=item)
    
    if form.validate_on_submit():
        try:
            item.item_name = form.item_name.data
            item.quantity_needed = form.quantity_needed.data
            item.unit_type = form.unit_type.data
            item.estimated_price = form.estimated_price.data
            item.priority = form.priority.data
            item.category = form.category.data
            item.supplier = form.supplier.data
            item.notes = form.notes.data
            
            db.session.commit()
            flash('تم تحديث المنتج بنجاح', 'success')
            return redirect(url_for('shopping_list'))
        except Exception as e:
            db.session.rollback()
            flash(f'خطأ في تحديث المنتج: {str(e)}', 'error')
    
    return render_template('shopping/edit.html', form=form, item=item)

@app.route('/shopping-list/<int:id>/delete', methods=['POST'])
@login_required
def delete_shopping_item(id):
    """حذف منتج من قائمة النواقص"""
    try:
        item = ShoppingList.query.get_or_404(id)
        db.session.delete(item)
        db.session.commit()
        flash('تم حذف المنتج من قائمة النواقص', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'خطأ في حذف المنتج: {str(e)}', 'error')
    
    return redirect(url_for('shopping_list'))

@app.route('/shopping-list/<int:id>/mark-purchased', methods=['POST'])
@login_required
def mark_purchased(id):
    """تحديد منتج كمُشترى"""
    try:
        item = ShoppingList.query.get_or_404(id)
        item.status = 'تم الشراء'
        item.purchased_date = datetime.utcnow()
        db.session.commit()
        flash('تم تحديد المنتج كمُشترى', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'خطأ في تحديث حالة المنتج: {str(e)}', 'error')
    
    return redirect(url_for('shopping_list'))

@app.route('/shopping-list/add-low-stock/<int:product_id>')
@login_required
def add_low_stock_product(product_id):
    """إضافة منتج منخفض المخزون لقائمة النواقص"""
    product = Product.query.get_or_404(product_id)
    
    # التحقق من عدم وجود المنتج في القائمة بالفعل
    existing = ShoppingList.query.filter_by(
        item_name=product.name_ar, 
        status='مطلوب'
    ).first()
    
    if existing:
        flash('هذا المنتج موجود بالفعل في قائمة النواقص', 'warning')
    else:
        try:
            # تحديد الكمية المقترحة (الحد الأدنى - الكمية الحالية)
            suggested_quantity = max(product.min_stock_threshold - product.stock_quantity, 10)
            
            shopping_item = ShoppingList(
                item_name=product.name_ar,
                quantity_needed=suggested_quantity,
                unit_type=product.unit_type,
                estimated_price=product.wholesale_price,
                priority='high' if product.is_out_of_stock else 'medium',
                category=product.category.name_ar if product.category else None,
                notes=f'منتج {"نفد" if product.is_out_of_stock else "منخفض"} من المخزون - الكمية الحالية: {product.stock_quantity}',
                user_id=current_user.id
            )
            db.session.add(shopping_item)
            db.session.commit()
            flash('تم إضافة المنتج لقائمة النواقص', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'خطأ في إضافة المنتج: {str(e)}', 'error')
    
    return redirect(url_for('shopping_list'))

@app.route('/test-export')
@login_required
def test_export():
    """صفحة اختبار وظائف التصدير"""
    return send_from_directory('.', 'test_export.html')

@app.route('/stock/update', methods=['GET', 'POST'])
@login_required
def update_stock():
    if not current_user.is_admin():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('dashboard'))
    
    form = StockUpdateForm()
    if form.validate_on_submit():
        product = Product.query.get(form.product_id.data)
        product.stock_quantity += form.quantity.data
        db.session.commit()
        flash(f'تم تحديث مخزون {product.name_ar} بنجاح', 'success')
        return redirect(url_for('products'))
    
    return render_template('stock/update.html', form=form)

def create_sample_data():
    """Create sample data for testing"""
    # Create default system admin user (hidden)
    system_admin = User.query.filter_by(username='araby').first()
    if not system_admin:
        system_admin = User(username='araby', role='admin', is_system=True)
        system_admin.set_password('92321066')
        db.session.add(system_admin)
    
    # Create admin user
    admin = User.query.filter_by(username='admin').first()
    if not admin:
        admin = User(username='admin', role='admin')
        admin.set_password('admin123')
        db.session.add(admin)
    
    # Create seller user
    seller = User.query.filter_by(username='seller').first()
    if not seller:
        seller = User(username='seller', role='seller')
        seller.set_password('seller123')
        db.session.add(seller)
    
    # Create comprehensive categories for bookstore and school supplies
    if Category.query.count() == 0:
        categories = [
            # كتب ومراجع
            Category(name_ar='كتب أدبية وروايات', description_ar='الروايات والقصص والشعر والأدب العربي والعالمي'),
            Category(name_ar='كتب علمية وتقنية', description_ar='الكتب العلمية والتقنية والهندسية والطبية'),
            Category(name_ar='كتب دراسية ومناهج', description_ar='المناهج الدراسية والكتب الجامعية والمدرسية'),
            Category(name_ar='كتب الأطفال', description_ar='القصص والكتب التعليمية للأطفال والرسوم المتحركة'),
            Category(name_ar='كتب دينية وإسلامية', description_ar='القرآن الكريم والتفاسير والكتب الدينية'),
            Category(name_ar='مجلات وصحف', description_ar='المجلات العلمية والثقافية والصحف اليومية'),
            
            # أدوات الكتابة
            Category(name_ar='أقلام حبر وجاف', description_ar='أقلام الحبر الجاف والسائل بألوان مختلفة'),
            Category(name_ar='أقلام رصاص وملونة', description_ar='أقلام الرصاص العادية والملونة وأقلام التلوين'),
            Category(name_ar='أقلام ماركر وتحديد', description_ar='أقلام الماركر والهايلايتر وأقلام التحديد'),
            Category(name_ar='محايات وبراية', description_ar='المحايات البيضاء والملونة وبرايات الأقلام'),
            
            # دفاتر وكشاكيل
            Category(name_ar='كشاكيل ودفاتر خانات', description_ar='الكشاكيل المخططة والمربعة وذات الخانات'),
            Category(name_ar='دفاتر مسطرة وسادة', description_ar='الدفاتر المسطرة والسادة للكتابة'),
            Category(name_ar='دفاتر رسم وفنية', description_ar='دفاتر الرسم والأوراق الفنية للرسم والتلوين'),
            Category(name_ar='بلوكات وأوراق لاصقة', description_ar='البلوكات والملاحظات اللاصقة بأحجام مختلفة'),
            
            # مجلدات وحفظ
            Category(name_ar='مجلدات وحافظات', description_ar='المجلدات البلاستيكية والكرتونية لحفظ الأوراق'),
            Category(name_ar='أكياس وحافظات شفافة', description_ar='الأكياس الشفافة وحافظات الأوراق البلاستيكية'),
            Category(name_ar='ملفات ومنظمات', description_ar='الملفات المعدنية والبلاستيكية ومنظمات المكتب'),
            
            # أدوات هندسية ورياضية
            Category(name_ar='مساطر وأدوات قياس', description_ar='المساطر والزوايا وأدوات القياس الهندسية'),
            Category(name_ar='برجل وكوسات هندسية', description_ar='البرجل والكوسات وأدوات الرسم الهندسي'),
            Category(name_ar='آلات حاسبة', description_ar='الآلات الحاسبة العلمية والعادية'),
            Category(name_ar='أدوات رياضية تعليمية', description_ar='النماذج الهندسية والأدوات التعليمية للرياضيات'),
            
            # أدوات فنية وإبداعية
            Category(name_ar='ألوان وطلاء', description_ar='الألوان المائية والزيتية وألوان الأطفال'),
            Category(name_ar='فرش ولوازم الرسم', description_ar='فرش الرسم وإسفنج التلوين واللوازم الفنية'),
            Category(name_ar='ورق ملون وكارتون', description_ar='الأوراق الملونة والكارتون المقوى للأعمال الفنية'),
            Category(name_ar='لاصق وصمغ', description_ar='أنواع اللاصق والصمغ والشريط اللاصق'),
            
            # أدوات مكتبية عامة
            Category(name_ar='مقصات وقطاعات', description_ar='المقصات بأحجام مختلفة وقطاعات الورق'),
            Category(name_ar='دباسة وخرامة', description_ar='الدباسات والخرامات ولوازم التثبيت'),
            Category(name_ar='مشابك ودبابيس', description_ar='مشابك الورق والدبابيس وأدوات التثبيت'),
            Category(name_ar='لوازم المكتب المختلفة', description_ar='منظمات المكتب وحوامل الأقلام والأدوات المكتبية'),
            
            # حقائب وأدوات حمل
            Category(name_ar='حقائب مدرسية', description_ar='الحقائب المدرسية بأحجام وأشكال مختلفة'),
            Category(name_ar='مقلمات وحافظات أقلام', description_ar='المقلمات وحافظات الأقلام والأدوات'),
            Category(name_ar='شنط لابتوب ووثائق', description_ar='حقائب اللابتوب وحافظات الوثائق والملفات'),
            
            # لوازم إلكترونية ومكتبية
            Category(name_ar='بطاريات وشواحن', description_ar='البطاريات والشواحن للأجهزة الإلكترونية'),
            Category(name_ar='فلاش ميموري وأقراص', description_ar='فلاش ميموري وأقراص التخزين والـ CD/DVD'),
            Category(name_ar='لوازم الكمبيوتر', description_ar='ماوس وكيبورد وإكسسوارات الكمبيوتر'),
            
            # متنوعات
            Category(name_ar='هدايا ولعب تعليمية', description_ar='الهدايا والألعاب التعليمية والترفيهية'),
            Category(name_ar='لوازم التغليف', description_ar='أكياس الهدايا وورق التغليف والشرائط'),
            Category(name_ar='منتجات موسمية', description_ar='المنتجات الخاصة بالمواسم والمناسبات المختلفة'),
        ]
        
        for category in categories:
            db.session.add(category)
    
    db.session.commit()
    
    # Create diverse sample products for different categories
    if Product.query.count() == 0:
        products = [
            # كتب أدبية وروايات
            Product(name_ar='رواية مئة عام من العزلة', category_id=1, wholesale_price=65.00, retail_price=85.00, price=85.00, stock_quantity=25, min_stock_threshold=5, description_ar='رواية للكاتب غابرييل غارسيا ماركيز'),
            Product(name_ar='ديوان محمود درويش', category_id=1, wholesale_price=35.00, retail_price=45.00, price=45.00, stock_quantity=30, min_stock_threshold=8, description_ar='مجموعة قصائد للشاعر محمود درويش'),
            Product(name_ar='رواية مدن الملح', category_id=1, wholesale_price=75.00, retail_price=95.00, price=95.00, stock_quantity=20, min_stock_threshold=5, description_ar='رواية عبد الرحمن منيف'),
            
            # كتب علمية وتقنية
            Product(name_ar='كتاب البرمجة بالبايثون', category_id=2, wholesale_price=95.00, retail_price=120.00, price=120.00, stock_quantity=15, min_stock_threshold=3, description_ar='دليل شامل لتعلم البرمجة'),
            Product(name_ar='أساسيات الرياضيات', category_id=2, wholesale_price=60.00, retail_price=80.00, price=80.00, stock_quantity=35, min_stock_threshold=10, description_ar='كتاب تعليمي في الرياضيات'),
            
            # كتب دراسية ومناهج
            Product(name_ar='منهج الرياضيات - الصف الثالث الثانوي', category_id=3, wholesale_price=42.00, retail_price=55.00, price=55.00, stock_quantity=50, min_stock_threshold=15, description_ar='منهج وزارة التربية والتعليم'),
            Product(name_ar='كتاب الفيزياء - الصف الثاني الثانوي', category_id=3, wholesale_price=36.00, retail_price=48.00, price=48.00, stock_quantity=40, min_stock_threshold=12, description_ar='منهج معتمد'),
            
            # أقلام حبر وجاف
            Product(name_ar='قلم حبر جاف أزرق', category_id=7, wholesale_price=2.50, retail_price=3.50, price=3.50, stock_quantity=200, min_stock_threshold=50, unit_type='جزئي', unit_description='قلم حبر جاف لون أزرق'),
            Product(name_ar='علبة أقلام حبر ملونة (12 قلم)', category_id=7, wholesale_price=25.00, retail_price=35.00, price=35.00, stock_quantity=80, min_stock_threshold=20, description_ar='مجموعة أقلام ملونة'),
            Product(name_ar='قلم حبر أحمر', category_id=7, wholesale_price=2.50, retail_price=3.50, price=3.50, stock_quantity=150, min_stock_threshold=40, unit_type='جزئي', unit_description='قلم'),
            Product(name_ar='قلم حبر أسود', category_id=7, wholesale_price=2.50, retail_price=3.50, price=3.50, stock_quantity=180, min_stock_threshold=45, unit_type='جزئي', unit_description='قلم'),
            
            # أقلام رصاص وملونة
            Product(name_ar='قلم رصاص HB', category_id=8, price=2.00, stock_quantity=250, min_stock_threshold=60, unit_type='جزئي', unit_description='قلم'),
            Product(name_ar='علبة أقلام ملونة خشبية (24 لون)', category_id=8, price=45.00, stock_quantity=60, min_stock_threshold=15, description_ar='أقلام تلوين خشبية عالية الجودة'),
            Product(name_ar='قلم رصاص 2B للرسم', category_id=8, price=4.00, stock_quantity=100, min_stock_threshold=25, unit_type='جزئي', unit_description='قلم'),
            
            # كشاكيل ودفاتر
            Product(name_ar='كشكول 100 ورقة مخطط', category_id=11, price=15.00, stock_quantity=120, min_stock_threshold=30, description_ar='كشكول مخطط للكتابة'),
            Product(name_ar='كشكول 200 ورقة مربعات', category_id=11, price=25.00, stock_quantity=90, min_stock_threshold=25, description_ar='كشكول مربعات للرياضيات'),
            Product(name_ar='دفتر 48 ورقة سادة', category_id=12, price=8.00, stock_quantity=200, min_stock_threshold=50, description_ar='دفتر سادة للكتابة الحرة'),
            Product(name_ar='كشكول سبايرال A4', category_id=11, price=28.00, stock_quantity=75, min_stock_threshold=20, description_ar='كشكول سبايرال حجم A4'),
            
            # محايات وبراية
            Product(name_ar='محاية بيضاء كبيرة', category_id=10, price=2.50, stock_quantity=300, min_stock_threshold=75, unit_type='جزئي', unit_description='قطعة'),
            Product(name_ar='براية معدنية', category_id=10, price=5.00, stock_quantity=150, min_stock_threshold=35, unit_type='جزئي', unit_description='قطعة'),
            Product(name_ar='محاية ملونة صغيرة', category_id=10, price=1.50, stock_quantity=400, min_stock_threshold=100, unit_type='جزئي', unit_description='قطعة'),
            
            # مجلدات وحافظات
            Product(name_ar='مجلد بلاستيكي A4', category_id=15, price=12.00, stock_quantity=80, min_stock_threshold=20, description_ar='مجلد شفاف لحفظ الأوراق'),
            Product(name_ar='حافظة أوراق شفافة (10 قطع)', category_id=16, price=8.00, stock_quantity=100, min_stock_threshold=25, description_ar='حافظات شفافة مثقبة'),
            Product(name_ar='مجلد كرتوني ملون', category_id=15, price=18.00, stock_quantity=60, min_stock_threshold=15, description_ar='مجلد كرتوني بألوان مختلفة'),
            
            # أدوات هندسية
            Product(name_ar='مسطرة 30 سم شفافة', category_id=18, price=8.00, stock_quantity=120, min_stock_threshold=30, unit_type='جزئي', unit_description='قطعة'),
            Product(name_ar='مجموعة أدوات هندسية (برجل + مسطرة + زاوية)', category_id=19, price=35.00, stock_quantity=40, min_stock_threshold=10, description_ar='مجموعة كاملة للرسم الهندسي'),
            Product(name_ar='آلة حاسبة علمية', category_id=20, price=85.00, stock_quantity=25, min_stock_threshold=5, description_ar='آلة حاسبة للطلاب والمهندسين'),
            
            # ألوان وطلاء
            Product(name_ar='علبة ألوان مائية (12 لون)', category_id=22, price=25.00, stock_quantity=50, min_stock_threshold=12, description_ar='ألوان مائية للرسم والفن'),
            Product(name_ar='ألوان فلوماستر (18 لون)', category_id=22, price=30.00, stock_quantity=70, min_stock_threshold=18, description_ar='أقلام ألوان فلوماستر'),
            
            # أدوات مكتبية
            Product(name_ar='مقص متوسط الحجم', category_id=26, price=12.00, stock_quantity=90, min_stock_threshold=20, unit_type='جزئي', unit_description='قطعة'),
            Product(name_ar='دباسة صغيرة + علبة دبابيس', category_id=27, price=15.00, stock_quantity=60, min_stock_threshold=15, description_ar='دباسة مع دبابيس للاستعمال المكتبي'),
            Product(name_ar='صمغ أبيض 50 مل', category_id=25, price=6.00, stock_quantity=150, min_stock_threshold=35, unit_type='جزئي', unit_description='أنبوبة'),
            
            # حقائب ومقلمات
            Product(name_ar='حقيبة مدرسية متوسطة', category_id=30, price=95.00, stock_quantity=30, min_stock_threshold=8, description_ar='حقيبة مدرسية بجيوب متعددة'),
            Product(name_ar='مقلمة بسحاب', category_id=31, price=18.00, stock_quantity=85, min_stock_threshold=20, description_ar='مقلمة لحفظ الأقلام والأدوات'),
            Product(name_ar='حقيبة لابتوب 15 بوصة', category_id=32, price=150.00, stock_quantity=20, min_stock_threshold=5, description_ar='حقيبة واقية للابتوب'),
            
            # منتجات إلكترونية
            Product(name_ar='فلاش ميموري 16 جيجا', category_id=34, price=45.00, stock_quantity=40, min_stock_threshold=10, unit_type='جزئي', unit_description='قطعة'),
            Product(name_ar='بطاريات AA (4 قطع)', category_id=33, price=12.00, stock_quantity=100, min_stock_threshold=25, description_ar='بطاريات قلوية عالية الجودة'),
            
            # هدايا ومتنوعات
            Product(name_ar='لعبة تعليمية للأطفال', category_id=36, price=35.00, stock_quantity=25, min_stock_threshold=5, description_ar='لعبة تعليمية تفاعلية'),
            Product(name_ar='كيس هدية ملون', category_id=37, price=3.00, stock_quantity=200, min_stock_threshold=50, unit_type='جزئي', unit_description='كيس'),
            
            # مجلات
            Product(name_ar='مجلة العلوم العدد الجديد', category_id=6, price=15.00, stock_quantity=35, min_stock_threshold=10, description_ar='مجلة شهرية علمية'),
            Product(name_ar='مجلة الأطفال المصورة', category_id=6, price=12.00, stock_quantity=50, min_stock_threshold=15, description_ar='مجلة أسبوعية للأطفال'),
        ]
        
        for product in products:
            db.session.add(product)
    
    db.session.commit()
    
    # Create sample customers
    if Customer.query.count() == 0:
        customers = [
            Customer(name='أحمد محمد علي', phone='01234567890', address='القاهرة - مصر الجديدة', notes='عميل دائم'),
            Customer(name='فاطمة أحمد حسن', phone='01098765432', address='الجيزة - الدقي'),
            Customer(name='محمد حسن إبراهيم', phone='01555666777', address='الإسكندرية - سيدي جابر', notes='عميل مميز'),
            Customer(name='سارة علي محمود', phone='01122334455', address='القاهرة - المعادي'),
            Customer(name='عمر خالد أحمد', phone='01199887766', address='الجيزة - المهندسين', notes='عميل جديد'),
        ]
        
        for customer in customers:
            db.session.add(customer)
        
        db.session.commit()

@app.route('/simple-export-test')
@login_required
def simple_export_test():
    """صفحة اختبار بسيطة للتصدير"""
    return """
    <!DOCTYPE html>
    <html>
    <head>
        <title>اختبار التصدير</title>
        <meta charset="UTF-8">
    </head>
    <body>
        <h1>اختبار وظيفة التصدير</h1>
        <button onclick="testExport()">اختبار تصدير Excel</button>
        <div id="result"></div>
        
        <script src="https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js"></script>
        <script>
            function testExport() {
                console.log('Testing export...');
                document.getElementById('result').innerHTML = 'جاري الاختبار...';
                
                if (typeof XLSX === 'undefined') {
                    document.getElementById('result').innerHTML = 'خطأ: مكتبة Excel غير محملة';
                    return;
                }
                
                try {
                    const testData = [
                        ['اختبار التصدير'],
                        ['المنتج', 'الكمية', 'السعر'],
                        ['منتج تجريبي', 10, 100]
                    ];
                    
                    const wb = XLSX.utils.book_new();
                    const ws = XLSX.utils.aoa_to_sheet(testData);
                    XLSX.utils.book_append_sheet(wb, ws, 'اختبار');
                    
                    const fileName = 'test_export_' + new Date().getTime() + '.xlsx';
                    XLSX.writeFile(wb, fileName);
                    
                    document.getElementById('result').innerHTML = 'نجح التصدير! ✅';
                } catch (error) {
                    document.getElementById('result').innerHTML = 'فشل التصدير: ' + error.message;
                }
            }
        </script>
    </body>
    </html>
    """

@app.route('/debug-export')
def debug_export():
    """صفحة تشخيص مشاكل التصدير - بدون تسجيل دخول للتشخيص"""
    return '''
    <!DOCTYPE html>
    <html lang="ar" dir="rtl">
    <head>
        <meta charset="UTF-8">
        <title>تشخيص مشكلة التصدير</title>
        <style>
            body { font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }
            .container { max-width: 800px; margin: 0 auto; }
            .step { background: white; padding: 20px; margin: 15px 0; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
            .success { border-left: 5px solid #28a745; }
            .error { border-left: 5px solid #dc3545; }
            .warning { border-left: 5px solid #ffc107; }
            button { 
                background: #007bff; color: white; border: none; 
                padding: 12px 24px; margin: 8px; 
                border-radius: 5px; cursor: pointer; font-size: 16px;
            }
            button:hover { background: #0056b3; }
            pre { background: #f8f9fa; padding: 15px; border-radius: 5px; overflow-x: auto; max-height: 300px; }
            .result { margin-top: 15px; padding: 10px; border-radius: 5px; }
            h1 { color: #333; text-align: center; }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>🔍 تشخيص مشكلة التصدير</h1>
            <p style="text-align: center; color: #666;">هذه الأداة ستساعدك في تحديد سبب عدم عمل التصدير</p>
            
            <div class="step">
                <h3>الخطوة 1: فحص المكتبات المطلوبة</h3>
                <button onclick="checkLibraries()">فحص المكتبات</button>
                <div id="libraries-result" class="result"></div>
            </div>
            
            <div class="step">
                <h3>الخطوة 2: اختبار تصدير Excel بسيط</h3>
                <button onclick="simpleExcelTest()">اختبار Excel</button>
                <div id="excel-result" class="result"></div>
            </div>
            
            <div class="step">
                <h3>الخطوة 3: اختبار تحميل ملف نصي</h3>
                <button onclick="simpleDownloadTest()">اختبار التحميل</button>
                <div id="download-result" class="result"></div>
            </div>
            
            <div class="step">
                <h3>الخطوة 4: معلومات المتصفح</h3>
                <button onclick="browserInfo()">معلومات المتصفح</button>
                <div id="browser-result" class="result"></div>
            </div>
            
            <div class="step">
                <h3>سجل العمليات والأخطاء</h3>
                <pre id="debug-log">انتظار بدء التشخيص...</pre>
                <button onclick="clearLog()">مسح السجل</button>
            </div>
        </div>

        <script src="https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js"></script>
        
        <script>
            let logs = [];
            
            function addLog(message) {
                const time = new Date().toLocaleTimeString();
                const logEntry = `[${time}] ${message}`;
                logs.push(logEntry);
                document.getElementById("debug-log").textContent = logs.join("\\n");
                console.log(logEntry);
            }
            
            function clearLog() {
                logs = [];
                document.getElementById("debug-log").textContent = "تم مسح السجل...";
            }
            
            function setResult(elementId, content, type = "info") {
                const element = document.getElementById(elementId);
                element.innerHTML = content;
                element.className = `result ${type}`;
            }
            
            function checkLibraries() {
                addLog("🔍 بدء فحص المكتبات...");
                
                let results = "<h4>نتائج الفحص:</h4>";
                let allOk = true;
                
                if (typeof XLSX !== "undefined") {
                    results += "✅ مكتبة XLSX: محملة بنجاح<br>";
                    results += `📋 الإصدار: ${XLSX.version || "غير محدد"}<br>`;
                    addLog("✅ XLSX library loaded");
                } else {
                    results += "❌ مكتبة XLSX: غير محملة<br>";
                    allOk = false;
                    addLog("❌ XLSX library missing");
                }
                
                if (typeof Blob !== "undefined") {
                    results += "✅ Blob API: متوفر<br>";
                } else {
                    results += "❌ Blob API: غير متوفر<br>";
                    allOk = false;
                }
                
                if (typeof URL.createObjectURL === "function") {
                    results += "✅ URL API: متوفر<br>";
                } else {
                    results += "❌ URL API: غير متوفر<br>";
                    allOk = false;
                }
                
                const downloadSupported = "download" in document.createElement("a");
                if (downloadSupported) {
                    results += "✅ Download Attribute: مدعوم<br>";
                } else {
                    results += "❌ Download Attribute: غير مدعوم<br>";
                    allOk = false;
                }
                
                setResult("libraries-result", results, allOk ? "success" : "error");
                addLog(`📊 فحص المكتبات: ${allOk ? "نجح" : "فشل"}`);
            }
            
            function simpleExcelTest() {
                addLog("📊 بدء اختبار Excel...");
                
                if (typeof XLSX === "undefined") {
                    setResult("excel-result", "❌ لا يمكن الاختبار - مكتبة XLSX غير محملة", "error");
                    addLog("❌ Excel test failed - no XLSX");
                    return;
                }
                
                try {
                    addLog("📝 إنشاء بيانات اختبار...");
                    const data = [
                        ["تجربة التصدير"],
                        ["العنصر", "القيمة", "التاريخ"],
                        ["اختبار 1", 100, new Date().toLocaleDateString()],
                        ["اختبار 2", 200, new Date().toLocaleDateString()],
                        ["المجموع", 300, ""]
                    ];
                    
                    addLog("🔧 إنشاء ملف Excel...");
                    const wb = XLSX.utils.book_new();
                    const ws = XLSX.utils.aoa_to_sheet(data);
                    XLSX.utils.book_append_sheet(wb, ws, "اختبار");
                    
                    addLog("💾 محاولة حفظ الملف...");
                    const fileName = `excel_test_${Date.now()}.xlsx`;
                    
                    XLSX.writeFile(wb, fileName);
                    
                    setResult("excel-result", "✅ تم إنشاء ملف Excel بنجاح! تحقق من مجلد التحميل.", "success");
                    addLog("✅ Excel file created successfully");
                    
                } catch (error) {
                    setResult("excel-result", `❌ خطأ في إنشاء Excel: ${error.message}`, "error");
                    addLog(`❌ Excel error: ${error.message}`);
                    console.error("Excel Error Details:", error);
                }
            }
            
            function simpleDownloadTest() {
                addLog("⬇️ بدء اختبار التحميل...");
                
                try {
                    const content = "اختبار التحميل\\nهذا ملف نصي تجريبي\\nالوقت: " + new Date().toLocaleString();
                    const blob = new Blob([content], { type: "text/plain;charset=utf-8" });
                    const url = URL.createObjectURL(blob);
                    
                    const link = document.createElement("a");
                    link.href = url;
                    link.download = `download_test_${Date.now()}.txt`;
                    link.style.display = "none";
                    
                    document.body.appendChild(link);
                    link.click();
                    document.body.removeChild(link);
                    
                    URL.revokeObjectURL(url);
                    
                    setResult("download-result", "✅ تم اختبار التحميل! تحقق من مجلد التحميل.", "success");
                    addLog("✅ Download test completed");
                    
                } catch (error) {
                    setResult("download-result", `❌ خطأ في التحميل: ${error.message}`, "error");
                    addLog(`❌ Download error: ${error.message}`);
                }
            }
            
            function browserInfo() {
                addLog("🌐 جمع معلومات المتصفح...");
                
                let info = "<h4>معلومات المتصفح:</h4>";
                info += `<strong>المتصفح:</strong> ${navigator.userAgent}<br><br>`;
                info += `<strong>النظام:</strong> ${navigator.platform}<br>`;
                info += `<strong>اللغة:</strong> ${navigator.language}<br>`;
                info += `<strong>Cookies مفعلة:</strong> ${navigator.cookieEnabled ? "نعم" : "لا"}<br>`;
                
                info += "<br><h5>🔧 نصائح لحل المشكلة:</h5>";
                info += "• تأكد أن مجلد التحميل محدد بشكل صحيح<br>";
                info += "• تحقق من إعدادات حظر النوافذ المنبثقة<br>";
                info += "• جرب متصفح آخر للمقارنة<br>";
                info += "• تأكد أن الإنترنت متصل لتحميل المكتبات<br>";
                info += "• اضغط F12 وافحص تبويب Console للأخطاء<br>";
                
                setResult("browser-result", info, "info");
                addLog("📋 Browser info collected");
            }
            
            window.addEventListener("load", function() {
                addLog("🚀 تم تحميل صفحة التشخيص");
                setTimeout(() => {
                    addLog("🔄 بدء الفحص التلقائي...");
                    checkLibraries();
                }, 500);
            });
        </script>
    </body>
    </html>
    '''

@app.route('/api/export/full-database')
@login_required
def api_export_full_database():
    if current_user.role not in ['admin', 'seller']:
        abort(403)
    
    try:
        data = {}
        
        # Basic report summary
        data['report_summary'] = [{
            'تاريخ التقرير': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'عدد المنتجات': Product.query.count(),
            'عدد العملاء': Customer.query.count(),
            'عدد المبيعات': Sale.query.count(),
            'عدد الفئات': Category.query.count()
        }]
        
        # Products - basic info only
        products_data = []
        for p in Product.query.all():
            products_data.append({
                'رقم المنتج': p.id,
                'اسم المنتج': p.name_ar,
                'سعر الجملة': float(p.wholesale_price),
                'سعر البيع': float(p.retail_price),
                'الكمية': float(p.stock_quantity),
                'نوع الوحدة': p.unit_type,
                'الحد الأدنى': float(p.min_stock_threshold)
            })
        data['products'] = products_data
        
        # Categories
        categories_data = []
        for c in Category.query.all():
            categories_data.append({
                'رقم الفئة': c.id,
                'اسم الفئة': c.name_ar,
                'الوصف': c.description_ar or ''
            })
        data['categories'] = categories_data
        
        # Customers
        customers_data = []
        for c in Customer.query.all():
            customers_data.append({
                'رقم العميل': c.id,
                'اسم العميل': c.name,
                'الهاتف': c.phone or '',
                'العنوان': c.address or '',
                'ملاحظات': c.notes or ''
            })
        data['customers'] = customers_data
        
        # Sales - basic info only
        sales_data = []
        for s in Sale.query.all():
            sales_data.append({
                'رقم البيع': s.id,
                'تاريخ البيع': str(s.sale_date),
                'المبلغ': float(s.total_amount),
                'نوع الدفع': s.payment_type,
                'حالة الدفع': s.payment_status,
                'ملاحظات': s.notes or ''
            })
        data['sales'] = sales_data
        
        # Sale Items
        sale_items_data = []
        for si in SaleItem.query.all():
            sale_items_data.append({
                'رقم البيع': si.sale_id,
                'رقم المنتج': si.product_id,
                'الكمية': float(si.quantity),
                'سعر الوحدة': float(si.unit_price),
                'الإجمالي': float(si.total_price)
            })
        data['sale_items'] = sale_items_data
        
        # Payments
        payments_data = []
        for p in Payment.query.all():
            payments_data.append({
                'رقم الدفعة': p.id,
                'رقم البيع': p.sale_id,
                'المبلغ': float(p.amount),
                'تاريخ الدفع': str(p.payment_date),
                'طريقة الدفع': p.payment_method,
                'ملاحظات': p.notes or ''
            })
        data['payments'] = payments_data
        
        # Expenses
        expenses_data = []
        for e in Expense.query.all():
            expenses_data.append({
                'رقم المصروف': e.id,
                'الوصف': e.description,
                'المبلغ': float(e.amount),
                'النوع': e.expense_type,
                'التاريخ': str(e.expense_date),
                'ملاحظات': e.notes or ''
            })
        data['expenses'] = expenses_data
        
        # Users (admin only)
        if current_user.role == 'admin':
            users_data = []
            for u in User.query.all():
                users_data.append({
                    'رقم المستخدم': u.id,
                    'اسم المستخدم': u.username,
                    'الدور': u.role,
                    'تاريخ الإنشاء': str(u.created_at)
                })
            data['users'] = users_data
        
        return jsonify(data)
        
    except Exception as e:
        # Log the actual error for debugging
        print(f"Export error: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return jsonify({
            'error': 'حدث خطأ في تصدير البيانات',
            'message': str(e)
        }), 500

@app.route('/api/export/test-database')
@login_required
def api_test_database_export():
    """Simple test endpoint for database export"""
    try:
        data = {
            'status': 'success',
            'message': 'Test endpoint working',
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'products_count': Product.query.count(),
            'customers_count': Customer.query.count(),
            'sales_count': Sale.query.count()
        }
        return jsonify(data)
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

# ================================
# Routes للمرتجعات
# ================================

@app.route('/returns')
@login_required
@seller_or_admin_required
def returns():
    """عرض قائمة المرتجعات"""
    page = request.args.get('page', 1, type=int)
    status_filter = request.args.get('status', 'all')
    
    query = Return.query.join(Sale).join(User)
    
    # تطبيق فلتر الحالة
    if status_filter != 'all':
        query = query.filter(Return.status == status_filter)
    
    # فلترة حسب صلاحية المستخدم
    if not current_user.is_admin():
        query = query.filter(Return.user_id == current_user.id)
    
    returns = query.order_by(Return.return_date.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    
    return render_template('returns/list.html', returns=returns, status_filter=status_filter)

@app.route('/returns/new/<int:sale_id>')
@login_required
@seller_or_admin_required
def new_return(sale_id):
    """إنشاء مرتجع جديد لبيعة معينة"""
    sale = Sale.query.get_or_404(sale_id)
    
    # التحقق من الصلاحية
    if not current_user.is_admin() and sale.user_id != current_user.id:
        flash('ليس لديك صلاحية لإنشاء مرتجع لهذا البيع', 'error')
        return redirect(url_for('sales'))
    
    return render_template('returns/new.html', sale=sale)

@app.route('/api/returns', methods=['POST'])
@login_required
@seller_or_admin_required
def api_create_return():
    """API لإنشاء مرتجع جديد"""
    try:
        data = request.get_json()
        sale_id = data.get('sale_id')
        reason = data.get('reason', '')
        refund_method = data.get('refund_method', 'نقدي')
        notes = data.get('notes', '')
        items = data.get('items', [])
        
        if not sale_id or not reason or not items:
            return jsonify({'success': False, 'message': 'بيانات غير مكتملة'}), 400
        
        # التحقق من البيع
        sale = Sale.query.get_or_404(sale_id)
        if not current_user.is_admin() and sale.user_id != current_user.id:
            return jsonify({'success': False, 'message': 'ليس لديك صلاحية لإنشاء مرتجع لهذا البيع'}), 403
        
        # إنشاء المرتجع
        return_obj = Return(
            sale_id=sale_id,
            customer_id=sale.customer_id,
            reason=reason,
            refund_method=refund_method,
            notes=notes,
            user_id=current_user.id,
            total_amount=0  # سيتم حسابه لاحقاً
        )
        
        db.session.add(return_obj)
        db.session.flush()  # للحصول على ID المرتجع
        
        total_amount = 0
        
        # إضافة أصناف المرتجع
        for item_data in items:
            sale_item_id = item_data.get('sale_item_id')
            quantity_returned = float(item_data.get('quantity_returned', 0))
            condition = item_data.get('condition', 'جيد')
            item_notes = item_data.get('notes', '')
            
            # التحقق من صنف البيع
            sale_item = SaleItem.query.get_or_404(sale_item_id)
            if sale_item.sale_id != sale_id:
                return jsonify({'success': False, 'message': 'خطأ في بيانات الصنف'}), 400
            
            # التحقق من الكمية
            if quantity_returned <= 0 or quantity_returned > sale_item.quantity:
                return jsonify({'success': False, 'message': f'كمية خاطئة للصنف {sale_item.product.name_ar}'}), 400
            
            # إنشاء صنف المرتجع
            return_item = ReturnItem(
                return_id=return_obj.id,
                sale_item_id=sale_item_id,
                product_id=sale_item.product_id,
                quantity_returned=quantity_returned,
                original_quantity=sale_item.quantity,
                unit_price=sale_item.unit_price,
                condition=condition,
                notes=item_notes
            )
            
            db.session.add(return_item)
            total_amount += return_item.total_refund
        
        # تحديث إجمالي المرتجع
        return_obj.total_amount = total_amount
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'تم إنشاء المرتجع بنجاح',
            'return_id': return_obj.id
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/returns/<int:return_id>')
@login_required
def api_return_details(return_id):
    """API لعرض تفاصيل المرتجع"""
    return_obj = Return.query.get_or_404(return_id)
    
    # التحقق من الصلاحية
    if not current_user.is_admin() and return_obj.user_id != current_user.id:
        return jsonify({'error': 'ليس لديك صلاحية لعرض هذا المرتجع'}), 403
    
    # تفاصيل أصناف المرتجع
    items = []
    for item in return_obj.return_items:
        items.append({
            'product_name': item.product.name_ar,
            'quantity_returned': float(item.quantity_returned),
            'original_quantity': float(item.original_quantity),
            'unit_price': float(item.unit_price),
            'total_refund': float(item.total_refund),
            'condition': item.condition,
            'condition_ar': item.condition_ar,
            'notes': item.notes or ''
        })
    
    # تحويل التاريخ إلى توقيت مصر
    egypt_time = get_egypt_time(return_obj.return_date)
    
    return jsonify({
        'id': return_obj.id,
        'sale_id': return_obj.sale_id,
        'customer_name': return_obj.customer.name if return_obj.customer else 'زبون نقدي',
        'total_amount': float(return_obj.total_amount),
        'return_date': egypt_time.strftime('%d/%m/%Y'),
        'return_time': egypt_time.strftime('%I:%M:%S %p').replace('AM', 'ص').replace('PM', 'م'),
        'reason': return_obj.reason,
        'status': return_obj.status,
        'status_ar': return_obj.status_ar,
        'refund_method': return_obj.refund_method,
        'notes': return_obj.notes or '',
        'user_name': return_obj.user.username,
        'processor_name': return_obj.processor.username if return_obj.processor else '',
        'processed_date': return_obj.processed_date.strftime('%d/%m/%Y') if return_obj.processed_date else '',
        'items': items
    })

@app.route('/api/returns/<int:return_id>/process', methods=['POST'])
@login_required
@admin_required
def api_process_return(return_id):
    """API لمعالجة المرتجع (قبول/رفض)"""
    try:
        data = request.get_json()
        action = data.get('action')  # 'approve' or 'reject'
        notes = data.get('notes', '')
        
        if action not in ['approve', 'reject']:
            return jsonify({'success': False, 'message': 'إجراء غير صحيح'}), 400
        
        return_obj = Return.query.get_or_404(return_id)
        
        if not return_obj.can_be_processed:
            return jsonify({'success': False, 'message': 'لا يمكن معالجة هذا المرتجع'}), 400
        
        # تحديث حالة المرتجع
        return_obj.status = 'approved' if action == 'approve' else 'rejected'
        return_obj.processed_by = current_user.id
        return_obj.processed_date = datetime.utcnow()
        if notes:
            return_obj.notes = (return_obj.notes or '') + f'\n\nملاحظات المعالجة: {notes}'
        
        # إذا تم قبول المرتجع، تحديث المخزون
        if action == 'approve':
            for return_item in return_obj.return_items:
                # إضافة الكمية المرتجعة إلى المخزون إذا كانت في حالة جيدة
                if return_item.condition in ['جيد', 'good']:
                    product = return_item.product
                    product.stock_quantity += return_item.quantity_returned
        
        db.session.commit()
        
        status_message = 'تم قبول المرتجع وإضافة الكمية للمخزون' if action == 'approve' else 'تم رفض المرتجع'
        
        return jsonify({
            'success': True,
            'message': status_message
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/sale/<int:sale_id>/items')
@login_required
@seller_or_admin_required
def api_sale_items(sale_id):
    """API لعرض أصناف البيع للمرتجع"""
    sale = Sale.query.get_or_404(sale_id)
    
    # التحقق من الصلاحية
    if not current_user.is_admin() and sale.user_id != current_user.id:
        return jsonify({'error': 'ليس لديك صلاحية لعرض هذا البيع'}), 403
    
    items = []
    for item in sale.sale_items:
        # حساب الكمية المرتجعة سابقاً
        returned_quantity = sum(
            ri.quantity_returned for ri in item.return_items 
            if ri.return_ref.status == 'approved'
        )
        
        available_quantity = item.quantity - returned_quantity
        
        if available_quantity > 0:  # فقط الأصناف التي يمكن إرجاعها
            items.append({
                'id': item.id,
                'product_name': item.product.name_ar,
                'quantity': float(item.quantity),
                'returned_quantity': float(returned_quantity),
                'available_quantity': float(available_quantity),
                'unit_price': float(item.unit_price),
                'total_price': float(item.total_price),
                'unit_type': item.product.unit_type
            })
    
    return jsonify(items)

@app.route('/api/products/excel-template')
@login_required
@admin_required
def api_products_excel_template():
    """تحميل نموذج Excel للمنتجات"""
    try:
        # إنشاء workbook جديد
        wb = Workbook()
        ws = wb.active
        ws.title = "المنتجات"
        
        # العناوين
        headers = [
            'اسم المنتج', 'وصف المنتج', 'الفئة', 'سعر الجملة', 
            'سعر البيع', 'الكمية', 'الحد الأدنى للمخزون', 'نوع الوحدة', 'وصف الوحدة'
        ]
        
        # إضافة العناوين
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # إضافة بيانات تجريبية
        sample_data = [
            ['منتج تجريبي 1', 'وصف المنتج الأول', 'قرطاسية', 10.50, 15.00, 100, 10, 'كامل', 'قطعة'],
            ['منتج تجريبي 2', 'وصف المنتج الثاني', 'كتب', 25.00, 35.00, 50, 5, 'كامل', 'كتاب']
        ]
        
        for row_num, row_data in enumerate(sample_data, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # ضبط عرض الأعمدة
        column_widths = [20, 25, 15, 12, 12, 10, 18, 12, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
        
        # حفظ في الذاكرة
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=products_template.xlsx'
        
        return response
        
    except Exception as e:
        app.logger.error(f"Error creating Excel template: {str(e)}")
        return jsonify({'error': 'حدث خطأ في إنشاء النموذج'}), 500

@app.route('/api/products/import-excel', methods=['POST'])
@login_required
@admin_required
def api_products_import_excel():
    """استيراد المنتجات من ملف Excel"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'لم يتم العثور على ملف'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'لم يتم اختيار ملف'}), 400
        
        # التحقق من نوع الملف
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': 'نوع الملف غير مدعوم. يرجى استخدام ملف Excel'}), 400
        
        # التحقق من حجم الملف (5 ميجابايت)
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)
        
        if file_size > 5 * 1024 * 1024:  # 5MB
            return jsonify({'success': False, 'message': 'حجم الملف كبير جداً. الحد الأقصى 5 ميجابايت'}), 400
        
        update_existing = request.form.get('update_existing') == 'true'
        
        # قراءة ملف Excel باستخدام openpyxl
        try:
            wb = load_workbook(file)
            ws = wb.active
        except Exception as e:
            return jsonify({'success': False, 'message': f'خطأ في قراءة ملف Excel: {str(e)}'}), 400
        
        # قراءة العناوين من الصف الأول
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                headers.append('')
        
        # التحقق من الأعمدة المطلوبة
        required_columns = ['اسم المنتج', 'الفئة', 'سعر الجملة', 'سعر البيع', 'الكمية']
        missing_columns = [col for col in required_columns if col not in headers]
        
        if missing_columns:
            return jsonify({
                'success': False, 
                'message': f'أعمدة مفقودة: {", ".join(missing_columns)}'
            }), 400
        
        # إنشاء فهرس للأعمدة
        column_index = {}
        for i, header in enumerate(headers):
            if header:
                column_index[header] = i
        
        # إحصائيات الاستيراد
        added_count = 0
        updated_count = 0
        skipped_count = 0
        errors = []
        
        # تعريف دالة مساعدة خارج الحلقة
        def get_cell_value(row, column_name, default=''):
            if column_name in column_index and column_index[column_name] < len(row):
                value = row[column_index[column_name]]
                if value is None:
                    return default
                value_str = str(value).strip()
                return value_str if value_str.lower() not in ['none', 'nan', ''] else default
            return default
        
        # إضافة logging للتشخيص
        app.logger.info(f"Starting import process. Total rows to process: {ws.max_row - 1}")
        
        # معالجة كل صف (بداية من الصف الثاني)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                # تجاهل الصفوف الفارغة
                if not any(row) or all(v is None or str(v).strip() == '' for v in row):
                    app.logger.debug(f"Skipping empty row {row_num}")
                    continue
                
                product_name = get_cell_value(row, 'اسم المنتج')
                category_name = get_cell_value(row, 'الفئة')
                
                app.logger.debug(f"Processing row {row_num}: product='{product_name}', category='{category_name}'")
                
                # التحقق من البيانات الأساسية
                if not product_name or product_name.lower() in ['none', 'nan', '']:
                    errors.append(f'الصف {row_num}: اسم المنتج فارغ أو غير صحيح')
                    skipped_count += 1
                    continue
                
                try:
                    wholesale_price_str = get_cell_value(row, 'سعر الجملة', '0')
                    retail_price_str = get_cell_value(row, 'سعر البيع', '0')
                    stock_quantity_str = get_cell_value(row, 'الكمية', '0')
                    
                    wholesale_price = float(wholesale_price_str)
                    retail_price = float(retail_price_str)
                    stock_quantity = float(stock_quantity_str)
                except (ValueError, TypeError) as e:
                    errors.append(f'الصف {row_num}: خطأ في تحويل الأرقام - {str(e)}')
                    skipped_count += 1
                    continue
                
                if wholesale_price <= 0 or retail_price <= 0:
                    errors.append(f'الصف {row_num}: أسعار غير صحيحة')
                    skipped_count += 1
                    continue
                
                if retail_price <= wholesale_price:
                    errors.append(f'الصف {row_num}: سعر البيع يجب أن يكون أكبر من سعر الجملة')
                    skipped_count += 1
                    continue
                
                # البحث عن الفئة
                category = Category.query.filter_by(name_ar=category_name).first()
                if not category:
                    errors.append(f'الصف {row_num}: الفئة "{category_name}" غير موجودة')
                    skipped_count += 1
                    continue
                
                # التحقق من وجود المنتج
                existing_product = Product.query.filter_by(name_ar=product_name).first()
                
                if existing_product and not update_existing:
                    skipped_count += 1
                    continue
                
                # البيانات الاختيارية
                description = get_cell_value(row, 'وصف المنتج')
                
                try:
                    min_stock = float(get_cell_value(row, 'الحد الأدنى للمخزون', '10'))
                except (ValueError, TypeError):
                    min_stock = 10
                
                unit_type = get_cell_value(row, 'نوع الوحدة', 'كامل')
                if unit_type not in ['كامل', 'جزئي']:
                    unit_type = 'كامل'
                
                unit_description = get_cell_value(row, 'وصف الوحدة')
                
                if existing_product and update_existing:
                    # تحديث المنتج الموجود
                    app.logger.info(f"Updating existing product: {product_name}")
                    existing_product.description_ar = description
                    existing_product.category_id = category.id
                    existing_product.wholesale_price = wholesale_price
                    existing_product.retail_price = retail_price
                    existing_product.price = retail_price
                    existing_product.stock_quantity = stock_quantity
                    existing_product.min_stock_threshold = min_stock
                    existing_product.unit_type = unit_type
                    existing_product.unit_description = unit_description
                    existing_product.updated_at = datetime.utcnow()
                    updated_count += 1
                else:
                    # إضافة منتج جديد
                    app.logger.info(f"Adding new product: {product_name}")
                    new_product = Product(
                        name_ar=product_name,
                        description_ar=description,
                        category_id=category.id,
                        wholesale_price=wholesale_price,
                        retail_price=retail_price,
                        price=retail_price,
                        stock_quantity=stock_quantity,
                        min_stock_threshold=min_stock,
                        unit_type=unit_type,
                        unit_description=unit_description
                    )
                    db.session.add(new_product)
                    added_count += 1
                    
                app.logger.info(f"Row {row_num} processed successfully")
                
            except Exception as e:
                app.logger.error(f'Error processing row {row_num}: {str(e)}')
                errors.append(f'الصف {row_num}: {str(e)}')
                skipped_count += 1
                continue
        
        # حفظ التغييرات
        try:
            app.logger.info(f"Import summary: added={added_count}, updated={updated_count}, skipped={skipped_count}, errors={len(errors)}")
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'تم استيراد المنتجات بنجاح',
                'added_count': added_count,
                'updated_count': updated_count,
                'skipped_count': skipped_count,
                'errors': errors[:10],  # أول 10 أخطاء فقط
                'total_rows_processed': added_count + updated_count + skipped_count
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({
                'success': False,
                'message': f'خطأ في حفظ البيانات: {str(e)}'
            }), 500
            
    except Exception as e:
        app.logger.error(f"Error importing Excel: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'خطأ في معالجة الملف: {str(e)}'
        }), 500

@app.route('/api/products/debug-excel', methods=['POST'])
@login_required
@admin_required
def api_debug_excel():
    """اختبار قراءة ملف Excel لتشخيص المشاكل"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'لم يتم العثور على ملف'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'لم يتم اختيار ملف'}), 400
        
        # قراءة ملف Excel
        wb = load_workbook(file)
        ws = wb.active
        
        # قراءة العناوين
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                headers.append('')
        
        # قراءة أول 5 صفوف للاختبار
        rows_data = []
        row_count = 0
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if row_count >= 5:  # اقرأ أول 5 صفوف فقط
                break
            
            row_data = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    value = row[i]
                    row_data[header] = str(value) if value is not None else 'None'
            
            rows_data.append({
                'row_number': row_num,
                'data': row_data,
                'raw_values': [str(v) if v is not None else 'None' for v in row[:len(headers)]]
            })
            row_count += 1
        
        return jsonify({
            'total_rows': ws.max_row,
            'headers': headers,
            'sample_rows': rows_data,
            'worksheet_name': ws.title
        })
        
    except Exception as e:
        return jsonify({'error': f'خطأ في قراءة الملف: {str(e)}'}), 500

@app.route('/qr-generator')
@login_required
def qr_generator():
    """صفحة توليد QR codes للينكات ومواقع التواصل الاجتماعي"""
    return render_template('qr_generator.html')

@app.route('/price-ticket')
@login_required
@seller_or_admin_required
def price_ticket():
    """صفحة إنتاج تيكت الأسعار"""
    products = Product.query.order_by(Product.name_ar).all()
    return render_template('price_ticket.html', products=products)

@app.route('/api/search-products')
@login_required
@seller_or_admin_required
def api_search_products():
    """البحث عن المنتجات لتيكت الأسعار"""
    query = request.args.get('q', '').strip()
    products = []
    
    if query:
        products = Product.query.filter(
            Product.name_ar.contains(query)
        ).order_by(Product.name_ar).limit(10).all()
    
    return jsonify([{
        'id': product.id,
        'name': product.name_ar,
        'retail_price': product.retail_price,
        'wholesale_price': product.wholesale_price
    } for product in products])

@app.route('/debug-auth')
def debug_auth():
    """صفحة تشخيص مشاكل المصادقة"""
    from flask_wtf.csrf import generate_csrf
    
    debug_info = {
        'is_authenticated': current_user.is_authenticated,
        'user_id': current_user.id if current_user.is_authenticated else None,
        'username': current_user.username if current_user.is_authenticated else None,
        'user_role': current_user.role if current_user.is_authenticated else None,
        'is_system_user': current_user.is_system if current_user.is_authenticated else None,
        'session_keys': list(session.keys()),
        'total_users': User.query.count(),
        'static_user_exists': User.query.filter_by(username='araby', is_system=True).first() is not None
    }
    
    # Get all users for debugging
    all_users = []
    try:
        for user in User.query.all():
            all_users.append({
                'id': user.id,
                'username': user.username,
                'role': user.role,
                'active': user.is_active,
                'system': user.is_system,
                'locked': user.is_account_locked(),
                'failed_attempts': user.failed_login_attempts,
                'has_password_hash': bool(user.password_hash)
            })
    except Exception as e:
        all_users = [{'error': str(e)}]
    
    csrf_token = generate_csrf()
    
    # Build user table rows
    user_rows = ""
    for u in all_users:
        active_class = "success" if u.get('active') else "error"
        locked_class = "error" if u.get('locked') else "success"
        hash_class = "success" if u.get('has_password_hash') else "error"
        
        user_rows += f"""<tr>
            <td>{u.get('id', 'N/A')}</td>
            <td>{u.get('username', 'N/A')}</td>
            <td>{u.get('role', 'N/A')}</td>
            <td class="{active_class}">{u.get('active', 'N/A')}</td>
            <td>{u.get('system', 'N/A')}</td>
            <td class="{locked_class}">{u.get('locked', 'N/A')}</td>
            <td>{u.get('failed_attempts', 'N/A')}</td>
            <td class="{hash_class}">{u.get('has_password_hash', 'N/A')}</td>
        </tr>"""
    
    return f"""
    <html dir="rtl">
    <head>
        <title>تشخيص المصادقة</title>
        <meta charset="utf-8">
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; }}
            .info {{ background: #f0f0f0; padding: 10px; margin: 10px 0; border-radius: 5px; }}
            .success {{ color: green; }}
            .error {{ color: red; }}
            .warning {{ color: orange; }}
            .btn {{ padding: 10px 15px; margin: 5px; border: none; border-radius: 5px; cursor: pointer; text-decoration: none; display: inline-block; }}
            .btn-green {{ background: green; color: white; }}
            .btn-blue {{ background: blue; color: white; }}
            table {{ border-collapse: collapse; width: 100%; }}
            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: center; }}
            th {{ background-color: #f2f2f2; }}
        </style>
    </head>
    <body>
        <h1>تشخيص نظام المصادقة</h1>
        
        <div class="info">
            <h3>معلومات المستخدم الحالي:</h3>
            <p><strong>مسجل الدخول:</strong> <span class="{'success' if debug_info['is_authenticated'] else 'error'}">{debug_info['is_authenticated']}</span></p>
            <p><strong>معرف المستخدم:</strong> {debug_info['user_id']}</p>
            <p><strong>اسم المستخدم:</strong> {debug_info['username']}</p>
            <p><strong>دور المستخدم:</strong> {debug_info['user_role']}</p>
            <p><strong>مستخدم نظام:</strong> {debug_info['is_system_user']}</p>
        </div>
        
        <div class="info">
            <h3>معلومات النظام:</h3>
            <p><strong>إجمالي المستخدمين:</strong> {debug_info['total_users']}</p>
            <p><strong>المستخدم الثابت موجود:</strong> <span class="{'success' if debug_info['static_user_exists'] else 'error'}">{debug_info['static_user_exists']}</span></p>
        </div>
        
        <div class="info">
            <h3>جميع المستخدمين:</h3>
            <table>
                <tr>
                    <th>المعرف</th>
                    <th>اسم المستخدم</th>
                    <th>الدور</th>
                    <th>نشط</th>
                    <th>نظام</th>
                    <th>مقفل</th>
                    <th>محاولات فاشلة</th>
                    <th>له كلمة مرور</th>
                </tr>
                {user_rows}
            </table>
        </div>
        
        <div class="info">
            <h3>الإجراءات:</h3>
            <a href="{url_for('index')}" class="btn btn-blue">صفحة تسجيل الدخول</a>
            <a href="{url_for('dashboard')}" class="btn btn-blue">لوحة التحكم</a>
        </div>
        
        <div class="info">
            <h3>اختبار المستخدم الثابت:</h3>
            <form method="POST" action="{url_for('index')}">
                <input type="hidden" name="csrf_token" value="{csrf_token}">
                <input type="hidden" name="username" value="araby">
                <input type="hidden" name="password" value="92321066">
                <button type="submit" class="btn btn-green">تسجيل دخول بالمستخدم الثابت</button>
            </form>
        </div>
        
        <div class="info">
            <h3>اختبار يدوي:</h3>
            <form method="POST" action="{url_for('index')}">
                <input type="hidden" name="csrf_token" value="{csrf_token}">
                <div style="margin: 10px 0;">
                    <label>اسم المستخدم:</label><br>
                    <input type="text" name="username" style="padding: 5px; width: 200px;" placeholder="ادخل اسم المستخدم">
                </div>
                <div style="margin: 10px 0;">
                    <label>كلمة المرور:</label><br>
                    <input type="password" name="password" style="padding: 5px; width: 200px;" placeholder="ادخل كلمة المرور">
                </div>
                <button type="submit" class="btn btn-green">تسجيل الدخول</button>
            </form>
        </div>
        
        <div class="info">
            <h3>أوامر التشخيص (للمطور):</h3>
            <p><code>python manage.py test-password</code> - اختبار كلمة مرور مستخدم</p>
            <p><code>python manage.py fix-password</code> - إصلاح كلمة مرور مستخدم</p>
            <p><code>python manage.py list-users</code> - عرض جميع المستخدمين</p>
        </div>
    </body>
    </html>
    """

@app.route('/api/sync', methods=['POST'])
@login_required
@seller_or_admin_required
def api_sync():
    """نقطة نهاية مخصصة لمزامنة البيانات غير المتصلة"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'البيانات مطلوبة'}), 400

        sync_type = data.get('type')
        sync_data = data.get('data', [])
        
        results = {
            'success': [],
            'errors': [],
            'total': len(sync_data)
        }

        if sync_type == 'sales':
            # مزامنة المبيعات المحفوظة محلياً
            for sale_data in sync_data:
                try:
                    # التحقق من البيانات المطلوبة
                    if not sale_data.get('items') or len(sale_data['items']) == 0:
                        results['errors'].append({
                            'data': sale_data,
                            'error': 'لا توجد عناصر في البيع'
                        })
                        continue

                    # إنشاء المبيعة
                    sale = Sale(
                        subtotal=sale_data.get('subtotal', 0),
                        total_amount=sale_data.get('total_amount', 0),
                        discount_type=sale_data.get('discount_type', 'none'),
                        discount_value=sale_data.get('discount_value', 0),
                        discount_amount=sale_data.get('discount_amount', 0),
                        user_id=current_user.id,
                        customer_id=sale_data.get('customer_id'),
                        payment_status=sale_data.get('payment_status', 'paid'),
                        payment_type=sale_data.get('payment_type', 'cash'),
                        notes=sale_data.get('notes', ''),
                        sale_date=datetime.utcnow()
                    )

                    db.session.add(sale)
                    db.session.flush()  # للحصول على sale.id

                    # إضافة عناصر البيع
                    for item_data in sale_data['items']:
                        product = Product.query.get(item_data['product_id'])
                        if not product:
                            raise ValueError(f"المنتج رقم {item_data['product_id']} غير موجود")

                        # التحقق من الكمية المتوفرة
                        if product.stock_quantity < item_data['quantity']:
                            raise ValueError(f"الكمية المطلوبة من {product.name_ar} غير متوفرة")

                        sale_item = SaleItem(
                            sale_id=sale.id,
                            product_id=item_data['product_id'],
                            quantity=item_data['quantity'],
                            unit_price=item_data['unit_price'],
                            total_price=item_data['total_price']
                        )

                        db.session.add(sale_item)

                        # تحديث المخزون
                        product.stock_quantity -= item_data['quantity']

                    db.session.commit()
                    
                    results['success'].append({
                        'local_id': sale_data.get('local_id'),
                        'server_id': sale.id,
                        'sale_date': sale.sale_date.isoformat()
                    })

                except Exception as e:
                    db.session.rollback()
                    results['errors'].append({
                        'data': sale_data,
                        'error': str(e)
                    })

        elif sync_type == 'customers':
            # مزامنة العملاء الجدد
            for customer_data in sync_data:
                try:
                    # التحقق من عدم وجود العميل مسبقاً
                    existing_customer = Customer.query.filter(
                        (Customer.name == customer_data['name']) |
                        (Customer.phone == customer_data.get('phone'))
                    ).first()

                    if existing_customer:
                        results['errors'].append({
                            'data': customer_data,
                            'error': 'العميل موجود مسبقاً'
                        })
                        continue

                    customer = Customer(
                        name=customer_data['name'],
                        phone=customer_data.get('phone'),
                        address=customer_data.get('address'),
                        notes=customer_data.get('notes')
                    )

                    db.session.add(customer)
                    db.session.commit()

                    results['success'].append({
                        'local_id': customer_data.get('local_id'),
                        'server_id': customer.id
                    })

                except Exception as e:
                    db.session.rollback()
                    results['errors'].append({
                        'data': customer_data,
                        'error': str(e)
                    })

        else:
            return jsonify({'error': 'نوع المزامنة غير مدعوم'}), 400

        return jsonify({
            'message': 'تمت المزامنة',
            'results': results,
            'success_count': len(results['success']),
            'error_count': len(results['errors'])
        })

    except Exception as e:
        app.logger.error(f"Sync error: {str(e)}")
        return jsonify({'error': 'حدث خطأ أثناء المزامنة'}), 500


@app.route('/api/offline-status')
@login_required
def api_offline_status():
    """إرجاع معلومات حالة التطبيق للوضع غير المتصل"""
    try:
        # إحصائيات المنتجات
        total_products = Product.query.count()
        low_stock_products = Product.query.filter(
            Product.stock_quantity <= Product.min_stock_threshold
        ).count()
        
        # إحصائيات العملاء
        total_customers = Customer.query.count()
        
        # إحصائيات المبيعات اليومية
        today = datetime.utcnow().date()
        today_sales = Sale.query.filter(
            func.date(Sale.sale_date) == today
        ).count()
        
        today_revenue = db.session.query(func.sum(Sale.total_amount)).filter(
            func.date(Sale.sale_date) == today
        ).scalar() or 0

        return jsonify({
            'status': 'online',
            'timestamp': datetime.utcnow().isoformat(),
            'stats': {
                'products': {
                    'total': total_products,
                    'low_stock': low_stock_products
                },
                'customers': {
                    'total': total_customers
                },
                'sales_today': {
                    'count': today_sales,
                    'revenue': float(today_revenue)
                }
            },
            'user': {
                'id': current_user.id,
                'username': current_user.username,
                'role': current_user.role
            }
        })

    except Exception as e:
        app.logger.error(f"Offline status error: {str(e)}")
        return jsonify({'error': 'حدث خطأ في الحصول على الحالة'}), 500


@app.route('/offline.html')
def offline_page():
    """صفحة الوضع غير المتصل"""
    return render_template('offline.html')


@app.route('/offline-demo')
@login_required
def offline_demo():
    """صفحة اختبار الوظائف غير المتصلة"""
    return render_template('offline-demo.html')


# إضافة route لدعم Service Worker
@app.route('/static/js/service-worker.js')
def service_worker():
    """تقديم Service Worker مع headers صحيحة"""
    response = make_response(send_from_directory('static/js', 'service-worker.js'))
    response.headers['Content-Type'] = 'application/javascript'
    response.headers['Service-Worker-Allowed'] = '/'
    return response

# Route إضافي للـ Service Worker في المسار الجذر
@app.route('/service-worker.js')
def service_worker_root():
    """تقديم Service Worker من المسار الجذر"""
    try:
        response = make_response(send_from_directory('static/js', 'service-worker.js'))
        response.headers['Content-Type'] = 'application/javascript; charset=utf-8'
        response.headers['Service-Worker-Allowed'] = '/'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return response
    except FileNotFoundError:
        # إنشاء service worker أساسي إذا لم يكن موجود
        basic_sw = """
// Basic Service Worker للطوارئ
const CACHE_NAME = 'norko-store-basic-v1';

self.addEventListener('install', event => {
    console.log('Basic Service Worker: Installing...');
    self.skipWaiting();
});

self.addEventListener('activate', event => {
    console.log('Basic Service Worker: Activating...');
    return self.clients.claim();
});

self.addEventListener('fetch', event => {
    // أساسي - لا يفعل شيء خاص
    return;
});
"""
        response = make_response(basic_sw)
        response.headers['Content-Type'] = 'application/javascript; charset=utf-8'
        response.headers['Service-Worker-Allowed'] = '/'
        return response

# Route للتحقق من Service Worker
@app.route('/sw-check')
def service_worker_check():
    """فحص حالة Service Worker"""
    import os
    sw_path = os.path.join('static', 'js', 'service-worker.js')
    exists = os.path.exists(sw_path)
    
    return jsonify({
        'service_worker_exists': exists,
        'path': sw_path,
        'routes_available': [
            '/service-worker.js',
            '/static/js/service-worker.js'
        ]
    })

# Route للتشخيص الشامل
@app.route('/offline-diagnostic')
@login_required
def offline_diagnostic():
    """تشخيص شامل للوظائف غير المتصلة"""
    import os
    
    # فحص الملفات
    js_files = {
        'service-worker.js': os.path.exists('static/js/service-worker.js'),
        'db-manager.js': os.path.exists('static/js/db-manager.js'),
        'sync-manager.js': os.path.exists('static/js/sync-manager.js'),
        'offline-handler.js': os.path.exists('static/js/offline-handler.js')
    }
    
    # فحص القوالب
    templates = {
        'offline.html': os.path.exists('templates/offline.html'),
        'offline-demo.html': os.path.exists('templates/offline-demo.html')
    }
    
    # معلومات الخادم
    server_info = {
        'host': request.host,
        'scheme': request.scheme,
        'user_agent': request.headers.get('User-Agent', ''),
        'secure': request.is_secure
    }
    
    return jsonify({
        'timestamp': datetime.utcnow().isoformat(),
        'files': js_files,
        'templates': templates,
        'server': server_info,
        'endpoints': {
            '/api/sync': True,
            '/api/offline-status': True,
            '/service-worker.js': True,
            '/static/js/service-worker.js': True
        }
    })

# Route بديل لملفات JavaScript مع MIME type صحيح
@app.route('/js/<filename>')
def serve_js(filename):
    """تقديم ملفات JavaScript مع MIME type صحيح"""
    try:
        response = make_response(send_from_directory('static/js', filename))
        response.headers['Content-Type'] = 'application/javascript; charset=utf-8'
        
        if 'service-worker' in filename:
            response.headers['Service-Worker-Allowed'] = '/'
            response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        
        return response
    except FileNotFoundError:
        return jsonify({'error': f'File {filename} not found'}), 404

# Route لـ PWA Manifest
@app.route('/manifest.json')
def serve_manifest():
    """تقديم PWA manifest مع headers صحيحة"""
    try:
        response = make_response(send_from_directory('static', 'manifest.json'))
        response.headers['Content-Type'] = 'application/manifest+json'
        return response
    except FileNotFoundError:
        return jsonify({'error': 'Manifest not found'}), 404

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_from_directory, session, make_response
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from functools import wraps
from flask_migrate import Migrate
from flask_wtf.csrf import CSRFProtect
from flask_mail import Mail, Message
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_talisman import Talisman
from datetime import datetime, timedelta
import pytz
from sqlalchemy import func, desc, and_
import json
import os
import logging
import secrets
from werkzeug.utils import secure_filename
import tempfile
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from config import config
from models import db, User, Category, Product, Sale, SaleItem, Customer, Payment, Expense, ShoppingList, Return, ReturnItem
from forms import LoginForm, UserForm, CategoryForm, ProductForm, SaleForm, SaleItemForm, StockUpdateForm, CustomerForm, PaymentForm, ExpenseForm, ShoppingListForm

app = Flask(__name__)

# Load configuration based on environment
config_name = os.environ.get('FLASK_CONFIG') or 'default'
app.config.from_object(config[config_name])

# تعيين إعدادات الترميز للنصوص العربية
app.config['JSON_AS_ASCII'] = False
app.config['JSONIFY_MIMETYPE'] = 'application/json; charset=utf-8'

# Initialize extensions
db.init_app(app)
migrate = Migrate(app, db)
csrf = CSRFProtect(app)

# Initialize security extensions
mail = Mail(app)

# Rate limiting (disabled in development)
if not app.debug:
    limiter = Limiter(
        key_func=get_remote_address,
        app=app,
        default_limits=["1000 per hour"]
    )
else:
    # Mock limiter for development
    class MockLimiter:
        def limit(self, *args, **kwargs):
            def decorator(f):
                return f
            return decorator
    limiter = MockLimiter()

# Security headers (production only)
if not app.debug and app.config.get('FLASK_CONFIG') != 'development':
    try:
        Talisman(
            app,
            force_https=app.config.get('SESSION_COOKIE_SECURE', False),
            content_security_policy=app.config.get('SECURITY_HEADERS', {}).get('Content-Security-Policy')
        )
    except Exception as e:
        app.logger.warning(f"Could not initialize Talisman: {e}")

# Login manager
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'index'
login_manager.login_message = 'يرجى تسجيل الدخول للوصول إلى هذه الصفحة'
login_manager.session_protection = 'strong'

@login_manager.user_loader
def load_user(user_id):
    try:
        return User.query.get(int(user_id))
    except:
        return None

def admin_required(f):
    """Decorator للتحقق من صلاحيات الأدمن"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('index'))
        if current_user.role != 'admin':
            flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def seller_or_admin_required(f):
    """Decorator للتحقق من صلاحيات البائع أو الأدمن"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('index'))
        if current_user.role not in ['admin', 'seller']:
            flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def format_currency(amount):
    """Format currency for Egyptian Pounds"""
    if amount is None:
        return "غير محدد"
    try:
        return f"{float(amount):.2f} ج.م"
    except (ValueError, TypeError):
        return "غير محدد"

def format_date(date):
    """تنسيق التاريخ بالعربية"""
    months = ['يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو', 
              'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر']
    return f"{date.day} {months[date.month-1]} {date.year}"

def get_egypt_time(utc_datetime=None):
    """تحويل التوقيت من UTC إلى توقيت مصر"""
    egypt_tz = pytz.timezone('Africa/Cairo')
    if utc_datetime is None:
        utc_datetime = datetime.utcnow()
    
    # إذا كان التاريخ لا يحتوي على معلومات المنطقة الزمنية، افترض أنه UTC
    if utc_datetime.tzinfo is None:
        utc_datetime = pytz.utc.localize(utc_datetime)
    
    # تحويل إلى توقيت مصر
    egypt_time = utc_datetime.astimezone(egypt_tz)
    return egypt_time

def format_egypt_datetime(utc_datetime):
    """تنسيق التاريخ والوقت بتوقيت مصر"""
    egypt_time = get_egypt_time(utc_datetime)
    return egypt_time.strftime('%Y-%m-%d %H:%M:%S')

def format_egypt_time_only(utc_datetime):
    """تنسيق الوقت فقط بتوقيت مصر بنظام 12 ساعة"""
    egypt_time = get_egypt_time(utc_datetime)
    time_str = egypt_time.strftime('%I:%M:%S %p')
    # تحويل AM/PM إلى العربية
    time_str = time_str.replace('AM', 'ص').replace('PM', 'م')
    return time_str

def format_egypt_date_only(utc_datetime):
    """تنسيق التاريخ فقط بتوقيت مصر"""
    egypt_time = get_egypt_time(utc_datetime)
    return egypt_time.strftime('%d/%m/%Y')

# Template filters
app.jinja_env.filters['currency'] = format_currency
app.jinja_env.filters['arabic_date'] = format_date
app.jinja_env.filters['egypt_datetime'] = format_egypt_datetime
app.jinja_env.filters['egypt_time'] = format_egypt_time_only
app.jinja_env.filters['egypt_date'] = format_egypt_date_only

# إضافة header لضمان الترميز الصحيح للنصوص العربية
@app.after_request
def after_request(response):
    """إضافة headers أمان وتحسين الأداء"""
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    
    # إصلاح MIME types لملفات JavaScript
    if response.mimetype == 'text/html' and '.js' in request.path:
        response.headers['Content-Type'] = 'application/javascript; charset=utf-8'
    
    # إضافة headers خاصة بـ Service Worker
    if 'service-worker' in request.path:
        response.headers['Content-Type'] = 'application/javascript; charset=utf-8'
        response.headers['Service-Worker-Allowed'] = '/'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    
    return response

@app.context_processor
def inject_debt_stats():
    """إضافة إحصائيات الديون إلى جميع القوالب"""
    if current_user.is_authenticated:
        try:
            # إحصائيات الديون
            # حساب إجمالي المبيعات الآجلة (غير المدفوعة بالكامل)
            unpaid_sales_total = db.session.query(func.sum(Sale.total_amount)).filter(
                Sale.payment_status != 'paid'
            ).scalar() or 0
            
            # حساب إجمالي الدفعات للمبيعات الآجلة
            total_payments = db.session.query(func.sum(Payment.amount)).join(Sale).filter(
                Sale.payment_status != 'paid'
            ).scalar() or 0
            
            # إجمالي الديون = إجمالي المبيعات الآجلة - إجمالي الدفعات
            total_debt = max(0, unpaid_sales_total - total_payments)
            
            customers_with_debt = Customer.query.filter(Customer.id.in_(
                db.session.query(Sale.customer_id).filter(Sale.payment_status != 'paid').distinct()
            )).count()
            return dict(global_total_debt=total_debt, global_customers_with_debt=customers_with_debt)
        except:
            return dict(global_total_debt=0, global_customers_with_debt=0)
    return dict(global_total_debt=0, global_customers_with_debt=0)

@app.route('/', methods=['GET', 'POST'])
@limiter.limit("10 per minute")
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    # عرض صفحة تسجيل الدخول مباشرة في الصفحة الرئيسية
    form = LoginForm()
    if form.validate_on_submit():
        try:
            # تصفية وتنظيف البيانات
            username = form.username.data.strip().lower() if form.username.data else ""
            password = form.password.data.strip() if form.password.data else ""
            
            # التحقق من وجود البيانات
            if not username or not password:
                flash('يرجى إدخال اسم المستخدم وكلمة المرور', 'error')
                return render_template('auth/login.html', form=form)
            
            app.logger.info(f'Login attempt for username: {username} from IP: {request.remote_addr}')
            
            # التحقق من المستخدم الثابت أولاً
            if username == "araby" and password == "92321066":
                try:
                    # إنشاء المستخدم الثابت إذا لم يكن موجوداً
                    from models import create_static_user
                    create_static_user()
                    
                    # البحث عن المستخدم الثابت
                    static_user = User.query.filter_by(username="araby", is_system=True).first()
                    if static_user:
                        # تنظيف الجلسة السابقة
                        session.clear()
                        
                        # تسجيل الدخول
                        login_result = login_user(static_user, remember=form.remember_me.data, force=True)
                        
                        if login_result:
                            # تأكيد تسجيل الدخول
                            session.permanent = True
                            session['user_id'] = static_user.id
                            
                            app.logger.info(f'Static user {username} logged in successfully from IP {request.remote_addr}')
                            flash('تم تسجيل الدخول بنجاح', 'success')
                            
                            next_page = request.args.get('next')
                            redirect_url = next_page if next_page else url_for('dashboard')
                            
                            # إضافة تأخير صغير للتأكد من حفظ الجلسة
                            db.session.commit()
                            return redirect(redirect_url)
                        else:
                            flash('فشل في تسجيل الدخول', 'error')
                    else:
                        flash('فشل في إنشاء المستخدم الثابت', 'error')
                except Exception as e:
                    app.logger.error(f'Error during static user login: {str(e)}')
                    flash('حدث خطأ أثناء تسجيل الدخول', 'error')
            else:
                # البحث عن المستخدم العادي
                user = User.query.filter_by(username=username).first()
                
                if user:
                    app.logger.info(f'User found: {user.username}, Active: {user.is_active}')
                    
                    # Check if account is locked
                    if user.is_account_locked():
                        flash('تم قفل الحساب مؤقتاً بسبب محاولات دخول خاطئة متعددة. يرجى المحاولة لاحقاً.', 'error')
                        return render_template('auth/login.html', form=form)
                    
                    # Check if account is active
                    if not user.is_active:
                        flash('هذا الحساب غير نشط. يرجى الاتصال بالمدير.', 'error')
                        return render_template('auth/login.html', form=form)
                    
                    # Verify password
                    app.logger.info(f'Checking password for user: {user.username}')
                    
                    if user.check_password(password):
                        # Check if password has expired
                        if user.is_password_expired():
                            flash('انتهت صلاحية كلمة المرور. يرجى تغييرها.', 'warning')
                            session['pending_user_id'] = user.id
                            return redirect(url_for('change_password'))
                        
                        try:
                            # تنظيف الجلسة السابقة
                            session.clear()
                            
                            # تسجيل الدخول مع خيار التذكر
                            login_result = login_user(user, remember=form.remember_me.data, force=True)
                            
                            if login_result:
                                # تأكيد تسجيل الدخول
                                session.permanent = True
                                session['user_id'] = user.id
                                
                                # Log successful login
                                app.logger.info(f'User {username} logged in successfully from IP {request.remote_addr}')
                                flash('تم تسجيل الدخول بنجاح', 'success')
                                
                                next_page = request.args.get('next')
                                redirect_url = next_page if next_page else url_for('dashboard')
                                
                                # إضافة تأخير صغير للتأكد من حفظ الجلسة
                                db.session.commit()
                                return redirect(redirect_url)
                            else:
                                flash('فشل في تسجيل الدخول', 'error')
                        except Exception as e:
                            app.logger.error(f'Error during user login: {str(e)}')
                            flash('حدث خطأ أثناء تسجيل الدخول', 'error')
                    else:
                        # Log failed login attempt
                        app.logger.warning(f'Failed password check for user {username} from IP {request.remote_addr}')
                        flash('اسم المستخدم أو كلمة المرور غير صحيحة', 'error')
                else:
                    # Log failed login attempt - user not found
                    app.logger.warning(f'User not found: {username} from IP {request.remote_addr}')
                    flash('اسم المستخدم أو كلمة المرور غير صحيحة', 'error')
                    
        except Exception as e:
            app.logger.error(f'Unexpected error during login: {str(e)}')
            flash('حدث خطأ غير متوقع. يرجى المحاولة مرة أخرى.', 'error')
    
    return render_template('auth/login.html', form=form)

@app.route('/logout')
@login_required
def logout():
    # Log user logout
    app.logger.info(f'User {current_user.username} logged out from IP {request.remote_addr}')
    logout_user()
    session.clear()  # Clear all session data
    flash('تم تسجيل الخروج بنجاح', 'success')
    return redirect(url_for('index'))

@app.route('/forgot-password', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def forgot_password():
    """Password reset request"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        user = User.query.filter_by(email=email).first()
        
        if user and user.is_active:
            # Generate reset token
            token = user.generate_password_reset_token()
            
            # Send password reset email
            try:
                send_password_reset_email(user, token)
                flash('تم إرسال رابط إعادة تعيين كلمة المرور إلى بريدك الإلكتروني', 'success')
            except Exception as e:
                app.logger.error(f'Failed to send password reset email: {str(e)}')
                flash('حدث خطأ في إرسال البريد الإلكتروني. يرجى المحاولة لاحقاً.', 'error')
        else:
            # Same message for security (don't reveal if email exists)
            flash('تم إرسال رابط إعادة تعيين كلمة المرور إلى بريدك الإلكتروني', 'success')
        
        return redirect(url_for('index'))
    
    return render_template('auth/forgot_password.html')

@app.route('/reset-password/<token>', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def reset_password(token):
    """Reset password with token"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    # Find user with valid token
    user = None
    for u in User.query.all():
        if u.verify_password_reset_token(token):
            user = u
            break
    
    if not user:
        flash('رابط إعادة تعيين كلمة المرور غير صالح أو منتهي الصلاحية', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        new_password = request.form.get('password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        
        # Validate password
        if len(new_password) < 8:
            flash('كلمة المرور يجب أن تكون 8 أحرف على الأقل', 'error')
        elif new_password != confirm_password:
            flash('كلمات المرور غير متطابقة', 'error')
        else:
            # Reset password
            user.reset_password(new_password)
            flash('تم تغيير كلمة المرور بنجاح. يمكنك الآن تسجيل الدخول.', 'success')
            return redirect(url_for('index'))
    
    return render_template('auth/reset_password.html', token=token)

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    """Change password for logged-in user"""
    if request.method == 'POST':
        current_password = request.form.get('current_password', '').strip()
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        
        # Verify current password
        if not current_user.check_password(current_password):
            flash('كلمة المرور الحالية غير صحيحة', 'error')
        elif len(new_password) < 8:
            flash('كلمة المرور الجديدة يجب أن تكون 8 أحرف على الأقل', 'error')
        elif new_password != confirm_password:
            flash('كلمات المرور الجديدة غير متطابقة', 'error')
        else:
            # Change password
            current_user.set_password(new_password)
            db.session.commit()
            flash('تم تغيير كلمة المرور بنجاح', 'success')
            
            # Clear pending user session if exists
            session.pop('pending_user_id', None)
            return redirect(url_for('dashboard'))
    
    return render_template('auth/change_password.html')

def send_password_reset_email(user, token):
    """Send password reset email"""
    if not app.config.get('MAIL_USERNAME'):
        raise Exception("Email not configured")
    
    subject = 'إعادة تعيين كلمة المرور - إدارة Sara Store'
    reset_url = url_for('reset_password', token=token, _external=True)
    
    body = f"""
    مرحباً {user.username},
    
    تم طلب إعادة تعيين كلمة المرور لحسابك.
    
    للمتابعة، اضغط على الرابط التالي:
    {reset_url}
    
    هذا الرابط صالح لمدة ساعة واحدة فقط.
    
    إذا لم تطلب إعادة تعيين كلمة المرور، يرجى تجاهل هذه الرسالة.
    
    تحياتنا،
    فريق إدارة Sara Store
    """
    
    msg = Message(
        subject=subject,
        recipients=[user.email],
        body=body
    )
    
    mail.send(msg)

@app.route('/dashboard')
@login_required
def dashboard():
    # Get statistics
    total_products = Product.query.count()
    low_stock_products = Product.query.filter(Product.stock_quantity <= Product.min_stock_threshold).count()
    out_of_stock_products = Product.query.filter(Product.stock_quantity <= 0).count()
    total_categories = Category.query.count()
    
    # Sales statistics
    today = datetime.now().date()
    today_sales = Sale.query.filter(func.date(Sale.sale_date) == today).all()
    today_revenue = sum(sale.total_amount for sale in today_sales)
    
    # This month sales
    month_start = datetime.now().replace(day=1).date()
    month_sales = Sale.query.filter(func.date(Sale.sale_date) >= month_start).all()
    month_revenue = sum(sale.total_amount for sale in month_sales)
    
    # حساب أرباح ومصاريف الشهر الحالي
    month_profit = sum(sale.total_profit for sale in month_sales)
    month_cost = sum(sale.cost_amount for sale in month_sales)
    
    # مصاريف الشهر الحالي
    month_expenses = Expense.query.filter(func.date(Expense.expense_date) >= month_start).all()
    month_total_expenses = sum(expense.amount for expense in month_expenses)
    
    # صافي ربح الشهر
    month_net_profit = month_profit - month_total_expenses
    
    # إحصائيات اليوم
    today_profit = sum(sale.total_profit for sale in today_sales)
    today_expenses = Expense.query.filter(func.date(Expense.expense_date) == today).all()
    today_total_expenses = sum(expense.amount for expense in today_expenses)
    today_net_profit = today_profit - today_total_expenses
    
    # Recent sales
    recent_sales = Sale.query.order_by(desc(Sale.sale_date)).limit(5).all()
    
    # Low stock alerts
    low_stock_alerts = Product.query.filter(Product.stock_quantity <= Product.min_stock_threshold).all()
    
    # Top selling products (this month)
    top_products = db.session.query(
        Product.name_ar,
        func.sum(SaleItem.quantity).label('total_sold')
    ).join(SaleItem).join(Sale).filter(
        func.date(Sale.sale_date) >= month_start
    ).group_by(Product.id).order_by(desc('total_sold')).limit(5).all()
    
    # إحصائيات الديون
    # حساب إجمالي المبيعات الآجلة (غير المدفوعة بالكامل)
    unpaid_sales_total = db.session.query(func.sum(Sale.total_amount)).filter(
        Sale.payment_status != 'paid'
    ).scalar() or 0
    
    # حساب إجمالي الدفعات للمبيعات الآجلة
    total_payments = db.session.query(func.sum(Payment.amount)).join(Sale).filter(
        Sale.payment_status != 'paid'
    ).scalar() or 0
    
    # إجمالي الديون = إجمالي المبيعات الآجلة - إجمالي الدفعات
    total_debt = max(0, unpaid_sales_total - total_payments)
    
    customers_with_debt = Customer.query.filter(Customer.id.in_(
        db.session.query(Sale.customer_id).filter(Sale.payment_status != 'paid').distinct()
    )).count()
    
    return render_template('dashboard.html', 
                         total_products=total_products,
                         low_stock_products=low_stock_products,
                         out_of_stock_products=out_of_stock_products,
                         total_categories=total_categories,
                         today_revenue=today_revenue,
                         month_revenue=month_revenue,
                         # إحصائيات الأرباح والمصاريف
                         month_profit=month_profit,
                         month_cost=month_cost,
                         month_total_expenses=month_total_expenses,
                         month_net_profit=month_net_profit,
                         today_profit=today_profit,
                         today_expenses=today_total_expenses,  # Fixed: added missing today_expenses
                         today_total_expenses=today_total_expenses,
                         today_net_profit=today_net_profit,
                         recent_sales=recent_sales,
                         low_stock_alerts=low_stock_alerts,
                         top_products=top_products,
                         total_debt=total_debt,
                         customers_with_debt=customers_with_debt)

@app.route('/products')
@login_required
@admin_required
def products():
    search = request.args.get('search', '', type=str)
    category_id = request.args.get('category', 0, type=int)
    stock_status = request.args.get('stock_status', '', type=str)
    min_price = request.args.get('min_price', type=float)
    max_price = request.args.get('max_price', type=float)
    unit_type = request.args.get('unit_type', '', type=str)
    sort_by = request.args.get('sort_by', 'name', type=str)
    
    query = Product.query
    
    # تطبيق فلاتر البحث
    if search:
        query = query.filter(
            Product.name_ar.contains(search) | 
            Product.description_ar.contains(search)
        )
    
    if category_id:
        query = query.filter(Product.category_id == category_id)
    
    if stock_status:
        if stock_status == 'available':
            query = query.filter(Product.stock_quantity > Product.min_stock_threshold)
        elif stock_status == 'low':
            query = query.filter(
                Product.stock_quantity <= Product.min_stock_threshold,
                Product.stock_quantity > 0
            )
        elif stock_status == 'out':
            query = query.filter(Product.stock_quantity <= 0)
    
    if min_price is not None:
        query = query.filter(Product.retail_price >= min_price)
    
    if max_price is not None:
        query = query.filter(Product.retail_price <= max_price)
    
    if unit_type:
        query = query.filter(Product.unit_type == unit_type)
    
    # ترتيب النتائج
    if sort_by == 'name':
        query = query.order_by(Product.name_ar)
    elif sort_by == 'price':
        query = query.order_by(desc(Product.retail_price))
    elif sort_by == 'stock':
        query = query.order_by(desc(Product.stock_quantity))
    elif sort_by == 'date':
        query = query.order_by(desc(Product.created_at))
    else:
        query = query.order_by(Product.name_ar)
    
    # الحصول على جميع المنتجات (بدون تقسيم)
    products = query.all()
    categories = Category.query.all()
    
    # حساب الإحصائيات
    total_wholesale_value = 0
    total_retail_value = 0
    total_profit = 0
    total_products_count = len(products)
    total_stock_quantity = 0
    
    for product in products:
        wholesale_price = product.wholesale_price or 0
        retail_price = product.retail_price or product.price or 0
        stock_quantity = product.stock_quantity or 0
        
        total_wholesale_value += wholesale_price * stock_quantity
        total_retail_value += retail_price * stock_quantity
        total_profit += (retail_price - wholesale_price) * stock_quantity
        total_stock_quantity += stock_quantity
    
    # إحصائيات إضافية
    low_stock_count = sum(1 for p in products if p.is_low_stock)
    out_of_stock_count = sum(1 for p in products if p.is_out_of_stock)
    
    product_stats = {
        'total_products_count': total_products_count,
        'total_wholesale_value': total_wholesale_value,
        'total_retail_value': total_retail_value,
        'total_profit': total_profit,
        'total_stock_quantity': total_stock_quantity,
        'low_stock_count': low_stock_count,
        'out_of_stock_count': out_of_stock_count,
        'profit_margin_percentage': (total_profit / total_wholesale_value * 100) if total_wholesale_value > 0 else 0
    }
    
    return render_template('products/list.html', 
                         products=products, 
                         categories=categories,
                         product_stats=product_stats)

@app.route('/products/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_product():
    if not current_user.is_admin():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('products'))
    
    form = ProductForm()
    if form.validate_on_submit():
        product = Product(
            name_ar=form.name_ar.data,
            description_ar=form.description_ar.data,
            category_id=form.category_id.data,
            wholesale_price=form.wholesale_price.data,
            retail_price=form.retail_price.data,
            price=form.retail_price.data,  # للتوافق مع الكود القديم
            stock_quantity=form.stock_quantity.data,
            min_stock_threshold=form.min_stock_threshold.data,
            unit_type=form.unit_type.data,
            unit_description=form.unit_description.data
        )
        db.session.add(product)
        db.session.commit()
        
        # التحقق من نوع الإجراء المطلوب
        action = request.form.get('action', 'save_and_exit')
        
        if action == 'save_and_continue':
            flash('تم إضافة المنتج بنجاح! يمكنك إضافة منتج آخر.', 'success')
            return redirect(url_for('add_product', success=1))
        else:
            flash('تم إضافة المنتج بنجاح', 'success')
            return redirect(url_for('products'))
    
    return render_template('products/form.html', form=form, title='إضافة منتج جديد')

@app.route('/products/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_product(id):
    if not current_user.is_admin():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('products'))
    
    product = Product.query.get_or_404(id)
    form = ProductForm(obj=product)
    
    if form.validate_on_submit():
        form.populate_obj(product)
        product.updated_at = datetime.utcnow()
        db.session.commit()
        
        # التحقق من نوع الإجراء المطلوب
        action = request.form.get('action', 'save_and_exit')
        
        if action == 'save_and_continue':
            flash('تم تحديث المنتج بنجاح! يمكنك إضافة منتج جديد.', 'success')
            return redirect(url_for('add_product', success=1))
        else:
            flash('تم تحديث المنتج بنجاح', 'success')
            return redirect(url_for('products'))
    
    return render_template('products/form.html', form=form, title='تعديل المنتج')

@app.route('/products/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_product(id):
    if not current_user.is_admin():
        flash('ليس لديك صلاحية لحذف المنتجات', 'error')
        return redirect(url_for('products'))
    
    product = Product.query.get_or_404(id)
    db.session.delete(product)
    db.session.commit()
    flash('تم حذف المنتج بنجاح', 'success')
    return redirect(url_for('products'))

@app.route('/categories')
@login_required
@admin_required
def categories():
    if not current_user.is_admin():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('dashboard'))
    
    categories = Category.query.order_by(Category.name_ar).all()
    return render_template('categories/list.html', categories=categories)

@app.route('/categories/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_category():
    if not current_user.is_admin():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('categories'))
    
    form = CategoryForm()
    if form.validate_on_submit():
        category = Category(
            name_ar=form.name_ar.data,
            description_ar=form.description_ar.data
        )
        db.session.add(category)
        db.session.commit()
        flash('تم إضافة الفئة بنجاح', 'success')
        return redirect(url_for('categories'))
    
    return render_template('categories/form.html', form=form, title='إضافة فئة جديدة')

@app.route('/categories/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_category(id):
    if not current_user.is_admin():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('categories'))
    
    category = Category.query.get_or_404(id)
    form = CategoryForm(obj=category)
    
    if form.validate_on_submit():
        form.populate_obj(category)
        db.session.commit()
        flash('تم تحديث الفئة بنجاح', 'success')
        return redirect(url_for('categories'))
    
    return render_template('categories/form.html', form=form, title='تعديل الفئة')

@app.route('/categories/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_category(id):
    """حذف فئة"""
    if not current_user.is_admin():
        flash('ليس لديك صلاحية لحذف الفئات', 'error')
        return redirect(url_for('categories'))
    
    category = Category.query.get_or_404(id)
    
    # التحقق من عدم وجود منتجات في الفئة
    if category.products:
        flash('لا يمكن حذف فئة تحتوي على منتجات', 'error')
        return redirect(url_for('categories'))
    
    category_name = category.name_ar
    db.session.delete(category)
    db.session.commit()
    flash(f'تم حذف الفئة "{category_name}" بنجاح', 'success')
    return redirect(url_for('categories'))

# User management routes
@app.route('/users')
@login_required
@admin_required
def users():
    """عرض قائمة المستخدمين"""
    if not current_user.is_admin():
        flash('ليس لديك صلاحية للوصول إلى إدارة المستخدمين', 'error')
        return redirect(url_for('dashboard'))
    
    # إخفاء مستخدمي النظام من القائمة
    users = User.query.filter_by(is_system=False).order_by(User.created_at.desc()).all()
    return render_template('users/list.html', users=users)

@app.route('/users/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_user():
    """إضافة مستخدم جديد"""
    if not current_user.is_admin():
        flash('ليس لديك صلاحية لإضافة المستخدمين', 'error')
        return redirect(url_for('dashboard'))
    
    form = UserForm()
    if form.validate_on_submit():
        user = User(
            username=form.username.data,
            role=form.role.data
        )
        user.set_password(form.password.data)
        
        try:
            db.session.add(user)
            db.session.commit()
            flash(f'تم إضافة المستخدم "{user.username}" بنجاح', 'success')
            return redirect(url_for('users'))
        except Exception as e:
            db.session.rollback()
            flash('حدث خطأ أثناء إضافة المستخدم', 'error')
    
    return render_template('users/add.html', form=form)

@app.route('/users/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(id):
    """تعديل مستخدم"""
    if not current_user.is_admin():
        flash('ليس لديك صلاحية لتعديل المستخدمين', 'error')
        return redirect(url_for('dashboard'))
    
    user = User.query.get_or_404(id)
    
    # منع تعديل مستخدم النظام
    if user.is_system:
        flash('لا يمكن تعديل مستخدم النظام', 'error')
        return redirect(url_for('users'))
    
    # منع تعديل نفس المستخدم
    if user.id == current_user.id:
        flash('لا يمكنك تعديل حسابك الشخصي من هنا', 'error')
        return redirect(url_for('users'))
    
    form = UserForm(original_username=user.username, is_edit=True)
    if form.validate_on_submit():
        user.username = form.username.data
        user.role = form.role.data
        
        # تحديث كلمة المرور إذا تم إدخال واحدة جديدة
        if form.password.data:
            user.set_password(form.password.data)
        
        try:
            db.session.commit()
            flash(f'تم تحديث بيانات المستخدم "{user.username}" بنجاح', 'success')
            return redirect(url_for('users'))
        except Exception as e:
            db.session.rollback()
            flash('حدث خطأ أثناء تحديث المستخدم', 'error')
    
    # ملء النموذج بالبيانات الحالية
    if request.method == 'GET':
        form.username.data = user.username
        form.role.data = user.role
        form.password.data = ''  # لا نعرض كلمة المرور
    
    return render_template('users/edit.html', form=form, user=user)

@app.route('/users/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_user(id):
    """حذف مستخدم"""
    if not current_user.is_admin():
        flash('ليس لديك صلاحية لحذف المستخدمين', 'error')
        return redirect(url_for('dashboard'))
    
    user = User.query.get_or_404(id)
    
    # منع حذف مستخدم النظام
    if user.is_system:
        flash('لا يمكن حذف مستخدم النظام', 'error')
        return redirect(url_for('users'))
    
    # منع حذف نفس المستخدم
    if user.id == current_user.id:
        flash('لا يمكنك حذف حسابك الشخصي', 'error')
        return redirect(url_for('users'))
    
    # التحقق من وجود مبيعات للمستخدم
    if user.sales:
        flash('لا يمكن حذف مستخدم لديه مبيعات مسجلة', 'error')
        return redirect(url_for('users'))
    
    username = user.username
    db.session.delete(user)
    db.session.commit()
    flash(f'تم حذف المستخدم "{username}" بنجاح', 'success')
    return redirect(url_for('users'))

# Customer management routes
@app.route('/customers')
@login_required
@admin_required
def customers():
    """عرض قائمة العملاء"""
    customers = Customer.query.order_by(Customer.name).all()
    return render_template('customers/list.html', customers=customers)

@app.route('/customers/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_customer():
    """إضافة عميل جديد"""
    if request.method == 'POST':
        # Check if it's a JSON request (for quick add from sales page)
        if request.is_json:
            data = request.get_json()
            customer = Customer(
                name=data['name'],
                phone=data.get('phone', ''),
                address=data.get('address', ''),
                notes=data.get('notes', '')
            )
            db.session.add(customer)
            db.session.commit()
            return jsonify({'success': True, 'customer_id': customer.id})
        
        # Regular form submission
        form = CustomerForm()
        if form.validate_on_submit():
            customer = Customer(
                name=form.name.data,
                phone=form.phone.data,
                address=form.address.data,
                notes=form.notes.data
            )
            db.session.add(customer)
            db.session.commit()
            flash(f'تم إضافة العميل "{customer.name}" بنجاح', 'success')
            return redirect(url_for('customers'))
        return render_template('customers/form.html', form=form, title='إضافة عميل جديد')
    
    # GET request
    form = CustomerForm()
    return render_template('customers/form.html', form=form, title='إضافة عميل جديد')

@app.route('/customers/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_customer(id):
    """تعديل بيانات عميل"""
    customer = Customer.query.get_or_404(id)
    form = CustomerForm(obj=customer)
    if form.validate_on_submit():
        customer.name = form.name.data
        customer.phone = form.phone.data
        customer.address = form.address.data
        customer.notes = form.notes.data
        db.session.commit()
        flash(f'تم تحديث بيانات العميل "{customer.name}" بنجاح', 'success')
        return redirect(url_for('customers'))
    return render_template('customers/form.html', form=form, title='تعديل بيانات العميل', customer=customer)

@app.route('/customers/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_customer(id):
    """حذف عميل"""
    customer = Customer.query.get_or_404(id)
    
    # Check if customer has any sales
    if customer.sales:
        flash('لا يمكن حذف العميل لأن لديه مبيعات مسجلة', 'error')
        return redirect(url_for('customers'))
    
    customer_name = customer.name
    db.session.delete(customer)
    db.session.commit()
    flash(f'تم حذف العميل "{customer_name}" بنجاح', 'success')
    return redirect(url_for('customers'))

@app.route('/customers/<int:id>/account')
@login_required
@admin_required
def customer_account(id):
    """عرض حساب العميل والديون"""
    customer = Customer.query.get_or_404(id)
    sales = Sale.query.filter_by(customer_id=id).order_by(desc(Sale.sale_date)).all()
    return render_template('customers/account.html', customer=customer, sales=sales)

@app.route('/customers/<int:customer_id>/sales/<int:sale_id>/payment', methods=['GET', 'POST'])
@login_required
@admin_required
def add_payment(customer_id, sale_id):
    """إضافة دفعة لمبيعة"""
    customer = Customer.query.get_or_404(customer_id)
    sale = Sale.query.get_or_404(sale_id)
    
    if sale.customer_id != customer_id:
        flash('خطأ في بيانات العميل أو المبيعة', 'error')
        return redirect(url_for('customers'))
    
    if sale.is_fully_paid:
        flash('هذه المبيعة مدفوعة بالكامل', 'info')
        return redirect(url_for('customer_account', id=customer_id))
    
    form = PaymentForm()
    if form.validate_on_submit():
        # التأكد من أن المبلغ لا يتجاوز المتبقي
        remaining = sale.remaining_amount
        if form.amount.data > remaining:
            flash(f'المبلغ المدخل أكبر من المتبقي ({remaining:.2f} ج.م)', 'error')
        else:
            payment = Payment(
                sale_id=sale_id,
                amount=form.amount.data,
                payment_method=form.payment_method.data,
                notes=form.notes.data,
                user_id=current_user.id
            )
            db.session.add(payment)
            
            # تحديث حالة الدفع
            total_paid = sale.paid_amount + form.amount.data
            if total_paid >= sale.total_amount:
                sale.payment_status = 'paid'
            else:
                sale.payment_status = 'partial'
            
            db.session.commit()
            flash(f'تم تسجيل دفعة بمبلغ {form.amount.data:.2f} ج.م بنجاح', 'success')
            return redirect(url_for('customer_account', id=customer_id))
    
    return render_template('customers/payment.html', form=form, customer=customer, sale=sale)

@app.route('/debts')
@login_required
@admin_required
def debts_report():
    """تقرير الديون"""
    # العملاء الذين لديهم ديون
    customers_with_debts = []
    customers = Customer.query.all()
    
    for customer in customers:
        debt = customer.total_debt
        if debt > 0:
            customers_with_debts.append({
                'customer': customer,
                'debt': debt,
                'unpaid_sales': [sale for sale in customer.sales if not sale.is_fully_paid]
            })
    
    # ترتيب حسب قيمة الدين (الأكبر أولاً)
    customers_with_debts.sort(key=lambda x: x['debt'], reverse=True)
    
    total_debts = sum(item['debt'] for item in customers_with_debts)
    
    return render_template('debts/report.html', 
                         customers_with_debts=customers_with_debts, 
                         total_debts=total_debts,
                         current_datetime=datetime.now())

@app.route('/sales')
@login_required
@seller_or_admin_required
def sales():
    page = request.args.get('page', 1, type=int)
    
    # Start with base query
    query = Sale.query
    
    # Apply filters from request arguments
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    product_search = request.args.get('product_search')
    seller_filter = request.args.get('seller_filter')
    amount_from = request.args.get('amount_from')
    amount_to = request.args.get('amount_to')
    
    # Date range filter
    if date_from:
        try:
            date_from_dt = datetime.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(func.date(Sale.sale_date) >= date_from_dt)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_dt = datetime.strptime(date_to, '%Y-%m-%d').date()
            query = query.filter(func.date(Sale.sale_date) <= date_to_dt)
        except ValueError:
            pass
    
    # Product search filter
    if product_search:
        # Join with SaleItem and Product to search in product names
        query = query.join(SaleItem).join(Product).filter(
            Product.name_ar.contains(product_search)
        ).distinct()
    
    # Seller filter
    if seller_filter:
        try:
            seller_id = int(seller_filter)
            query = query.filter(Sale.user_id == seller_id)
        except ValueError:
            pass
    
    # Amount range filter
    if amount_from:
        try:
            amount_from_val = float(amount_from)
            query = query.filter(Sale.total_amount >= amount_from_val)
        except ValueError:
            pass
    
    if amount_to:
        try:
            amount_to_val = float(amount_to)
            query = query.filter(Sale.total_amount <= amount_to_val)
        except ValueError:
            pass
    
    # Apply user permissions
    if not current_user.is_admin():
        query = query.filter(Sale.user_id == current_user.id)
    
    # Order by newest first and paginate
    sales = query.order_by(desc(Sale.sale_date)).paginate(
        page=page, per_page=20, error_out=False)
    
    # Get all users for seller filter dropdown
    users = User.query.all() if current_user.is_admin() else [current_user]
    
    return render_template('sales/list.html', sales=sales, users=users)

@app.route('/sales/new')
@login_required
@seller_or_admin_required
def new_sale():
    return render_template('sales/new.html')

@app.route('/api/products')
@login_required
@seller_or_admin_required
def api_products():
    """API endpoint to get all products"""
    try:
        app.logger.info(f"API products called by user: {current_user.username}")
        products = Product.query.all()
        app.logger.info(f"Found {len(products)} products in database")
        
        result = []
        for p in products:
            try:
                # Use safe attribute access with fallbacks
                wholesale_price = p.wholesale_price if p.wholesale_price else (p.price if p.price else 0)
                retail_price = p.retail_price if p.retail_price else (p.price if p.price else 0)
                
                product_data = {
                    'id': p.id,
                    'name': p.name_ar or 'منتج غير محدد',
                    'wholesale_price': float(wholesale_price),
                    'retail_price': float(retail_price),
                    'price': float(retail_price),  # Use retail_price as the main price
                    'stock': float(p.stock_quantity or 0),
                    'unit_type': p.unit_type or 'كامل',
                    'category': p.category.name_ar if p.category else 'غير محدد',
                    'min_stock_threshold': float(p.min_stock_threshold or 10),
                    'profit_margin': 0,
                    'profit_percentage': 0
                }
                
                # Add profit calculations if possible
                try:
                    if wholesale_price > 0 and retail_price > 0:
                        profit_margin = retail_price - wholesale_price
                        profit_percentage = (profit_margin / wholesale_price) * 100
                        product_data['profit_margin'] = float(profit_margin)
                        product_data['profit_percentage'] = float(profit_percentage)
                except:
                    pass
                
                result.append(product_data)
            except Exception as e:
                # Skip problematic products but log the error
                app.logger.error(f"Error processing product {p.id}: {str(e)}")
                continue
        
        app.logger.info(f"Returning {len(result)} products to client")
        return jsonify(result)
    except Exception as e:
        app.logger.error(f"Error in api_products: {str(e)}")
        return jsonify({'error': 'حدث خطأ في تحميل المنتجات'}), 500

@app.route('/api/categories')
@login_required
def api_categories():
    """API endpoint to get all categories"""
    categories = Category.query.all()
    return jsonify([{
        'id': c.id,
        'name': c.name_ar,
        'description': c.description_ar or '',
        'product_count': len(c.products),
        'created_at': c.created_at.isoformat() if hasattr(c, 'created_at') and c.created_at else None
    } for c in categories])

@app.route('/api/customers')
@login_required
@seller_or_admin_required
def api_customers():
    """API endpoint to get all customers"""
    customers = Customer.query.order_by(Customer.name).all()
    return jsonify([{
        'id': c.id,
        'name': c.name,
        'phone': c.phone or '',
        'debt': c.total_debt
    } for c in customers])

@app.route('/api/sales', methods=['POST'])
@login_required
@seller_or_admin_required
def api_create_sale():
    """API endpoint to create a new sale"""
    data = request.get_json()
    
    if not data.get('items'):
        return jsonify({'error': 'لا توجد عناصر في البيع'}), 400
    
    # Validate stock availability
    for item in data['items']:
        product = Product.query.get(item['product_id'])
        if not product:
            return jsonify({'error': f'المنتج غير موجود'}), 400
        if product.stock_quantity < item['quantity']:
            return jsonify({'error': f'الكمية المطلوبة غير متوفرة للمنتج {product.name_ar}'}), 400
    
    # Get payment info
    payment_type = data.get('payment_type', 'cash')  # 'cash' or 'credit'
    customer_id = data.get('customer_id') if payment_type == 'credit' else None
    paid_amount = float(data.get('paid_amount', 0))
    
    # Get discount info
    subtotal = float(data.get('subtotal', 0))
    discount_type = data.get('discount_type', 'none')
    discount_value = float(data.get('discount_value', 0))
    discount_amount = float(data.get('discount_amount', 0))
    total_amount = float(data['total_amount'])
    
    # Validate customer for credit sales
    if payment_type == 'credit' and not customer_id:
        return jsonify({'error': 'يجب اختيار عميل للبيع الآجل'}), 400
    
    # تحديد حالة الدفع
    if payment_type == 'cash':
        payment_status = 'paid'
        paid_amount = total_amount  # في البيع النقدي يكون المبلغ مدفوع بالكامل
    else:
        if paid_amount >= total_amount:
            payment_status = 'paid'
        elif paid_amount > 0:
            payment_status = 'partial'
        else:
            payment_status = 'unpaid'
    
    # Create sale with discount info
    sale = Sale(
        subtotal=subtotal,
        discount_type=discount_type,
        discount_value=discount_value,
        discount_amount=discount_amount,
        total_amount=total_amount,
        user_id=current_user.id,
        customer_id=customer_id,
        payment_type=payment_type,
        payment_status=payment_status,
        notes=data.get('notes', '')
    )
    db.session.add(sale)
    db.session.flush()  # Get sale.id
    
    # Create sale items and update stock
    for item in data['items']:
        sale_item = SaleItem(
            sale_id=sale.id,
            product_id=item['product_id'],
            quantity=item['quantity'],
            unit_price=item['unit_price'],
            total_price=item['total_price']
        )
        db.session.add(sale_item)
        
        # Update product stock
        product = Product.query.get(item['product_id'])
        product.stock_quantity -= item['quantity']
    
    # إضافة دفعة في حالة البيع الآجل مع دفعة مقدمة
    if payment_type == 'credit' and paid_amount > 0:
        payment = Payment(
            sale_id=sale.id,
            amount=paid_amount,
            payment_method='نقدي',
            notes='دفعة مقدمة مع البيع',
            user_id=current_user.id
        )
        db.session.add(payment)
    
    db.session.commit()
    
    # تحديد الرسالة
    if payment_type == 'cash':
        message = 'تم تسجيل البيع نقداً بنجاح'
    elif payment_status == 'paid':
        message = 'تم تسجيل البيع وتم دفع المبلغ كاملاً'
    elif payment_status == 'partial':
        remaining = total_amount - paid_amount
        message = f'تم تسجيل البيع مع دفعة مقدمة {paid_amount:.2f} ج.م - المتبقي: {remaining:.2f} ج.م'
    else:
        message = 'تم تسجيل البيع آجلاً'
    
    return jsonify({
        'success': True,
        'sale_id': sale.id,
        'message': message,
        'payment_status': payment_status,
        'paid_amount': paid_amount,
        'remaining_amount': total_amount - paid_amount
    })

@app.route('/api/sales/<int:sale_id>')
@login_required
def api_sale_details(sale_id):
    """API endpoint to get sale details"""
    sale = Sale.query.get_or_404(sale_id)
    
    # Check permissions - sellers can only view their own sales, admins can view all
    if not current_user.is_admin() and sale.user_id != current_user.id:
        return jsonify({'error': 'ليس لديك صلاحية لعرض هذا البيع'}), 403
    
    # Get sale items
    sale_items = []
    for item in sale.sale_items:
        sale_items.append({
            'product_name': item.product.name_ar,
            'quantity': float(item.quantity),
            'unit_price': float(item.unit_price),
            'total_price': float(item.total_price),
            'unit_type': item.product.unit_type
        })
    
    # تحويل التاريخ إلى توقيت مصر
    egypt_time = get_egypt_time(sale.sale_date)
    time_str = egypt_time.strftime('%I:%M:%S %p').replace('AM', 'ص').replace('PM', 'م')
    datetime_str = egypt_time.strftime('%d/%m/%Y %I:%M:%S %p').replace('AM', 'ص').replace('PM', 'م')
    
    return jsonify({
        'id': sale.id,
        'sale_date': egypt_time.strftime('%d/%m/%Y'),
        'sale_time': time_str,
        'sale_datetime': datetime_str,
        'subtotal': float(sale.subtotal or sale.total_amount),
        'discount_type': sale.discount_type or 'none',
        'discount_value': float(sale.discount_value or 0),
        'discount_amount': float(sale.discount_amount or 0),
        'total_amount': float(sale.total_amount),
        'notes': sale.notes,
        'user_name': sale.user.username,
        'user_role': sale.user.role,
        'items': sale_items
    })

# New API endpoints for data export
@app.route('/api/export/products')
@login_required
def api_export_products():
    if current_user.role not in ['admin', 'seller']:
        abort(403)
    
    products = Product.query.join(Category).all()
    return jsonify([{
        'id': p.id,
        'name': p.name_ar,
        'category': p.category.name_ar if p.category else 'غير محدد',
        'price': float(p.retail_price or p.price or 0),
        'stock': float(p.stock_quantity),
        'unit_type': p.unit_type,
        'is_whole_unit': p.is_whole_unit,
        'created_date': p.id  # Using id as proxy for creation order
    } for p in products])

@app.route('/api/export/sales')
@login_required
def api_export_sales():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = Sale.query.join(User)
    
    if start_date:
        query = query.filter(Sale.sale_date >= start_date)
    if end_date:
        query = query.filter(Sale.sale_date <= end_date + ' 23:59:59')
    
    # Filter by user role
    if current_user.role != 'admin':
        query = query.filter(Sale.user_id == current_user.id)
    
    sales = query.order_by(Sale.sale_date.desc()).all()
    
    sales_data = []
    for sale in sales:
        for item in sale.sale_items:
            sales_data.append({
                'sale_id': sale.id,
                'sale_date': format_egypt_date_only(sale.sale_date),
                'sale_time': format_egypt_time_only(sale.sale_date),
                'seller_name': sale.user.username,
                'seller_role': sale.user.role,
                'product_name': item.product.name_ar,
                'product_category': item.product.category.name_ar if item.product.category else 'غير محدد',
                'quantity': float(item.quantity),
                'unit_price': float(item.unit_price),
                'total_price': float(item.total_price),
                'unit_type': item.product.unit_type,
                'sale_total': float(sale.total_amount),
                'notes': sale.notes or ''
            })
    
    return jsonify(sales_data)

@app.route('/api/quick-payment', methods=['POST'])
@login_required
def api_quick_payment():
    """API endpoint for quick debt payment"""
    try:
        data = request.get_json()
        customer_id = data.get('customer_id')
        amount = float(data.get('amount', 0))
        payment_method = data.get('payment_method', 'نقدي')
        notes = data.get('notes', '')
        
        if not customer_id or amount <= 0:
            return jsonify({'success': False, 'message': 'بيانات غير صحيحة'}), 400
        
        # Get customer
        customer = Customer.query.get_or_404(customer_id)
        
        # Get unpaid sales for this customer
        unpaid_sales = Sale.query.filter(
            Sale.customer_id == customer_id,
            Sale.payment_status.in_(['unpaid', 'partial'])
        ).order_by(Sale.sale_date.asc()).all()
        
        if not unpaid_sales:
            return jsonify({'success': False, 'message': 'لا توجد ديون لهذا العميل'}), 400
        
        remaining_amount = amount
        payments_made = []
        
        # Distribute payment across unpaid sales
        for sale in unpaid_sales:
            if remaining_amount <= 0:
                break
            
            sale_remaining = sale.remaining_amount
            if sale_remaining <= 0:
                continue
            
            payment_amount = min(remaining_amount, sale_remaining)
            
            # Create payment record
            payment = Payment(
                sale_id=sale.id,
                amount=payment_amount,
                payment_method=payment_method,
                notes=f"{notes} - تسديد سريع",
                user_id=current_user.id
            )
            
            db.session.add(payment)
            
            # Update sale payment status
            sale_total_paid = sale.paid_amount + payment_amount
            if sale_total_paid >= sale.total_amount:
                sale.payment_status = 'paid'
            else:
                sale.payment_status = 'partial'
            
            payments_made.append({
                'sale_id': sale.id,
                'amount': payment_amount
            })
            
            remaining_amount -= payment_amount
        
        db.session.commit()
        
        message = f"تم تسديد {amount:.2f} ج.م بنجاح"
        if remaining_amount > 0:
            message += f" (متبقي {remaining_amount:.2f} ج.م كرصيد)"
        
        return jsonify({
            'success': True,
            'message': message,
            'payments_made': payments_made
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/stock-status')
@login_required
def api_stock_status():
    """API endpoint for real-time stock monitoring"""
    try:
        # Get products with low stock or out of stock
        critical_products = Product.query.filter(Product.stock_quantity <= 0).all()
        low_stock_products = Product.query.filter(
            and_(Product.stock_quantity > 0, 
                 Product.stock_quantity <= Product.min_stock_threshold)
        ).all()
        
        new_alerts = []
        
        # Add critical stock alerts (out of stock)
        for product in critical_products:
            new_alerts.append({
                'type': 'critical',
                'product_name': product.name_ar,
                'current_stock': product.stock_quantity,
                'min_threshold': product.min_stock_threshold or 10,
                'message': f'نفدت كمية {product.name_ar} من المخزون'
            })
        
        # Add low stock alerts
        for product in low_stock_products:
            new_alerts.append({
                'type': 'warning',
                'product_name': product.name_ar,
                'current_stock': product.stock_quantity,
                'min_threshold': product.min_stock_threshold or 10,
                'message': f'كمية {product.name_ar} منخفضة في المخزون'
            })
        
        return jsonify({
            'status': 'success',
            'new_alerts': new_alerts,
            'critical_count': len(critical_products),
            'low_stock_count': len(low_stock_products),
            'timestamp': datetime.now().isoformat()
        })
    
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': 'حدث خطأ في مراقبة المخزون',
            'new_alerts': []
        }), 500

@app.route('/api/export/inventory')
@login_required
def api_export_inventory():
    if current_user.role not in ['admin', 'seller']:
        abort(403)
    
    # Get products with stock status
    products = Product.query.join(Category).all()
    
    inventory_data = []
    for product in products:
        # Calculate stock status
        if product.stock_quantity <= 0:
            status = 'نفدت الكمية'
            status_en = 'Out of Stock'
        elif product.stock_quantity <= 10:
            status = 'كمية قليلة'
            status_en = 'Low Stock'
        else:
            status = 'متوفر'
            status_en = 'Available'
        
        # Calculate total sales for this product
        total_sold = db.session.query(db.func.sum(SaleItem.quantity)).filter(
            SaleItem.product_id == product.id
        ).scalar() or 0
        
        total_revenue = db.session.query(db.func.sum(SaleItem.total_price)).filter(
            SaleItem.product_id == product.id
        ).scalar() or 0
        
        inventory_data.append({
            'product_id': product.id,
            'product_name': product.name_ar,
            'category': product.category.name_ar if product.category else 'غير محدد',
            'current_stock': float(product.stock_quantity),
            'unit_type': product.unit_type,
            'unit_price': float(product.price),
            'stock_value': float(product.stock_quantity * product.price),
            'is_whole_unit': product.is_whole_unit,
            'status_ar': status,
            'status_en': status_en,
            'total_sold': float(total_sold),
            'total_revenue': float(total_revenue)
        })
    
    return jsonify(inventory_data)

@app.route('/reports')
@login_required
@admin_required
def reports():
    if current_user.role not in ['admin', 'seller']:
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('dashboard'))
    
    # Get date range from request
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if not start_date:
        start_date = datetime.now().replace(day=1).strftime('%Y-%m-%d')
    if not end_date:
        end_date = datetime.now().strftime('%Y-%m-%d')
    
    # Convert to datetime objects
    start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Sales in date range
    sales = Sale.query.filter(
        and_(func.date(Sale.sale_date) >= start_dt,
             func.date(Sale.sale_date) <= end_dt)
    ).all()
    
    total_revenue = sum(sale.total_amount for sale in sales)
    total_sales_count = len(sales)
    
    # حساب الأرباح والتكاليف
    total_profit = sum(sale.total_profit for sale in sales)
    total_cost = sum(sale.cost_amount for sale in sales)
    profit_margin = (total_profit / total_revenue * 100) if total_revenue > 0 else 0
    
    # حساب المصاريف في نفس الفترة
    expenses = Expense.query.filter(
        and_(func.date(Expense.expense_date) >= start_dt,
             func.date(Expense.expense_date) <= end_dt)
    ).all()
    
    total_expenses = sum(expense.amount for expense in expenses)
    
    # صافي الربح = إجمالي الأرباح - المصاريف
    net_profit = total_profit - total_expenses
    
    # تصنيف المصاريف حسب النوع
    expenses_by_type = {}
    for expense in expenses:
        expense_type = expense.expense_type_ar
        if expense_type not in expenses_by_type:
            expenses_by_type[expense_type] = 0
        expenses_by_type[expense_type] += expense.amount
    
    # Top products in date range
    top_products = db.session.query(
        Product.name_ar,
        func.sum(SaleItem.quantity).label('total_sold'),
        func.sum(SaleItem.total_price).label('total_revenue')
    ).join(SaleItem).join(Sale).filter(
        and_(func.date(Sale.sale_date) >= start_dt,
             func.date(Sale.sale_date) <= end_dt)
    ).group_by(Product.id).order_by(desc('total_sold')).all()
    
    # Daily sales chart data
    daily_sales_raw = db.session.query(
        func.date(Sale.sale_date).label('date'),
        func.sum(Sale.total_amount).label('total')
    ).filter(
        and_(func.date(Sale.sale_date) >= start_dt,
             func.date(Sale.sale_date) <= end_dt)
    ).group_by(func.date(Sale.sale_date)).order_by('date').all()
    
    # Convert to JSON-serializable format
    daily_sales = []
    for row in daily_sales_raw:
        date_str = row.date if isinstance(row.date, str) else row.date.strftime('%Y-%m-%d')
        daily_sales.append({'date': date_str, 'total': float(row.total or 0)})
    
    # Debt-related statistics
    # Total debts across all customers
    total_debts = sum(customer.total_debt for customer in Customer.query.all())
    
    # Count customers with debts
    customers_with_debts = Customer.query.filter(
        Customer.id.in_(
            db.session.query(Sale.customer_id).filter(
                Sale.payment_status.in_(['unpaid', 'partial'])
            ).distinct()
        )
    ).count()
    
    # Credit sales in date range
    credit_sales = Sale.query.filter(
        and_(func.date(Sale.sale_date) >= start_dt,
             func.date(Sale.sale_date) <= end_dt,
             Sale.payment_type == 'credit')
    ).all()
    total_credit_sales = len(credit_sales)
    
    # Total payments in date range
    total_payments = db.session.query(func.sum(Payment.amount)).join(Sale).filter(
        and_(func.date(Payment.payment_date) >= start_dt,
             func.date(Payment.payment_date) <= end_dt)
    ).scalar() or 0
    
    # Payment rate calculation
    total_credit_amount = sum(sale.total_amount for sale in credit_sales)
    payment_rate = (total_payments / total_credit_amount * 100) if total_credit_amount > 0 else 0
    
    # Top debtors
    top_debtors = []
    customers_with_debt = Customer.query.join(Sale).filter(
        Sale.payment_status.in_(['unpaid', 'partial'])
    ).distinct().all()
    
    for customer in customers_with_debt:
        if customer.total_debt > 0:
            unpaid_sales_count = Sale.query.filter(
                Sale.customer_id == customer.id,
                Sale.payment_status.in_(['unpaid', 'partial'])
            ).count()
            
            last_sale = Sale.query.filter(
                Sale.customer_id == customer.id
            ).order_by(Sale.sale_date.desc()).first()
            
            top_debtors.append((
                customer,
                customer.total_debt,
                unpaid_sales_count,
                last_sale.sale_date if last_sale else None
            ))
    
    # Sort by debt amount (highest first) and take top 10
    top_debtors.sort(key=lambda x: x[1], reverse=True)
    top_debtors = top_debtors[:10]
    
    return render_template('reports/index.html',
                         start_date=start_date,
                         end_date=end_date,
                         total_revenue=total_revenue,
                         total_sales_count=total_sales_count,
                         top_products=top_products,
                         daily_sales=daily_sales,
                         # Profit and expense data
                         total_profit=total_profit,
                         total_cost=total_cost,
                         profit_margin=profit_margin,
                         total_expenses=total_expenses,
                         net_profit=net_profit,
                         expenses_by_type=expenses_by_type,
                         # Debt-related data
                         total_debts=total_debts,
                         customers_with_debts=customers_with_debts,
                         total_credit_sales=total_credit_sales,
                         total_payments=total_payments,
                         payment_rate=payment_rate,
                         top_debtors=top_debtors)

# إدارة المصاريف
@app.route('/expenses')
@login_required
@admin_required
def expenses():
    """عرض قائمة المصاريف"""
    if not current_user.is_admin():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('dashboard'))
    
    page = request.args.get('page', 1, type=int)
    expense_type = request.args.get('type', '', type=str)
    start_date = request.args.get('start_date', '', type=str)
    end_date = request.args.get('end_date', '', type=str)
    
    query = Expense.query
    
    if expense_type:
        query = query.filter(Expense.expense_type == expense_type)
    
    if start_date:
        query = query.filter(func.date(Expense.expense_date) >= start_date)
    
    if end_date:
        query = query.filter(func.date(Expense.expense_date) <= end_date)
    
    expenses = query.order_by(desc(Expense.expense_date)).paginate(
        page=page, per_page=20, error_out=False)
    
    # حساب إجمالي المصاريف
    total_expenses = query.with_entities(func.sum(Expense.amount)).scalar() or 0
    
    return render_template('expenses/list.html', 
                         expenses=expenses, 
                         total_expenses=total_expenses,
                         expense_type=expense_type,
                         start_date=start_date,
                         end_date=end_date)

@app.route('/expenses/add', methods=['GET', 'POST'])
@login_required
def add_expense():
    """إضافة مصروف جديد"""
    if not current_user.is_admin():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('dashboard'))
    
    form = ExpenseForm()
    if form.validate_on_submit():
        expense = Expense(
            description=form.description.data,
            amount=form.amount.data,
            expense_type=form.expense_type.data,
            category=form.category.data,
            expense_date=form.expense_date.data or datetime.utcnow(),
            notes=form.notes.data,
            user_id=current_user.id
        )
        db.session.add(expense)
        db.session.commit()
        flash('تم إضافة المصروف بنجاح', 'success')
        return redirect(url_for('expenses'))
    
    return render_template('expenses/add.html', form=form, title='إضافة مصروف جديد')

@app.route('/expenses/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_expense(id):
    """تعديل مصروف"""
    if not current_user.is_admin():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('expenses'))
    
    expense = Expense.query.get_or_404(id)
    form = ExpenseForm(obj=expense)
    
    if form.validate_on_submit():
        expense.description = form.description.data
        expense.amount = form.amount.data
        expense.expense_type = form.expense_type.data
        expense.category = form.category.data
        expense.expense_date = form.expense_date.data or expense.expense_date
        expense.notes = form.notes.data
        db.session.commit()
        flash('تم تحديث المصروف بنجاح', 'success')
        return redirect(url_for('expenses'))
    
    return render_template('expenses/edit.html', form=form, title='تعديل المصروف')

@app.route('/expenses/<int:id>/delete', methods=['POST'])
@login_required
def delete_expense(id):
    """حذف مصروف"""
    if not current_user.is_admin():
        flash('ليس لديك صلاحية لحذف المصاريف', 'error')
        return redirect(url_for('expenses'))
    
    expense = Expense.query.get_or_404(id)
    description = expense.description
    db.session.delete(expense)
    db.session.commit()
    flash(f'تم حذف المصروف "{description}" بنجاح', 'success')
    return redirect(url_for('expenses'))

# ==================== النواقص ====================

@app.route('/shopping-list')
@login_required
@admin_required
def shopping_list():
    """صفحة النواقص"""
    # الحصول على قائمة النواقص المطلوبة
    needed_items = ShoppingList.query.filter_by(status='مطلوب').order_by(
        ShoppingList.priority.desc(), ShoppingList.created_at.desc()
    ).all()
    
    # الحصول على المنتجات التي نفدت أو قاربت على النفاد
    out_of_stock = Product.query.filter(Product.stock_quantity <= 0).all()
    low_stock = Product.query.filter(
        and_(Product.stock_quantity > 0, 
             Product.stock_quantity <= Product.min_stock_threshold)
    ).all()
    
    # حساب إجمالي التكلفة المتوقعة
    total_estimated_cost = sum(item.total_estimated_cost for item in needed_items if item.estimated_price)
    
    return render_template('shopping/list.html', 
                         needed_items=needed_items,
                         out_of_stock=out_of_stock,
                         low_stock=low_stock,
                         total_estimated_cost=total_estimated_cost)

@app.route('/shopping-list/add', methods=['GET', 'POST'])
@login_required
def add_shopping_item():
    """إضافة منتج لقائمة النواقص"""
    form = ShoppingListForm()
    if form.validate_on_submit():
        try:
            shopping_item = ShoppingList(
                item_name=form.item_name.data,
                quantity_needed=form.quantity_needed.data,
                unit_type=form.unit_type.data,
                estimated_price=form.estimated_price.data,
                priority=form.priority.data,
                category=form.category.data,
                supplier=form.supplier.data,
                notes=form.notes.data,
                user_id=current_user.id
            )
            db.session.add(shopping_item)
            db.session.commit()
            flash('تم إضافة المنتج لقائمة النواقص بنجاح', 'success')
            return redirect(url_for('shopping_list'))
        except Exception as e:
            db.session.rollback()
            flash(f'خطأ في إضافة المنتج: {str(e)}', 'error')
    
    return render_template('shopping/add.html', form=form)

@app.route('/shopping-list/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_shopping_item(id):
    """تعديل منتج في قائمة النواقص"""
    item = ShoppingList.query.get_or_404(id)
    form = ShoppingListForm(obj=item)
    
    if form.validate_on_submit():
        try:
            item.item_name = form.item_name.data
            item.quantity_needed = form.quantity_needed.data
            item.unit_type = form.unit_type.data
            item.estimated_price = form.estimated_price.data
            item.priority = form.priority.data
            item.category = form.category.data
            item.supplier = form.supplier.data
            item.notes = form.notes.data
            
            db.session.commit()
            flash('تم تحديث المنتج بنجاح', 'success')
            return redirect(url_for('shopping_list'))
        except Exception as e:
            db.session.rollback()
            flash(f'خطأ في تحديث المنتج: {str(e)}', 'error')
    
    return render_template('shopping/edit.html', form=form, item=item)

@app.route('/shopping-list/<int:id>/delete', methods=['POST'])
@login_required
def delete_shopping_item(id):
    """حذف منتج من قائمة النواقص"""
    try:
        item = ShoppingList.query.get_or_404(id)
        db.session.delete(item)
        db.session.commit()
        flash('تم حذف المنتج من قائمة النواقص', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'خطأ في حذف المنتج: {str(e)}', 'error')
    
    return redirect(url_for('shopping_list'))

@app.route('/shopping-list/<int:id>/mark-purchased', methods=['POST'])
@login_required
def mark_purchased(id):
    """تحديد منتج كمُشترى"""
    try:
        item = ShoppingList.query.get_or_404(id)
        item.status = 'تم الشراء'
        item.purchased_date = datetime.utcnow()
        db.session.commit()
        flash('تم تحديد المنتج كمُشترى', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'خطأ في تحديث حالة المنتج: {str(e)}', 'error')
    
    return redirect(url_for('shopping_list'))

@app.route('/shopping-list/add-low-stock/<int:product_id>')
@login_required
def add_low_stock_product(product_id):
    """إضافة منتج منخفض المخزون لقائمة النواقص"""
    product = Product.query.get_or_404(product_id)
    
    # التحقق من عدم وجود المنتج في القائمة بالفعل
    existing = ShoppingList.query.filter_by(
        item_name=product.name_ar, 
        status='مطلوب'
    ).first()
    
    if existing:
        flash('هذا المنتج موجود بالفعل في قائمة النواقص', 'warning')
    else:
        try:
            # تحديد الكمية المقترحة (الحد الأدنى - الكمية الحالية)
            suggested_quantity = max(product.min_stock_threshold - product.stock_quantity, 10)
            
            shopping_item = ShoppingList(
                item_name=product.name_ar,
                quantity_needed=suggested_quantity,
                unit_type=product.unit_type,
                estimated_price=product.wholesale_price,
                priority='high' if product.is_out_of_stock else 'medium',
                category=product.category.name_ar if product.category else None,
                notes=f'منتج {"نفد" if product.is_out_of_stock else "منخفض"} من المخزون - الكمية الحالية: {product.stock_quantity}',
                user_id=current_user.id
            )
            db.session.add(shopping_item)
            db.session.commit()
            flash('تم إضافة المنتج لقائمة النواقص', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'خطأ في إضافة المنتج: {str(e)}', 'error')
    
    return redirect(url_for('shopping_list'))

@app.route('/test-export')
@login_required
def test_export():
    """صفحة اختبار وظائف التصدير"""
    return send_from_directory('.', 'test_export.html')

@app.route('/stock/update', methods=['GET', 'POST'])
@login_required
def update_stock():
    if not current_user.is_admin():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('dashboard'))
    
    form = StockUpdateForm()
    if form.validate_on_submit():
        product = Product.query.get(form.product_id.data)
        product.stock_quantity += form.quantity.data
        db.session.commit()
        flash(f'تم تحديث مخزون {product.name_ar} بنجاح', 'success')
        return redirect(url_for('products'))
    
    return render_template('stock/update.html', form=form)

def create_sample_data():
    """Create sample data for testing"""
    # Create default system admin user (hidden)
    system_admin = User.query.filter_by(username='araby').first()
    if not system_admin:
        system_admin = User(username='araby', role='admin', is_system=True)
        system_admin.set_password('92321066')
        db.session.add(system_admin)
    
    # Create admin user
    admin = User.query.filter_by(username='admin').first()
    if not admin:
        admin = User(username='admin', role='admin')
        admin.set_password('admin123')
        db.session.add(admin)
    
    # Create seller user
    seller = User.query.filter_by(username='seller').first()
    if not seller:
        seller = User(username='seller', role='seller')
        seller.set_password('seller123')
        db.session.add(seller)
    
    # Create comprehensive categories for bookstore and school supplies
    if Category.query.count() == 0:
        categories = [
            # كتب ومراجع
            Category(name_ar='كتب أدبية وروايات', description_ar='الروايات والقصص والشعر والأدب العربي والعالمي'),
            Category(name_ar='كتب علمية وتقنية', description_ar='الكتب العلمية والتقنية والهندسية والطبية'),
            Category(name_ar='كتب دراسية ومناهج', description_ar='المناهج الدراسية والكتب الجامعية والمدرسية'),
            Category(name_ar='كتب الأطفال', description_ar='القصص والكتب التعليمية للأطفال والرسوم المتحركة'),
            Category(name_ar='كتب دينية وإسلامية', description_ar='القرآن الكريم والتفاسير والكتب الدينية'),
            Category(name_ar='مجلات وصحف', description_ar='المجلات العلمية والثقافية والصحف اليومية'),
            
            # أدوات الكتابة
            Category(name_ar='أقلام حبر وجاف', description_ar='أقلام الحبر الجاف والسائل بألوان مختلفة'),
            Category(name_ar='أقلام رصاص وملونة', description_ar='أقلام الرصاص العادية والملونة وأقلام التلوين'),
            Category(name_ar='أقلام ماركر وتحديد', description_ar='أقلام الماركر والهايلايتر وأقلام التحديد'),
            Category(name_ar='محايات وبراية', description_ar='المحايات البيضاء والملونة وبرايات الأقلام'),
            
            # دفاتر وكشاكيل
            Category(name_ar='كشاكيل ودفاتر خانات', description_ar='الكشاكيل المخططة والمربعة وذات الخانات'),
            Category(name_ar='دفاتر مسطرة وسادة', description_ar='الدفاتر المسطرة والسادة للكتابة'),
            Category(name_ar='دفاتر رسم وفنية', description_ar='دفاتر الرسم والأوراق الفنية للرسم والتلوين'),
            Category(name_ar='بلوكات وأوراق لاصقة', description_ar='البلوكات والملاحظات اللاصقة بأحجام مختلفة'),
            
            # مجلدات وحفظ
            Category(name_ar='مجلدات وحافظات', description_ar='المجلدات البلاستيكية والكرتونية لحفظ الأوراق'),
            Category(name_ar='أكياس وحافظات شفافة', description_ar='الأكياس الشفافة وحافظات الأوراق البلاستيكية'),
            Category(name_ar='ملفات ومنظمات', description_ar='الملفات المعدنية والبلاستيكية ومنظمات المكتب'),
            
            # أدوات هندسية ورياضية
            Category(name_ar='مساطر وأدوات قياس', description_ar='المساطر والزوايا وأدوات القياس الهندسية'),
            Category(name_ar='برجل وكوسات هندسية', description_ar='البرجل والكوسات وأدوات الرسم الهندسي'),
            Category(name_ar='آلات حاسبة', description_ar='الآلات الحاسبة العلمية والعادية'),
            Category(name_ar='أدوات رياضية تعليمية', description_ar='النماذج الهندسية والأدوات التعليمية للرياضيات'),
            
            # أدوات فنية وإبداعية
            Category(name_ar='ألوان وطلاء', description_ar='الألوان المائية والزيتية وألوان الأطفال'),
            Category(name_ar='فرش ولوازم الرسم', description_ar='فرش الرسم وإسفنج التلوين واللوازم الفنية'),
            Category(name_ar='ورق ملون وكارتون', description_ar='الأوراق الملونة والكارتون المقوى للأعمال الفنية'),
            Category(name_ar='لاصق وصمغ', description_ar='أنواع اللاصق والصمغ والشريط اللاصق'),
            
            # أدوات مكتبية عامة
            Category(name_ar='مقصات وقطاعات', description_ar='المقصات بأحجام مختلفة وقطاعات الورق'),
            Category(name_ar='دباسة وخرامة', description_ar='الدباسات والخرامات ولوازم التثبيت'),
            Category(name_ar='مشابك ودبابيس', description_ar='مشابك الورق والدبابيس وأدوات التثبيت'),
            Category(name_ar='لوازم المكتب المختلفة', description_ar='منظمات المكتب وحوامل الأقلام والأدوات المكتبية'),
            
            # حقائب وأدوات حمل
            Category(name_ar='حقائب مدرسية', description_ar='الحقائب المدرسية بأحجام وأشكال مختلفة'),
            Category(name_ar='مقلمات وحافظات أقلام', description_ar='المقلمات وحافظات الأقلام والأدوات'),
            Category(name_ar='شنط لابتوب ووثائق', description_ar='حقائب اللابتوب وحافظات الوثائق والملفات'),
            
            # لوازم إلكترونية ومكتبية
            Category(name_ar='بطاريات وشواحن', description_ar='البطاريات والشواحن للأجهزة الإلكترونية'),
            Category(name_ar='فلاش ميموري وأقراص', description_ar='فلاش ميموري وأقراص التخزين والـ CD/DVD'),
            Category(name_ar='لوازم الكمبيوتر', description_ar='ماوس وكيبورد وإكسسوارات الكمبيوتر'),
            
            # متنوعات
            Category(name_ar='هدايا ولعب تعليمية', description_ar='الهدايا والألعاب التعليمية والترفيهية'),
            Category(name_ar='لوازم التغليف', description_ar='أكياس الهدايا وورق التغليف والشرائط'),
            Category(name_ar='منتجات موسمية', description_ar='المنتجات الخاصة بالمواسم والمناسبات المختلفة'),
        ]
        
        for category in categories:
            db.session.add(category)
    
    db.session.commit()
    
    # Create diverse sample products for different categories
    if Product.query.count() == 0:
        products = [
            # كتب أدبية وروايات
            Product(name_ar='رواية مئة عام من العزلة', category_id=1, wholesale_price=65.00, retail_price=85.00, price=85.00, stock_quantity=25, min_stock_threshold=5, description_ar='رواية للكاتب غابرييل غارسيا ماركيز'),
            Product(name_ar='ديوان محمود درويش', category_id=1, wholesale_price=35.00, retail_price=45.00, price=45.00, stock_quantity=30, min_stock_threshold=8, description_ar='مجموعة قصائد للشاعر محمود درويش'),
            Product(name_ar='رواية مدن الملح', category_id=1, wholesale_price=75.00, retail_price=95.00, price=95.00, stock_quantity=20, min_stock_threshold=5, description_ar='رواية عبد الرحمن منيف'),
            
            # كتب علمية وتقنية
            Product(name_ar='كتاب البرمجة بالبايثون', category_id=2, wholesale_price=95.00, retail_price=120.00, price=120.00, stock_quantity=15, min_stock_threshold=3, description_ar='دليل شامل لتعلم البرمجة'),
            Product(name_ar='أساسيات الرياضيات', category_id=2, wholesale_price=60.00, retail_price=80.00, price=80.00, stock_quantity=35, min_stock_threshold=10, description_ar='كتاب تعليمي في الرياضيات'),
            
            # كتب دراسية ومناهج
            Product(name_ar='منهج الرياضيات - الصف الثالث الثانوي', category_id=3, wholesale_price=42.00, retail_price=55.00, price=55.00, stock_quantity=50, min_stock_threshold=15, description_ar='منهج وزارة التربية والتعليم'),
            Product(name_ar='كتاب الفيزياء - الصف الثاني الثانوي', category_id=3, wholesale_price=36.00, retail_price=48.00, price=48.00, stock_quantity=40, min_stock_threshold=12, description_ar='منهج معتمد'),
            
            # أقلام حبر وجاف
            Product(name_ar='قلم حبر جاف أزرق', category_id=7, wholesale_price=2.50, retail_price=3.50, price=3.50, stock_quantity=200, min_stock_threshold=50, unit_type='جزئي', unit_description='قلم حبر جاف لون أزرق'),
            Product(name_ar='علبة أقلام حبر ملونة (12 قلم)', category_id=7, wholesale_price=25.00, retail_price=35.00, price=35.00, stock_quantity=80, min_stock_threshold=20, description_ar='مجموعة أقلام ملونة'),
            Product(name_ar='قلم حبر أحمر', category_id=7, wholesale_price=2.50, retail_price=3.50, price=3.50, stock_quantity=150, min_stock_threshold=40, unit_type='جزئي', unit_description='قلم'),
            Product(name_ar='قلم حبر أسود', category_id=7, wholesale_price=2.50, retail_price=3.50, price=3.50, stock_quantity=180, min_stock_threshold=45, unit_type='جزئي', unit_description='قلم'),
            
            # أقلام رصاص وملونة
            Product(name_ar='قلم رصاص HB', category_id=8, price=2.00, stock_quantity=250, min_stock_threshold=60, unit_type='جزئي', unit_description='قلم'),
            Product(name_ar='علبة أقلام ملونة خشبية (24 لون)', category_id=8, price=45.00, stock_quantity=60, min_stock_threshold=15, description_ar='أقلام تلوين خشبية عالية الجودة'),
            Product(name_ar='قلم رصاص 2B للرسم', category_id=8, price=4.00, stock_quantity=100, min_stock_threshold=25, unit_type='جزئي', unit_description='قلم'),
            
            # كشاكيل ودفاتر
            Product(name_ar='كشكول 100 ورقة مخطط', category_id=11, price=15.00, stock_quantity=120, min_stock_threshold=30, description_ar='كشكول مخطط للكتابة'),
            Product(name_ar='كشكول 200 ورقة مربعات', category_id=11, price=25.00, stock_quantity=90, min_stock_threshold=25, description_ar='كشكول مربعات للرياضيات'),
            Product(name_ar='دفتر 48 ورقة سادة', category_id=12, price=8.00, stock_quantity=200, min_stock_threshold=50, description_ar='دفتر سادة للكتابة الحرة'),
            Product(name_ar='كشكول سبايرال A4', category_id=11, price=28.00, stock_quantity=75, min_stock_threshold=20, description_ar='كشكول سبايرال حجم A4'),
            
            # محايات وبراية
            Product(name_ar='محاية بيضاء كبيرة', category_id=10, price=2.50, stock_quantity=300, min_stock_threshold=75, unit_type='جزئي', unit_description='قطعة'),
            Product(name_ar='براية معدنية', category_id=10, price=5.00, stock_quantity=150, min_stock_threshold=35, unit_type='جزئي', unit_description='قطعة'),
            Product(name_ar='محاية ملونة صغيرة', category_id=10, price=1.50, stock_quantity=400, min_stock_threshold=100, unit_type='جزئي', unit_description='قطعة'),
            
            # مجلدات وحافظات
            Product(name_ar='مجلد بلاستيكي A4', category_id=15, price=12.00, stock_quantity=80, min_stock_threshold=20, description_ar='مجلد شفاف لحفظ الأوراق'),
            Product(name_ar='حافظة أوراق شفافة (10 قطع)', category_id=16, price=8.00, stock_quantity=100, min_stock_threshold=25, description_ar='حافظات شفافة مثقبة'),
            Product(name_ar='مجلد كرتوني ملون', category_id=15, price=18.00, stock_quantity=60, min_stock_threshold=15, description_ar='مجلد كرتوني بألوان مختلفة'),
            
            # أدوات هندسية
            Product(name_ar='مسطرة 30 سم شفافة', category_id=18, price=8.00, stock_quantity=120, min_stock_threshold=30, unit_type='جزئي', unit_description='قطعة'),
            Product(name_ar='مجموعة أدوات هندسية (برجل + مسطرة + زاوية)', category_id=19, price=35.00, stock_quantity=40, min_stock_threshold=10, description_ar='مجموعة كاملة للرسم الهندسي'),
            Product(name_ar='آلة حاسبة علمية', category_id=20, price=85.00, stock_quantity=25, min_stock_threshold=5, description_ar='آلة حاسبة للطلاب والمهندسين'),
            
            # ألوان وطلاء
            Product(name_ar='علبة ألوان مائية (12 لون)', category_id=22, price=25.00, stock_quantity=50, min_stock_threshold=12, description_ar='ألوان مائية للرسم والفن'),
            Product(name_ar='ألوان فلوماستر (18 لون)', category_id=22, price=30.00, stock_quantity=70, min_stock_threshold=18, description_ar='أقلام ألوان فلوماستر'),
            
            # أدوات مكتبية
            Product(name_ar='مقص متوسط الحجم', category_id=26, price=12.00, stock_quantity=90, min_stock_threshold=20, unit_type='جزئي', unit_description='قطعة'),
            Product(name_ar='دباسة صغيرة + علبة دبابيس', category_id=27, price=15.00, stock_quantity=60, min_stock_threshold=15, description_ar='دباسة مع دبابيس للاستعمال المكتبي'),
            Product(name_ar='صمغ أبيض 50 مل', category_id=25, price=6.00, stock_quantity=150, min_stock_threshold=35, unit_type='جزئي', unit_description='أنبوبة'),
            
            # حقائب ومقلمات
            Product(name_ar='حقيبة مدرسية متوسطة', category_id=30, price=95.00, stock_quantity=30, min_stock_threshold=8, description_ar='حقيبة مدرسية بجيوب متعددة'),
            Product(name_ar='مقلمة بسحاب', category_id=31, price=18.00, stock_quantity=85, min_stock_threshold=20, description_ar='مقلمة لحفظ الأقلام والأدوات'),
            Product(name_ar='حقيبة لابتوب 15 بوصة', category_id=32, price=150.00, stock_quantity=20, min_stock_threshold=5, description_ar='حقيبة واقية للابتوب'),
            
            # منتجات إلكترونية
            Product(name_ar='فلاش ميموري 16 جيجا', category_id=34, price=45.00, stock_quantity=40, min_stock_threshold=10, unit_type='جزئي', unit_description='قطعة'),
            Product(name_ar='بطاريات AA (4 قطع)', category_id=33, price=12.00, stock_quantity=100, min_stock_threshold=25, description_ar='بطاريات قلوية عالية الجودة'),
            
            # هدايا ومتنوعات
            Product(name_ar='لعبة تعليمية للأطفال', category_id=36, price=35.00, stock_quantity=25, min_stock_threshold=5, description_ar='لعبة تعليمية تفاعلية'),
            Product(name_ar='كيس هدية ملون', category_id=37, price=3.00, stock_quantity=200, min_stock_threshold=50, unit_type='جزئي', unit_description='كيس'),
            
            # مجلات
            Product(name_ar='مجلة العلوم العدد الجديد', category_id=6, price=15.00, stock_quantity=35, min_stock_threshold=10, description_ar='مجلة شهرية علمية'),
            Product(name_ar='مجلة الأطفال المصورة', category_id=6, price=12.00, stock_quantity=50, min_stock_threshold=15, description_ar='مجلة أسبوعية للأطفال'),
        ]
        
        for product in products:
            db.session.add(product)
    
    db.session.commit()
    
    # Create sample customers
    if Customer.query.count() == 0:
        customers = [
            Customer(name='أحمد محمد علي', phone='01234567890', address='القاهرة - مصر الجديدة', notes='عميل دائم'),
            Customer(name='فاطمة أحمد حسن', phone='01098765432', address='الجيزة - الدقي'),
            Customer(name='محمد حسن إبراهيم', phone='01555666777', address='الإسكندرية - سيدي جابر', notes='عميل مميز'),
            Customer(name='سارة علي محمود', phone='01122334455', address='القاهرة - المعادي'),
            Customer(name='عمر خالد أحمد', phone='01199887766', address='الجيزة - المهندسين', notes='عميل جديد'),
        ]
        
        for customer in customers:
            db.session.add(customer)
        
        db.session.commit()

@app.route('/simple-export-test')
@login_required
def simple_export_test():
    """صفحة اختبار بسيطة للتصدير"""
    return """
    <!DOCTYPE html>
    <html>
    <head>
        <title>اختبار التصدير</title>
        <meta charset="UTF-8">
    </head>
    <body>
        <h1>اختبار وظيفة التصدير</h1>
        <button onclick="testExport()">اختبار تصدير Excel</button>
        <div id="result"></div>
        
        <script src="https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js"></script>
        <script>
            function testExport() {
                console.log('Testing export...');
                document.getElementById('result').innerHTML = 'جاري الاختبار...';
                
                if (typeof XLSX === 'undefined') {
                    document.getElementById('result').innerHTML = 'خطأ: مكتبة Excel غير محملة';
                    return;
                }
                
                try {
                    const testData = [
                        ['اختبار التصدير'],
                        ['المنتج', 'الكمية', 'السعر'],
                        ['منتج تجريبي', 10, 100]
                    ];
                    
                    const wb = XLSX.utils.book_new();
                    const ws = XLSX.utils.aoa_to_sheet(testData);
                    XLSX.utils.book_append_sheet(wb, ws, 'اختبار');
                    
                    const fileName = 'test_export_' + new Date().getTime() + '.xlsx';
                    XLSX.writeFile(wb, fileName);
                    
                    document.getElementById('result').innerHTML = 'نجح التصدير! ✅';
                } catch (error) {
                    document.getElementById('result').innerHTML = 'فشل التصدير: ' + error.message;
                }
            }
        </script>
    </body>
    </html>
    """

@app.route('/debug-export')
def debug_export():
    """صفحة تشخيص مشاكل التصدير - بدون تسجيل دخول للتشخيص"""
    return '''
    <!DOCTYPE html>
    <html lang="ar" dir="rtl">
    <head>
        <meta charset="UTF-8">
        <title>تشخيص مشكلة التصدير</title>
        <style>
            body { font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }
            .container { max-width: 800px; margin: 0 auto; }
            .step { background: white; padding: 20px; margin: 15px 0; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
            .success { border-left: 5px solid #28a745; }
            .error { border-left: 5px solid #dc3545; }
            .warning { border-left: 5px solid #ffc107; }
            button { 
                background: #007bff; color: white; border: none; 
                padding: 12px 24px; margin: 8px; 
                border-radius: 5px; cursor: pointer; font-size: 16px;
            }
            button:hover { background: #0056b3; }
            pre { background: #f8f9fa; padding: 15px; border-radius: 5px; overflow-x: auto; max-height: 300px; }
            .result { margin-top: 15px; padding: 10px; border-radius: 5px; }
            h1 { color: #333; text-align: center; }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>🔍 تشخيص مشكلة التصدير</h1>
            <p style="text-align: center; color: #666;">هذه الأداة ستساعدك في تحديد سبب عدم عمل التصدير</p>
            
            <div class="step">
                <h3>الخطوة 1: فحص المكتبات المطلوبة</h3>
                <button onclick="checkLibraries()">فحص المكتبات</button>
                <div id="libraries-result" class="result"></div>
            </div>
            
            <div class="step">
                <h3>الخطوة 2: اختبار تصدير Excel بسيط</h3>
                <button onclick="simpleExcelTest()">اختبار Excel</button>
                <div id="excel-result" class="result"></div>
            </div>
            
            <div class="step">
                <h3>الخطوة 3: اختبار تحميل ملف نصي</h3>
                <button onclick="simpleDownloadTest()">اختبار التحميل</button>
                <div id="download-result" class="result"></div>
            </div>
            
            <div class="step">
                <h3>الخطوة 4: معلومات المتصفح</h3>
                <button onclick="browserInfo()">معلومات المتصفح</button>
                <div id="browser-result" class="result"></div>
            </div>
            
            <div class="step">
                <h3>سجل العمليات والأخطاء</h3>
                <pre id="debug-log">انتظار بدء التشخيص...</pre>
                <button onclick="clearLog()">مسح السجل</button>
            </div>
        </div>

        <script src="https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js"></script>
        
        <script>
            let logs = [];
            
            function addLog(message) {
                const time = new Date().toLocaleTimeString();
                const logEntry = `[${time}] ${message}`;
                logs.push(logEntry);
                document.getElementById("debug-log").textContent = logs.join("\\n");
                console.log(logEntry);
            }
            
            function clearLog() {
                logs = [];
                document.getElementById("debug-log").textContent = "تم مسح السجل...";
            }
            
            function setResult(elementId, content, type = "info") {
                const element = document.getElementById(elementId);
                element.innerHTML = content;
                element.className = `result ${type}`;
            }
            
            function checkLibraries() {
                addLog("🔍 بدء فحص المكتبات...");
                
                let results = "<h4>نتائج الفحص:</h4>";
                let allOk = true;
                
                if (typeof XLSX !== "undefined") {
                    results += "✅ مكتبة XLSX: محملة بنجاح<br>";
                    results += `📋 الإصدار: ${XLSX.version || "غير محدد"}<br>`;
                    addLog("✅ XLSX library loaded");
                } else {
                    results += "❌ مكتبة XLSX: غير محملة<br>";
                    allOk = false;
                    addLog("❌ XLSX library missing");
                }
                
                if (typeof Blob !== "undefined") {
                    results += "✅ Blob API: متوفر<br>";
                } else {
                    results += "❌ Blob API: غير متوفر<br>";
                    allOk = false;
                }
                
                if (typeof URL.createObjectURL === "function") {
                    results += "✅ URL API: متوفر<br>";
                } else {
                    results += "❌ URL API: غير متوفر<br>";
                    allOk = false;
                }
                
                const downloadSupported = "download" in document.createElement("a");
                if (downloadSupported) {
                    results += "✅ Download Attribute: مدعوم<br>";
                } else {
                    results += "❌ Download Attribute: غير مدعوم<br>";
                    allOk = false;
                }
                
                setResult("libraries-result", results, allOk ? "success" : "error");
                addLog(`📊 فحص المكتبات: ${allOk ? "نجح" : "فشل"}`);
            }
            
            function simpleExcelTest() {
                addLog("📊 بدء اختبار Excel...");
                
                if (typeof XLSX === "undefined") {
                    setResult("excel-result", "❌ لا يمكن الاختبار - مكتبة XLSX غير محملة", "error");
                    addLog("❌ Excel test failed - no XLSX");
                    return;
                }
                
                try {
                    addLog("📝 إنشاء بيانات اختبار...");
                    const data = [
                        ["تجربة التصدير"],
                        ["العنصر", "القيمة", "التاريخ"],
                        ["اختبار 1", 100, new Date().toLocaleDateString()],
                        ["اختبار 2", 200, new Date().toLocaleDateString()],
                        ["المجموع", 300, ""]
                    ];
                    
                    addLog("🔧 إنشاء ملف Excel...");
                    const wb = XLSX.utils.book_new();
                    const ws = XLSX.utils.aoa_to_sheet(data);
                    XLSX.utils.book_append_sheet(wb, ws, "اختبار");
                    
                    addLog("💾 محاولة حفظ الملف...");
                    const fileName = `excel_test_${Date.now()}.xlsx`;
                    
                    XLSX.writeFile(wb, fileName);
                    
                    setResult("excel-result", "✅ تم إنشاء ملف Excel بنجاح! تحقق من مجلد التحميل.", "success");
                    addLog("✅ Excel file created successfully");
                    
                } catch (error) {
                    setResult("excel-result", `❌ خطأ في إنشاء Excel: ${error.message}`, "error");
                    addLog(`❌ Excel error: ${error.message}`);
                    console.error("Excel Error Details:", error);
                }
            }
            
            function simpleDownloadTest() {
                addLog("⬇️ بدء اختبار التحميل...");
                
                try {
                    const content = "اختبار التحميل\\nهذا ملف نصي تجريبي\\nالوقت: " + new Date().toLocaleString();
                    const blob = new Blob([content], { type: "text/plain;charset=utf-8" });
                    const url = URL.createObjectURL(blob);
                    
                    const link = document.createElement("a");
                    link.href = url;
                    link.download = `download_test_${Date.now()}.txt`;
                    link.style.display = "none";
                    
                    document.body.appendChild(link);
                    link.click();
                    document.body.removeChild(link);
                    
                    URL.revokeObjectURL(url);
                    
                    setResult("download-result", "✅ تم اختبار التحميل! تحقق من مجلد التحميل.", "success");
                    addLog("✅ Download test completed");
                    
                } catch (error) {
                    setResult("download-result", `❌ خطأ في التحميل: ${error.message}`, "error");
                    addLog(`❌ Download error: ${error.message}`);
                }
            }
            
            function browserInfo() {
                addLog("🌐 جمع معلومات المتصفح...");
                
                let info = "<h4>معلومات المتصفح:</h4>";
                info += `<strong>المتصفح:</strong> ${navigator.userAgent}<br><br>`;
                info += `<strong>النظام:</strong> ${navigator.platform}<br>`;
                info += `<strong>اللغة:</strong> ${navigator.language}<br>`;
                info += `<strong>Cookies مفعلة:</strong> ${navigator.cookieEnabled ? "نعم" : "لا"}<br>`;
                
                info += "<br><h5>🔧 نصائح لحل المشكلة:</h5>";
                info += "• تأكد أن مجلد التحميل محدد بشكل صحيح<br>";
                info += "• تحقق من إعدادات حظر النوافذ المنبثقة<br>";
                info += "• جرب متصفح آخر للمقارنة<br>";
                info += "• تأكد أن الإنترنت متصل لتحميل المكتبات<br>";
                info += "• اضغط F12 وافحص تبويب Console للأخطاء<br>";
                
                setResult("browser-result", info, "info");
                addLog("📋 Browser info collected");
            }
            
            window.addEventListener("load", function() {
                addLog("🚀 تم تحميل صفحة التشخيص");
                setTimeout(() => {
                    addLog("🔄 بدء الفحص التلقائي...");
                    checkLibraries();
                }, 500);
            });
        </script>
    </body>
    </html>
    '''

@app.route('/api/export/full-database')
@login_required
def api_export_full_database():
    if current_user.role not in ['admin', 'seller']:
        abort(403)
    
    try:
        data = {}
        
        # Basic report summary
        data['report_summary'] = [{
            'تاريخ التقرير': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'عدد المنتجات': Product.query.count(),
            'عدد العملاء': Customer.query.count(),
            'عدد المبيعات': Sale.query.count(),
            'عدد الفئات': Category.query.count()
        }]
        
        # Products - basic info only
        products_data = []
        for p in Product.query.all():
            products_data.append({
                'رقم المنتج': p.id,
                'اسم المنتج': p.name_ar,
                'سعر الجملة': float(p.wholesale_price),
                'سعر البيع': float(p.retail_price),
                'الكمية': float(p.stock_quantity),
                'نوع الوحدة': p.unit_type,
                'الحد الأدنى': float(p.min_stock_threshold)
            })
        data['products'] = products_data
        
        # Categories
        categories_data = []
        for c in Category.query.all():
            categories_data.append({
                'رقم الفئة': c.id,
                'اسم الفئة': c.name_ar,
                'الوصف': c.description_ar or ''
            })
        data['categories'] = categories_data
        
        # Customers
        customers_data = []
        for c in Customer.query.all():
            customers_data.append({
                'رقم العميل': c.id,
                'اسم العميل': c.name,
                'الهاتف': c.phone or '',
                'العنوان': c.address or '',
                'ملاحظات': c.notes or ''
            })
        data['customers'] = customers_data
        
        # Sales - basic info only
        sales_data = []
        for s in Sale.query.all():
            sales_data.append({
                'رقم البيع': s.id,
                'تاريخ البيع': str(s.sale_date),
                'المبلغ': float(s.total_amount),
                'نوع الدفع': s.payment_type,
                'حالة الدفع': s.payment_status,
                'ملاحظات': s.notes or ''
            })
        data['sales'] = sales_data
        
        # Sale Items
        sale_items_data = []
        for si in SaleItem.query.all():
            sale_items_data.append({
                'رقم البيع': si.sale_id,
                'رقم المنتج': si.product_id,
                'الكمية': float(si.quantity),
                'سعر الوحدة': float(si.unit_price),
                'الإجمالي': float(si.total_price)
            })
        data['sale_items'] = sale_items_data
        
        # Payments
        payments_data = []
        for p in Payment.query.all():
            payments_data.append({
                'رقم الدفعة': p.id,
                'رقم البيع': p.sale_id,
                'المبلغ': float(p.amount),
                'تاريخ الدفع': str(p.payment_date),
                'طريقة الدفع': p.payment_method,
                'ملاحظات': p.notes or ''
            })
        data['payments'] = payments_data
        
        # Expenses
        expenses_data = []
        for e in Expense.query.all():
            expenses_data.append({
                'رقم المصروف': e.id,
                'الوصف': e.description,
                'المبلغ': float(e.amount),
                'النوع': e.expense_type,
                'التاريخ': str(e.expense_date),
                'ملاحظات': e.notes or ''
            })
        data['expenses'] = expenses_data
        
        # Users (admin only)
        if current_user.role == 'admin':
            users_data = []
            for u in User.query.all():
                users_data.append({
                    'رقم المستخدم': u.id,
                    'اسم المستخدم': u.username,
                    'الدور': u.role,
                    'تاريخ الإنشاء': str(u.created_at)
                })
            data['users'] = users_data
        
        return jsonify(data)
        
    except Exception as e:
        # Log the actual error for debugging
        print(f"Export error: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return jsonify({
            'error': 'حدث خطأ في تصدير البيانات',
            'message': str(e)
        }), 500

@app.route('/api/export/test-database')
@login_required
def api_test_database_export():
    """Simple test endpoint for database export"""
    try:
        data = {
            'status': 'success',
            'message': 'Test endpoint working',
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'products_count': Product.query.count(),
            'customers_count': Customer.query.count(),
            'sales_count': Sale.query.count()
        }
        return jsonify(data)
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

# ================================
# Routes للمرتجعات
# ================================

@app.route('/returns')
@login_required
@seller_or_admin_required
def returns():
    """عرض قائمة المرتجعات"""
    page = request.args.get('page', 1, type=int)
    status_filter = request.args.get('status', 'all')
    
    query = Return.query.join(Sale).join(User)
    
    # تطبيق فلتر الحالة
    if status_filter != 'all':
        query = query.filter(Return.status == status_filter)
    
    # فلترة حسب صلاحية المستخدم
    if not current_user.is_admin():
        query = query.filter(Return.user_id == current_user.id)
    
    returns = query.order_by(Return.return_date.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    
    return render_template('returns/list.html', returns=returns, status_filter=status_filter)

@app.route('/returns/new/<int:sale_id>')
@login_required
@seller_or_admin_required
def new_return(sale_id):
    """إنشاء مرتجع جديد لبيعة معينة"""
    sale = Sale.query.get_or_404(sale_id)
    
    # التحقق من الصلاحية
    if not current_user.is_admin() and sale.user_id != current_user.id:
        flash('ليس لديك صلاحية لإنشاء مرتجع لهذا البيع', 'error')
        return redirect(url_for('sales'))
    
    return render_template('returns/new.html', sale=sale)

@app.route('/api/returns', methods=['POST'])
@login_required
@seller_or_admin_required
def api_create_return():
    """API لإنشاء مرتجع جديد"""
    try:
        data = request.get_json()
        sale_id = data.get('sale_id')
        reason = data.get('reason', '')
        refund_method = data.get('refund_method', 'نقدي')
        notes = data.get('notes', '')
        items = data.get('items', [])
        
        if not sale_id or not reason or not items:
            return jsonify({'success': False, 'message': 'بيانات غير مكتملة'}), 400
        
        # التحقق من البيع
        sale = Sale.query.get_or_404(sale_id)
        if not current_user.is_admin() and sale.user_id != current_user.id:
            return jsonify({'success': False, 'message': 'ليس لديك صلاحية لإنشاء مرتجع لهذا البيع'}), 403
        
        # إنشاء المرتجع
        return_obj = Return(
            sale_id=sale_id,
            customer_id=sale.customer_id,
            reason=reason,
            refund_method=refund_method,
            notes=notes,
            user_id=current_user.id,
            total_amount=0  # سيتم حسابه لاحقاً
        )
        
        db.session.add(return_obj)
        db.session.flush()  # للحصول على ID المرتجع
        
        total_amount = 0
        
        # إضافة أصناف المرتجع
        for item_data in items:
            sale_item_id = item_data.get('sale_item_id')
            quantity_returned = float(item_data.get('quantity_returned', 0))
            condition = item_data.get('condition', 'جيد')
            item_notes = item_data.get('notes', '')
            
            # التحقق من صنف البيع
            sale_item = SaleItem.query.get_or_404(sale_item_id)
            if sale_item.sale_id != sale_id:
                return jsonify({'success': False, 'message': 'خطأ في بيانات الصنف'}), 400
            
            # التحقق من الكمية
            if quantity_returned <= 0 or quantity_returned > sale_item.quantity:
                return jsonify({'success': False, 'message': f'كمية خاطئة للصنف {sale_item.product.name_ar}'}), 400
            
            # إنشاء صنف المرتجع
            return_item = ReturnItem(
                return_id=return_obj.id,
                sale_item_id=sale_item_id,
                product_id=sale_item.product_id,
                quantity_returned=quantity_returned,
                original_quantity=sale_item.quantity,
                unit_price=sale_item.unit_price,
                condition=condition,
                notes=item_notes
            )
            
            db.session.add(return_item)
            total_amount += return_item.total_refund
        
        # تحديث إجمالي المرتجع
        return_obj.total_amount = total_amount
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'تم إنشاء المرتجع بنجاح',
            'return_id': return_obj.id
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/returns/<int:return_id>')
@login_required
def api_return_details(return_id):
    """API لعرض تفاصيل المرتجع"""
    return_obj = Return.query.get_or_404(return_id)
    
    # التحقق من الصلاحية
    if not current_user.is_admin() and return_obj.user_id != current_user.id:
        return jsonify({'error': 'ليس لديك صلاحية لعرض هذا المرتجع'}), 403
    
    # تفاصيل أصناف المرتجع
    items = []
    for item in return_obj.return_items:
        items.append({
            'product_name': item.product.name_ar,
            'quantity_returned': float(item.quantity_returned),
            'original_quantity': float(item.original_quantity),
            'unit_price': float(item.unit_price),
            'total_refund': float(item.total_refund),
            'condition': item.condition,
            'condition_ar': item.condition_ar,
            'notes': item.notes or ''
        })
    
    # تحويل التاريخ إلى توقيت مصر
    egypt_time = get_egypt_time(return_obj.return_date)
    
    return jsonify({
        'id': return_obj.id,
        'sale_id': return_obj.sale_id,
        'customer_name': return_obj.customer.name if return_obj.customer else 'زبون نقدي',
        'total_amount': float(return_obj.total_amount),
        'return_date': egypt_time.strftime('%d/%m/%Y'),
        'return_time': egypt_time.strftime('%I:%M:%S %p').replace('AM', 'ص').replace('PM', 'م'),
        'reason': return_obj.reason,
        'status': return_obj.status,
        'status_ar': return_obj.status_ar,
        'refund_method': return_obj.refund_method,
        'notes': return_obj.notes or '',
        'user_name': return_obj.user.username,
        'processor_name': return_obj.processor.username if return_obj.processor else '',
        'processed_date': return_obj.processed_date.strftime('%d/%m/%Y') if return_obj.processed_date else '',
        'items': items
    })

@app.route('/api/returns/<int:return_id>/process', methods=['POST'])
@login_required
@admin_required
def api_process_return(return_id):
    """API لمعالجة المرتجع (قبول/رفض)"""
    try:
        data = request.get_json()
        action = data.get('action')  # 'approve' or 'reject'
        notes = data.get('notes', '')
        
        if action not in ['approve', 'reject']:
            return jsonify({'success': False, 'message': 'إجراء غير صحيح'}), 400
        
        return_obj = Return.query.get_or_404(return_id)
        
        if not return_obj.can_be_processed:
            return jsonify({'success': False, 'message': 'لا يمكن معالجة هذا المرتجع'}), 400
        
        # تحديث حالة المرتجع
        return_obj.status = 'approved' if action == 'approve' else 'rejected'
        return_obj.processed_by = current_user.id
        return_obj.processed_date = datetime.utcnow()
        if notes:
            return_obj.notes = (return_obj.notes or '') + f'\n\nملاحظات المعالجة: {notes}'
        
        # إذا تم قبول المرتجع، تحديث المخزون
        if action == 'approve':
            for return_item in return_obj.return_items:
                # إضافة الكمية المرتجعة إلى المخزون إذا كانت في حالة جيدة
                if return_item.condition in ['جيد', 'good']:
                    product = return_item.product
                    product.stock_quantity += return_item.quantity_returned
        
        db.session.commit()
        
        status_message = 'تم قبول المرتجع وإضافة الكمية للمخزون' if action == 'approve' else 'تم رفض المرتجع'
        
        return jsonify({
            'success': True,
            'message': status_message
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/sale/<int:sale_id>/items')
@login_required
@seller_or_admin_required
def api_sale_items(sale_id):
    """API لعرض أصناف البيع للمرتجع"""
    sale = Sale.query.get_or_404(sale_id)
    
    # التحقق من الصلاحية
    if not current_user.is_admin() and sale.user_id != current_user.id:
        return jsonify({'error': 'ليس لديك صلاحية لعرض هذا البيع'}), 403
    
    items = []
    for item in sale.sale_items:
        # حساب الكمية المرتجعة سابقاً
        returned_quantity = sum(
            ri.quantity_returned for ri in item.return_items 
            if ri.return_ref.status == 'approved'
        )
        
        available_quantity = item.quantity - returned_quantity
        
        if available_quantity > 0:  # فقط الأصناف التي يمكن إرجاعها
            items.append({
                'id': item.id,
                'product_name': item.product.name_ar,
                'quantity': float(item.quantity),
                'returned_quantity': float(returned_quantity),
                'available_quantity': float(available_quantity),
                'unit_price': float(item.unit_price),
                'total_price': float(item.total_price),
                'unit_type': item.product.unit_type
            })
    
    return jsonify(items)

@app.route('/api/products/excel-template')
@login_required
@admin_required
def api_products_excel_template():
    """تحميل نموذج Excel للمنتجات"""
    try:
        # إنشاء workbook جديد
        wb = Workbook()
        ws = wb.active
        ws.title = "المنتجات"
        
        # العناوين
        headers = [
            'اسم المنتج', 'وصف المنتج', 'الفئة', 'سعر الجملة', 
            'سعر البيع', 'الكمية', 'الحد الأدنى للمخزون', 'نوع الوحدة', 'وصف الوحدة'
        ]
        
        # إضافة العناوين
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # إضافة بيانات تجريبية
        sample_data = [
            ['منتج تجريبي 1', 'وصف المنتج الأول', 'قرطاسية', 10.50, 15.00, 100, 10, 'كامل', 'قطعة'],
            ['منتج تجريبي 2', 'وصف المنتج الثاني', 'كتب', 25.00, 35.00, 50, 5, 'كامل', 'كتاب']
        ]
        
        for row_num, row_data in enumerate(sample_data, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # ضبط عرض الأعمدة
        column_widths = [20, 25, 15, 12, 12, 10, 18, 12, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
        
        # حفظ في الذاكرة
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=products_template.xlsx'
        
        return response
        
    except Exception as e:
        app.logger.error(f"Error creating Excel template: {str(e)}")
        return jsonify({'error': 'حدث خطأ في إنشاء النموذج'}), 500

@app.route('/api/products/import-excel', methods=['POST'])
@login_required
@admin_required
def api_products_import_excel():
    """استيراد المنتجات من ملف Excel"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'لم يتم العثور على ملف'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'لم يتم اختيار ملف'}), 400
        
        # التحقق من نوع الملف
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': 'نوع الملف غير مدعوم. يرجى استخدام ملف Excel'}), 400
        
        # التحقق من حجم الملف (5 ميجابايت)
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)
        
        if file_size > 5 * 1024 * 1024:  # 5MB
            return jsonify({'success': False, 'message': 'حجم الملف كبير جداً. الحد الأقصى 5 ميجابايت'}), 400
        
        update_existing = request.form.get('update_existing') == 'true'
        
        # قراءة ملف Excel باستخدام openpyxl
        try:
            wb = load_workbook(file)
            ws = wb.active
        except Exception as e:
            return jsonify({'success': False, 'message': f'خطأ في قراءة ملف Excel: {str(e)}'}), 400
        
        # قراءة العناوين من الصف الأول
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                headers.append('')
        
        # التحقق من الأعمدة المطلوبة
        required_columns = ['اسم المنتج', 'الفئة', 'سعر الجملة', 'سعر البيع', 'الكمية']
        missing_columns = [col for col in required_columns if col not in headers]
        
        if missing_columns:
            return jsonify({
                'success': False, 
                'message': f'أعمدة مفقودة: {", ".join(missing_columns)}'
            }), 400
        
        # إنشاء فهرس للأعمدة
        column_index = {}
        for i, header in enumerate(headers):
            if header:
                column_index[header] = i
        
        # إحصائيات الاستيراد
        added_count = 0
        updated_count = 0
        skipped_count = 0
        errors = []
        
        # تعريف دالة مساعدة خارج الحلقة
        def get_cell_value(row, column_name, default=''):
            if column_name in column_index and column_index[column_name] < len(row):
                value = row[column_index[column_name]]
                if value is None:
                    return default
                value_str = str(value).strip()
                return value_str if value_str.lower() not in ['none', 'nan', ''] else default
            return default
        
        # إضافة logging للتشخيص
        app.logger.info(f"Starting import process. Total rows to process: {ws.max_row - 1}")
        
        # معالجة كل صف (بداية من الصف الثاني)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                # تجاهل الصفوف الفارغة
                if not any(row) or all(v is None or str(v).strip() == '' for v in row):
                    app.logger.debug(f"Skipping empty row {row_num}")
                    continue
                
                product_name = get_cell_value(row, 'اسم المنتج')
                category_name = get_cell_value(row, 'الفئة')
                
                app.logger.debug(f"Processing row {row_num}: product='{product_name}', category='{category_name}'")
                
                # التحقق من البيانات الأساسية
                if not product_name or product_name.lower() in ['none', 'nan', '']:
                    errors.append(f'الصف {row_num}: اسم المنتج فارغ أو غير صحيح')
                    skipped_count += 1
                    continue
                
                try:
                    wholesale_price_str = get_cell_value(row, 'سعر الجملة', '0')
                    retail_price_str = get_cell_value(row, 'سعر البيع', '0')
                    stock_quantity_str = get_cell_value(row, 'الكمية', '0')
                    
                    wholesale_price = float(wholesale_price_str)
                    retail_price = float(retail_price_str)
                    stock_quantity = float(stock_quantity_str)
                except (ValueError, TypeError) as e:
                    errors.append(f'الصف {row_num}: خطأ في تحويل الأرقام - {str(e)}')
                    skipped_count += 1
                    continue
                
                if wholesale_price <= 0 or retail_price <= 0:
                    errors.append(f'الصف {row_num}: أسعار غير صحيحة')
                    skipped_count += 1
                    continue
                
                if retail_price <= wholesale_price:
                    errors.append(f'الصف {row_num}: سعر البيع يجب أن يكون أكبر من سعر الجملة')
                    skipped_count += 1
                    continue
                
                # البحث عن الفئة
                category = Category.query.filter_by(name_ar=category_name).first()
                if not category:
                    errors.append(f'الصف {row_num}: الفئة "{category_name}" غير موجودة')
                    skipped_count += 1
                    continue
                
                # التحقق من وجود المنتج
                existing_product = Product.query.filter_by(name_ar=product_name).first()
                
                if existing_product and not update_existing:
                    skipped_count += 1
                    continue
                
                # البيانات الاختيارية
                description = get_cell_value(row, 'وصف المنتج')
                
                try:
                    min_stock = float(get_cell_value(row, 'الحد الأدنى للمخزون', '10'))
                except (ValueError, TypeError):
                    min_stock = 10
                
                unit_type = get_cell_value(row, 'نوع الوحدة', 'كامل')
                if unit_type not in ['كامل', 'جزئي']:
                    unit_type = 'كامل'
                
                unit_description = get_cell_value(row, 'وصف الوحدة')
                
                if existing_product and update_existing:
                    # تحديث المنتج الموجود
                    app.logger.info(f"Updating existing product: {product_name}")
                    existing_product.description_ar = description
                    existing_product.category_id = category.id
                    existing_product.wholesale_price = wholesale_price
                    existing_product.retail_price = retail_price
                    existing_product.price = retail_price
                    existing_product.stock_quantity = stock_quantity
                    existing_product.min_stock_threshold = min_stock
                    existing_product.unit_type = unit_type
                    existing_product.unit_description = unit_description
                    existing_product.updated_at = datetime.utcnow()
                    updated_count += 1
                else:
                    # إضافة منتج جديد
                    app.logger.info(f"Adding new product: {product_name}")
                    new_product = Product(
                        name_ar=product_name,
                        description_ar=description,
                        category_id=category.id,
                        wholesale_price=wholesale_price,
                        retail_price=retail_price,
                        price=retail_price,
                        stock_quantity=stock_quantity,
                        min_stock_threshold=min_stock,
                        unit_type=unit_type,
                        unit_description=unit_description
                    )
                    db.session.add(new_product)
                    added_count += 1
                    
                app.logger.info(f"Row {row_num} processed successfully")
                
            except Exception as e:
                app.logger.error(f'Error processing row {row_num}: {str(e)}')
                errors.append(f'الصف {row_num}: {str(e)}')
                skipped_count += 1
                continue
        
        # حفظ التغييرات
        try:
            app.logger.info(f"Import summary: added={added_count}, updated={updated_count}, skipped={skipped_count}, errors={len(errors)}")
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'تم استيراد المنتجات بنجاح',
                'added_count': added_count,
                'updated_count': updated_count,
                'skipped_count': skipped_count,
                'errors': errors[:10],  # أول 10 أخطاء فقط
                'total_rows_processed': added_count + updated_count + skipped_count
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({
                'success': False,
                'message': f'خطأ في حفظ البيانات: {str(e)}'
            }), 500
            
    except Exception as e:
        app.logger.error(f"Error importing Excel: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'خطأ في معالجة الملف: {str(e)}'
        }), 500

@app.route('/api/products/debug-excel', methods=['POST'])
@login_required
@admin_required
def api_debug_excel():
    """اختبار قراءة ملف Excel لتشخيص المشاكل"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'لم يتم العثور على ملف'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'لم يتم اختيار ملف'}), 400
        
        # قراءة ملف Excel
        wb = load_workbook(file)
        ws = wb.active
        
        # قراءة العناوين
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                headers.append('')
        
        # قراءة أول 5 صفوف للاختبار
        rows_data = []
        row_count = 0
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if row_count >= 5:  # اقرأ أول 5 صفوف فقط
                break
            
            row_data = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    value = row[i]
                    row_data[header] = str(value) if value is not None else 'None'
            
            rows_data.append({
                'row_number': row_num,
                'data': row_data,
                'raw_values': [str(v) if v is not None else 'None' for v in row[:len(headers)]]
            })
            row_count += 1
        
        return jsonify({
            'total_rows': ws.max_row,
            'headers': headers,
            'sample_rows': rows_data,
            'worksheet_name': ws.title
        })
        
    except Exception as e:
        return jsonify({'error': f'خطأ في قراءة الملف: {str(e)}'}), 500

@app.route('/qr-generator')
@login_required
def qr_generator():
    """صفحة توليد QR codes للينكات ومواقع التواصل الاجتماعي"""
    return render_template('qr_generator.html')

@app.route('/price-ticket')
@login_required
@seller_or_admin_required
def price_ticket():
    """صفحة إنتاج تيكت الأسعار"""
    products = Product.query.order_by(Product.name_ar).all()
    return render_template('price_ticket.html', products=products)

@app.route('/api/search-products')
@login_required
@seller_or_admin_required
def api_search_products():
    """البحث عن المنتجات لتيكت الأسعار"""
    query = request.args.get('q', '').strip()
    products = []
    
    if query:
        products = Product.query.filter(
            Product.name_ar.contains(query)
        ).order_by(Product.name_ar).limit(10).all()
    
    return jsonify([{
        'id': product.id,
        'name': product.name_ar,
        'retail_price': product.retail_price,
        'wholesale_price': product.wholesale_price
    } for product in products])

@app.route('/debug-auth')
def debug_auth():
    """صفحة تشخيص مشاكل المصادقة"""
    from flask_wtf.csrf import generate_csrf
    
    debug_info = {
        'is_authenticated': current_user.is_authenticated,
        'user_id': current_user.id if current_user.is_authenticated else None,
        'username': current_user.username if current_user.is_authenticated else None,
        'user_role': current_user.role if current_user.is_authenticated else None,
        'is_system_user': current_user.is_system if current_user.is_authenticated else None,
        'session_keys': list(session.keys()),
        'total_users': User.query.count(),
        'static_user_exists': User.query.filter_by(username='araby', is_system=True).first() is not None
    }
    
    # Get all users for debugging
    all_users = []
    try:
        for user in User.query.all():
            all_users.append({
                'id': user.id,
                'username': user.username,
                'role': user.role,
                'active': user.is_active,
                'system': user.is_system,
                'locked': user.is_account_locked(),
                'failed_attempts': user.failed_login_attempts,
                'has_password_hash': bool(user.password_hash)
            })
    except Exception as e:
        all_users = [{'error': str(e)}]
    
    csrf_token = generate_csrf()
    
    # Build user table rows
    user_rows = ""
    for u in all_users:
        active_class = "success" if u.get('active') else "error"
        locked_class = "error" if u.get('locked') else "success"
        hash_class = "success" if u.get('has_password_hash') else "error"
        
        user_rows += f"""<tr>
            <td>{u.get('id', 'N/A')}</td>
            <td>{u.get('username', 'N/A')}</td>
            <td>{u.get('role', 'N/A')}</td>
            <td class="{active_class}">{u.get('active', 'N/A')}</td>
            <td>{u.get('system', 'N/A')}</td>
            <td class="{locked_class}">{u.get('locked', 'N/A')}</td>
            <td>{u.get('failed_attempts', 'N/A')}</td>
            <td class="{hash_class}">{u.get('has_password_hash', 'N/A')}</td>
        </tr>"""
    
    return f"""
    <html dir="rtl">
    <head>
        <title>تشخيص المصادقة</title>
        <meta charset="utf-8">
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; }}
            .info {{ background: #f0f0f0; padding: 10px; margin: 10px 0; border-radius: 5px; }}
            .success {{ color: green; }}
            .error {{ color: red; }}
            .warning {{ color: orange; }}
            .btn {{ padding: 10px 15px; margin: 5px; border: none; border-radius: 5px; cursor: pointer; text-decoration: none; display: inline-block; }}
            .btn-green {{ background: green; color: white; }}
            .btn-blue {{ background: blue; color: white; }}
            table {{ border-collapse: collapse; width: 100%; }}
            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: center; }}
            th {{ background-color: #f2f2f2; }}
        </style>
    </head>
    <body>
        <h1>تشخيص نظام المصادقة</h1>
        
        <div class="info">
            <h3>معلومات المستخدم الحالي:</h3>
            <p><strong>مسجل الدخول:</strong> <span class="{'success' if debug_info['is_authenticated'] else 'error'}">{debug_info['is_authenticated']}</span></p>
            <p><strong>معرف المستخدم:</strong> {debug_info['user_id']}</p>
            <p><strong>اسم المستخدم:</strong> {debug_info['username']}</p>
            <p><strong>دور المستخدم:</strong> {debug_info['user_role']}</p>
            <p><strong>مستخدم نظام:</strong> {debug_info['is_system_user']}</p>
        </div>
        
        <div class="info">
            <h3>معلومات النظام:</h3>
            <p><strong>إجمالي المستخدمين:</strong> {debug_info['total_users']}</p>
            <p><strong>المستخدم الثابت موجود:</strong> <span class="{'success' if debug_info['static_user_exists'] else 'error'}">{debug_info['static_user_exists']}</span></p>
        </div>
        
        <div class="info">
            <h3>جميع المستخدمين:</h3>
            <table>
                <tr>
                    <th>المعرف</th>
                    <th>اسم المستخدم</th>
                    <th>الدور</th>
                    <th>نشط</th>
                    <th>نظام</th>
                    <th>مقفل</th>
                    <th>محاولات فاشلة</th>
                    <th>له كلمة مرور</th>
                </tr>
                {user_rows}
            </table>
        </div>
        
        <div class="info">
            <h3>الإجراءات:</h3>
            <a href="{url_for('index')}" class="btn btn-blue">صفحة تسجيل الدخول</a>
            <a href="{url_for('dashboard')}" class="btn btn-blue">لوحة التحكم</a>
        </div>
        
        <div class="info">
            <h3>اختبار المستخدم الثابت:</h3>
            <form method="POST" action="{url_for('index')}">
                <input type="hidden" name="csrf_token" value="{csrf_token}">
                <input type="hidden" name="username" value="araby">
                <input type="hidden" name="password" value="92321066">
                <button type="submit" class="btn btn-green">تسجيل دخول بالمستخدم الثابت</button>
            </form>
        </div>
        
        <div class="info">
            <h3>اختبار يدوي:</h3>
            <form method="POST" action="{url_for('index')}">
                <input type="hidden" name="csrf_token" value="{csrf_token}">
                <div style="margin: 10px 0;">
                    <label>اسم المستخدم:</label><br>
                    <input type="text" name="username" style="padding: 5px; width: 200px;" placeholder="ادخل اسم المستخدم">
                </div>
                <div style="margin: 10px 0;">
                    <label>كلمة المرور:</label><br>
                    <input type="password" name="password" style="padding: 5px; width: 200px;" placeholder="ادخل كلمة المرور">
                </div>
                <button type="submit" class="btn btn-green">تسجيل الدخول</button>
            </form>
        </div>
        
        <div class="info">
            <h3>أوامر التشخيص (للمطور):</h3>
            <p><code>python manage.py test-password</code> - اختبار كلمة مرور مستخدم</p>
            <p><code>python manage.py fix-password</code> - إصلاح كلمة مرور مستخدم</p>
            <p><code>python manage.py list-users</code> - عرض جميع المستخدمين</p>
        </div>
    </body>
    </html>
    """

@app.route('/api/sync', methods=['POST'])
@login_required
@seller_or_admin_required
def api_sync():
    """نقطة نهاية مخصصة لمزامنة البيانات غير المتصلة"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'البيانات مطلوبة'}), 400

        sync_type = data.get('type')
        sync_data = data.get('data', [])
        
        results = {
            'success': [],
            'errors': [],
            'total': len(sync_data)
        }

        if sync_type == 'sales':
            # مزامنة المبيعات المحفوظة محلياً
            for sale_data in sync_data:
                try:
                    # التحقق من البيانات المطلوبة
                    if not sale_data.get('items') or len(sale_data['items']) == 0:
                        results['errors'].append({
                            'data': sale_data,
                            'error': 'لا توجد عناصر في البيع'
                        })
                        continue

                    # إنشاء المبيعة
                    sale = Sale(
                        subtotal=sale_data.get('subtotal', 0),
                        total_amount=sale_data.get('total_amount', 0),
                        discount_type=sale_data.get('discount_type', 'none'),
                        discount_value=sale_data.get('discount_value', 0),
                        discount_amount=sale_data.get('discount_amount', 0),
                        user_id=current_user.id,
                        customer_id=sale_data.get('customer_id'),
                        payment_status=sale_data.get('payment_status', 'paid'),
                        payment_type=sale_data.get('payment_type', 'cash'),
                        notes=sale_data.get('notes', ''),
                        sale_date=datetime.utcnow()
                    )

                    db.session.add(sale)
                    db.session.flush()  # للحصول على sale.id

                    # إضافة عناصر البيع
                    for item_data in sale_data['items']:
                        product = Product.query.get(item_data['product_id'])
                        if not product:
                            raise ValueError(f"المنتج رقم {item_data['product_id']} غير موجود")

                        # التحقق من الكمية المتوفرة
                        if product.stock_quantity < item_data['quantity']:
                            raise ValueError(f"الكمية المطلوبة من {product.name_ar} غير متوفرة")

                        sale_item = SaleItem(
                            sale_id=sale.id,
                            product_id=item_data['product_id'],
                            quantity=item_data['quantity'],
                            unit_price=item_data['unit_price'],
                            total_price=item_data['total_price']
                        )

                        db.session.add(sale_item)

                        # تحديث المخزون
                        product.stock_quantity -= item_data['quantity']

                    db.session.commit()
                    
                    results['success'].append({
                        'local_id': sale_data.get('local_id'),
                        'server_id': sale.id,
                        'sale_date': sale.sale_date.isoformat()
                    })

                except Exception as e:
                    db.session.rollback()
                    results['errors'].append({
                        'data': sale_data,
                        'error': str(e)
                    })

        elif sync_type == 'customers':
            # مزامنة العملاء الجدد
            for customer_data in sync_data:
                try:
                    # التحقق من عدم وجود العميل مسبقاً
                    existing_customer = Customer.query.filter(
                        (Customer.name == customer_data['name']) |
                        (Customer.phone == customer_data.get('phone'))
                    ).first()

                    if existing_customer:
                        results['errors'].append({
                            'data': customer_data,
                            'error': 'العميل موجود مسبقاً'
                        })
                        continue

                    customer = Customer(
                        name=customer_data['name'],
                        phone=customer_data.get('phone'),
                        address=customer_data.get('address'),
                        notes=customer_data.get('notes')
                    )

                    db.session.add(customer)
                    db.session.commit()

                    results['success'].append({
                        'local_id': customer_data.get('local_id'),
                        'server_id': customer.id
                    })

                except Exception as e:
                    db.session.rollback()
                    results['errors'].append({
                        'data': customer_data,
                        'error': str(e)
                    })

        else:
            return jsonify({'error': 'نوع المزامنة غير مدعوم'}), 400

        return jsonify({
            'message': 'تمت المزامنة',
            'results': results,
            'success_count': len(results['success']),
            'error_count': len(results['errors'])
        })

    except Exception as e:
        app.logger.error(f"Sync error: {str(e)}")
        return jsonify({'error': 'حدث خطأ أثناء المزامنة'}), 500


@app.route('/api/offline-status')
@login_required
def api_offline_status():
    """إرجاع معلومات حالة التطبيق للوضع غير المتصل"""
    try:
        # إحصائيات المنتجات
        total_products = Product.query.count()
        low_stock_products = Product.query.filter(
            Product.stock_quantity <= Product.min_stock_threshold
        ).count()
        
        # إحصائيات العملاء
        total_customers = Customer.query.count()
        
        # إحصائيات المبيعات اليومية
        today = datetime.utcnow().date()
        today_sales = Sale.query.filter(
            func.date(Sale.sale_date) == today
        ).count()
        
        today_revenue = db.session.query(func.sum(Sale.total_amount)).filter(
            func.date(Sale.sale_date) == today
        ).scalar() or 0

        return jsonify({
            'status': 'online',
            'timestamp': datetime.utcnow().isoformat(),
            'stats': {
                'products': {
                    'total': total_products,
                    'low_stock': low_stock_products
                },
                'customers': {
                    'total': total_customers
                },
                'sales_today': {
                    'count': today_sales,
                    'revenue': float(today_revenue)
                }
            },
            'user': {
                'id': current_user.id,
                'username': current_user.username,
                'role': current_user.role
            }
        })

    except Exception as e:
        app.logger.error(f"Offline status error: {str(e)}")
        return jsonify({'error': 'حدث خطأ في الحصول على الحالة'}), 500


@app.route('/offline.html')
def offline_page():
    """صفحة الوضع غير المتصل"""
    return render_template('offline.html')


@app.route('/offline-demo')
@login_required
def offline_demo():
    """صفحة اختبار الوظائف غير المتصلة"""
    return render_template('offline-demo.html')


# إضافة route لدعم Service Worker
@app.route('/static/js/service-worker.js')
def service_worker():
    """تقديم Service Worker مع headers صحيحة"""
    response = make_response(send_from_directory('static/js', 'service-worker.js'))
    response.headers['Content-Type'] = 'application/javascript'
    response.headers['Service-Worker-Allowed'] = '/'
    return response

# Route إضافي للـ Service Worker في المسار الجذر
@app.route('/service-worker.js')
def service_worker_root():
    """تقديم Service Worker من المسار الجذر"""
    try:
        response = make_response(send_from_directory('static/js', 'service-worker.js'))
        response.headers['Content-Type'] = 'application/javascript; charset=utf-8'
        response.headers['Service-Worker-Allowed'] = '/'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return response
    except FileNotFoundError:
        # إنشاء service worker أساسي إذا لم يكن موجود
        basic_sw = """
// Basic Service Worker للطوارئ
const CACHE_NAME = 'norko-store-basic-v1';

self.addEventListener('install', event => {
    console.log('Basic Service Worker: Installing...');
    self.skipWaiting();
});

self.addEventListener('activate', event => {
    console.log('Basic Service Worker: Activating...');
    return self.clients.claim();
});

self.addEventListener('fetch', event => {
    // أساسي - لا يفعل شيء خاص
    return;
});
"""
        response = make_response(basic_sw)
        response.headers['Content-Type'] = 'application/javascript; charset=utf-8'
        response.headers['Service-Worker-Allowed'] = '/'
        return response

# Route للتحقق من Service Worker
@app.route('/sw-check')
def service_worker_check():
    """فحص حالة Service Worker"""
    import os
    sw_path = os.path.join('static', 'js', 'service-worker.js')
    exists = os.path.exists(sw_path)
    
    return jsonify({
        'service_worker_exists': exists,
        'path': sw_path,
        'routes_available': [
            '/service-worker.js',
            '/static/js/service-worker.js'
        ]
    })

# Route للتشخيص الشامل
@app.route('/offline-diagnostic')
@login_required
def offline_diagnostic():
    """تشخيص شامل للوظائف غير المتصلة"""
    import os
    
    # فحص الملفات
    js_files = {
        'service-worker.js': os.path.exists('static/js/service-worker.js'),
        'db-manager.js': os.path.exists('static/js/db-manager.js'),
        'sync-manager.js': os.path.exists('static/js/sync-manager.js'),
        'offline-handler.js': os.path.exists('static/js/offline-handler.js')
    }
    
    # فحص القوالب
    templates = {
        'offline.html': os.path.exists('templates/offline.html'),
        'offline-demo.html': os.path.exists('templates/offline-demo.html')
    }
    
    # معلومات الخادم
    server_info = {
        'host': request.host,
        'scheme': request.scheme,
        'user_agent': request.headers.get('User-Agent', ''),
        'secure': request.is_secure
    }
    
    return jsonify({
        'timestamp': datetime.utcnow().isoformat(),
        'files': js_files,
        'templates': templates,
        'server': server_info,
        'endpoints': {
            '/api/sync': True,
            '/api/offline-status': True,
            '/service-worker.js': True,
            '/static/js/service-worker.js': True
        }
    })

# Route بديل لملفات JavaScript مع MIME type صحيح
@app.route('/js/<filename>')
def serve_js(filename):
    """تقديم ملفات JavaScript مع MIME type صحيح"""
    try:
        response = make_response(send_from_directory('static/js', filename))
        response.headers['Content-Type'] = 'application/javascript; charset=utf-8'
        
        if 'service-worker' in filename:
            response.headers['Service-Worker-Allowed'] = '/'
            response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        
        return response
    except FileNotFoundError:
        return jsonify({'error': f'File {filename} not found'}), 404

# Route لـ PWA Manifest
@app.route('/manifest.json')
def serve_manifest():
    """تقديم PWA manifest مع headers صحيحة"""
    try:
        response = make_response(send_from_directory('static', 'manifest.json'))
        response.headers['Content-Type'] = 'application/manifest+json'
        return response
    except FileNotFoundError:
        return jsonify({'error': 'Manifest not found'}), 404


# Error Handlers
@app.errorhandler(404)
def not_found_error(error):
    """معالج خطأ 404 - الصفحة غير موجودة"""
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    """معالج خطأ 500 - خطأ داخلي في الخادم"""
    db.session.rollback()
    return render_template('500.html'), 500

@app.errorhandler(403)
def forbidden_error(error):
    """معالج خطأ 403 - ممنوع الوصول"""
    return render_template('403.html'), 403


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        # إنشاء المستخدم الثابت عند بدء التطبيق
        from models import create_static_user
        create_static_user()
    app.run(debug=True) 
